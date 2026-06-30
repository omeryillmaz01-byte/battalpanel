# -*- coding: utf-8 -*-
"""
TOPLU HESAP PLANI KURULUMU — HESAP PLANLARI klasöründeki tüm firmaları kurar.
Mevcut veritabanı firmalarının IBAN'larını korur (hesap planı kodlarıyla birleştirir).
"""
from __future__ import annotations
import re
from pathlib import Path
import openpyxl

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent))
from motor.firma_kur import firma_kur, FIRMA_DIZIN
from motor.eslestirici import _ilk_satir_basliklar

KAYNAK = Path(r"C:/Users/omery/OneDrive/Desktop/HESAP PLANLARI")

# dosya adı (küçük) → firma kodu (mevcut DB klasörüyle eşleşenler aynı kod, yeniler yeni)
DOSYA_KOD = {
    "2aa hesap planı": "2AA_YAYINCILIK",
    "akça ağız hesap planı": "AKCA_AGIZ",
    "besa gıda hesap planı": "BESA_GIDA",
    "bilpark bilişim hesap planı": "BILPARK_BILISIM",
    "bİlpark yazilim hesap plani": "BILPARK_YAZILIM",
    "bodur hesap planı": "BODUR_TEKNIK",
    "burak cet adi ort hesap planı": "BURAK_CET_ADI_ORT",
    "cet enerji hesap planı": "CET_ENERJI",
    "decor people hesap planı": "DECOR_PEOPLE",
    "dtb yazılım hesap planı": "DTB_YAZILIM",
    "erda gümrük hesap planı": "ERDA_GUMRUK",
    "erdem özşen hesap planı": "ERDEM_OZSEN",
    "erlamer hesap planı": "ERLAMER",
    "esni bilişim hesap planı": "ESNI_BILISIM",
    "gizem göker adi ort hesap planı": "GIZEM_GOKER_ADI_ORT",
    "hacıkerimoğlu gıda hesap planı": "HACIKERIMOGLU_GIDA",
    "hülya hatun hesap planı": "HULYA_HATUN_AKPINAR",
    "istanbul plastik hesap planı": "ISTANBUL_PLASTIK",
    "kariyerküre hesap planı": "KARIYERKURE",
    "karmakent hesap planı": "KARMAKENT",
    "kirpi hesap planı": "KIRPI_YAYINCILIK",
    "kna hesap planı": "KNA_BILISIM",
    "marmarages hesap planı": "MARMARAGES_GUNES",
    "md hesap planı": "MD_DERMATOLOJI",
    "nen hesap planı": "NEN_SIGORTA",
    "nes güzellik hesap planı": "NES_GUZELLIK",
    "ogutmen hesap planı": "OGUTMEN_SAGLIK",
    "osnak hesap planı": "OSNAK_NAKLIYAT",
    "sambaz sukusu hesap planı": "SAMBAZ_SUKUSU",
    "scf bilişim hesap planı": "SCF_BILISIM",
    "sİnan yilmaz hesap plani": "SINAN_YILMAZ",
    "ssc gÜvenlİk hesap plani": "SSC_GUVENLIK",
    "sts hesap plani": "STS_SAGLIK",
    "svi hesap plani": "SVI",
}


def _mevcut_ibanlari_oku(firma_kodu: str) -> list[tuple[str, str]]:
    """Mevcut banka_hesaplari'ndaki (banka, IBAN) çiftlerini döner (kod birleştirme öncesi)."""
    yol = FIRMA_DIZIN / firma_kodu / "01_banka_hesaplari.xlsx"
    if not yol.exists():
        return []
    wb = openpyxl.load_workbook(yol, data_only=True)
    ws = wb.active
    bas = _ilk_satir_basliklar(ws)
    out = []
    for row in ws.iter_rows(min_row=bas + 1, values_only=True):
        if not row:
            continue
        iban = str(row[4] or "").strip() if len(row) > 4 else ""
        if iban:
            out.append((str(row[1] or "").strip(), iban))
    return out


def _ibanlari_birlestir(firma_kodu: str, eski_ibanlar: list[tuple[str, str]]) -> int:
    """Yeni banka_hesaplari'na (102.xx kodlu) eski IBAN'ları hesap no eşleşmesiyle yazar."""
    if not eski_ibanlar:
        return 0
    yol = FIRMA_DIZIN / firma_kodu / "01_banka_hesaplari.xlsx"
    wb = openpyxl.load_workbook(yol)
    ws = wb.active
    bas = _ilk_satir_basliklar(ws)
    # iban tail digits hazırla
    iban_map = []
    for banka, iban in eski_ibanlar:
        digits = re.sub(r"\D", "", iban)
        iban_map.append((digits, iban))
    sayac = 0
    for r in range(bas + 1, ws.max_row + 1):
        kod = ws.cell(row=r, column=1).value
        if not kod or not str(kod).startswith("102"):
            continue
        hesap_no = ws.cell(row=r, column=4).value
        no = re.sub(r"\D", "", str(hesap_no or ""))
        if not no:
            continue
        for digits, iban in iban_map:
            if no and (no in digits or digits.endswith(no.zfill(len(no)))):
                if not ws.cell(row=r, column=5).value:
                    ws.cell(row=r, column=5, value=iban)
                    sayac += 1
                break
    if sayac:
        wb.save(yol)
    return sayac


def main():
    dosyalar = sorted(KAYNAK.glob("*.xlsx"))
    print(f"{len(dosyalar)} hesap planı bulundu.\n")
    ok = 0
    bilinmeyen = []
    for yol in dosyalar:
        ad = yol.stem.lower()
        kod = DOSYA_KOD.get(ad)
        if not kod:
            bilinmeyen.append(yol.name)
            continue
        eski = _mevcut_ibanlari_oku(kod)
        try:
            s = firma_kur(kod, yol)
            n_iban = _ibanlari_birlestir(kod, eski)
            print(f"  [OK] {kod:22s} | banka:{s['banka']:3d} pos:{s['pos']:2d} cari:{s['cari']:4d} "
                  f"kural:{s['kural']:2d} POS→{s['pos_hesabi']:8s} | IBAN birleşti:{n_iban}")
            ok += 1
        except Exception as e:
            print(f"  [HATA] {kod}: {e}")
    print(f"\n{ok}/{len(dosyalar)} firma kuruldu.")
    if bilinmeyen:
        print("Eşlenemeyen dosyalar:", bilinmeyen)


if __name__ == "__main__":
    main()
