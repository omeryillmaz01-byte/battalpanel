# -*- coding: utf-8 -*-
"""
Eşleştirme motoru — Banka hareketini hesap koduna dönüştürür.

Akış:
  1. Firma referans Excel'leri (banka, POS, cari, kural) yüklenir.
  2. Her hareket için sıralı kural uygulanır:
     (a) Kural sözlüğünde anahtar kelime eşleşir mi? (komisyon, BSMV, masraf vs. → 780.02)
     (b) Cari adı eşleşir mi? (120/320)
     (c) Hiçbir kural eşleşmedi → MANUEL bayrağı + en yakın 3 öneri
  3. Güven skoru: 100 (exact keyword) | 80 (cari tam) | 60 (cari kısmi) | 0 (eşleşmedi)
  4. Bakiye kontrolü: açılış + Σ(tutar) = kapanış mı?
"""
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

import openpyxl

BASE = Path(__file__).resolve().parent.parent
FIRMA_DIZIN = BASE / "firmalar"


# ============================================================================= #
# Yardımcılar
# ============================================================================= #
def normalize(s: str) -> str:
    """Türkçe karakter, boşluk, noktalama temizliği — eşleştirme için."""
    if s is None:
        return ""
    s = str(s).upper()
    degisim = str.maketrans({"İ": "I", "I": "I", "Ş": "S", "Ğ": "G", "Ü": "U", "Ö": "O", "Ç": "C"})
    s = s.translate(degisim)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _ilk_satir_basliklar(ws) -> int:
    """Açıklama satırlarını atla, başlık satırının indexini döner.
    Başlık satırı: ilk hücresi TAM olarak bilinen bir başlık olan satır."""
    HEADER_TOKENLARI = {"HESAP KODU", "SIRA", "SΙRA", "SIRA"}
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if not row or row[0] is None:
            continue
        ilk = str(row[0]).strip().upper()
        if ilk in HEADER_TOKENLARI:
            return i
    return 1


# Cari adı eşleştirmede göz ardı edilecek SADECE jenerik kelimeler.
# DİKKAT: Sektör/ürün kelimeleri (MEŞRUBAT, OTO, GIDA, TÜTÜN...) BİLEREK listede DEĞİL —
# bunlar firma adının ayırt edici parçası; çıkarılırsa "Nokta Meşrubat" eşleşemez.
DURDURMA_KELIMELERI = {
    # Müşterinin kendi ünvanı (her açıklamada geçer)
    "ISIK", "PETROL", "KOMUR", "LASTIK", "PARCA", "ELKTRONIK", "ELEKTRONIK",
    "MADEN", "NAK", "DAYANIKLI",
    # Şirket türü / hukuki ekler
    "LTD", "STI", "STII", "SAN", "SANAYI", "TIC", "TICARET", "TICARETI",
    "VE", "ILE", "ANONIM", "AS", "SIRKETI", "SIRKET", "LIMITED", "KOLL",
    # Banka/işlem gürültüsü
    "MBL", "HAV", "FAST", "ODEME", "ODE", "ODEMESI", "PES", "VIRMAN", "TR",
    "ABONE", "FATURA", "TL", "KS", "PARA", "EFT", "BSMV", "HESABI", "HESAP",
    "NOLU", "NO", "ADET",
}


def cari_adi_cikar(aciklama: str) -> str:
    """
    Akbank açıklama kalıbından cari adını çıkarır.
      '7777/MBL-3530082-Osmanlı Tedarik Dağıtım-Işık petrol nestle' → 'Osmanlı Tedarik Dağıtım'
      '7777/MBL-HAV.MODÜL KİMYA PETR...-...'                        → 'MODÜL KİMYA PETR...'
      '7777/MBL-HAV.REMZİ ÇAKAR GIDA...-128437...'                  → 'REMZİ ÇAKAR GIDA...'
    Kalıba uymuyorsa açıklamanın tümünü döner.
    """
    s = aciklama.strip()
    # 'HAV.' kalıbı
    m = re.search(r"MBL-HAV\.(.+?)(?:-[0-9]|$)", s, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    # 'MBL-<rakam>-<isim>-<not>' kalıbı
    m = re.search(r"MBL-\d+-(.+?)(?:-|$)", s)
    if m:
        return m.group(1).strip()
    # 'MBL-<isim>' (rakamsız)
    m = re.search(r"MBL-(.+?)(?:-|$)", s)
    if m:
        return m.group(1).strip()
    return s


def _anlamli_tokenlar(metin: str) -> set[str]:
    return {t for t in normalize(metin).split()
            if len(t) >= 3 and t not in DURDURMA_KELIMELERI}


# Banka adı (normalize, ön ek) → o bankanın 103 VERİLEN ÇEKLER hesabı (çek ÖDEMESİ, -)
CEK_HESAP_HARITASI = {
    "AKBANK":    "103.01.003.0001",
    "GARANTI":   "103.01.002.0001",
    "HALKBANK":  "103.01.011.0001",
    "YAPIKREDI": "103.01.006.0001",
    "ZIRAAT":    "103.01.009.0001",
    "IS BANK":   "103.01.001.0001",
}
# Banka adı → o bankanın 101 ALINAN ÇEKLER hesabı (çek TAHSİLATI, +)
CEK_TAHSIL_HARITASI = {
    "AKBANK":    "101.01.001.0003",
    "GARANTI":   "101.01.001.0002",
    "HALKBANK":  "101.01.001.0011",
    "YAPIKREDI": "101.01.001.0006",
    "ZIRAAT":    "101.01.001.0009",
    "IS BANK":   "101.01.001.0001",
}


# ============================================================================= #
# Referans veri sınıfları
# ============================================================================= #
@dataclass
class BankaHesabi:
    hesap_kodu: str
    banka: str
    hesap_adi: str
    hesap_no: str
    iban: str
    anahtarlar: list[str] = field(default_factory=list)


@dataclass
class PosHesabi:
    hesap_kodu: str
    pos_adi: str
    banka: str
    anahtarlar: list[str] = field(default_factory=list)


@dataclass
class Cari:
    hesap_kodu: str
    tip: str  # "120 ALICI" / "320 SATICI"
    ad: str
    anahtarlar: list[str] = field(default_factory=list)
    aktif: bool = True


@dataclass
class Kural:
    sira: int
    anahtarlar: list[str]
    hesap_kodu: str
    hesap_adi: str
    tutar_isareti: str  # "NEGATIF" | "POZITIF" | "HER İKİSİ"
    guven: int = 100
    not_: str = ""
    aktif: bool = True


@dataclass
class Ortak:
    ad: str
    kod_131: str          # ortaktan alacak (para çıkışı)
    kod_331: str          # ortağa borç (para girişi)
    anahtarlar: list[str] = field(default_factory=list)


@dataclass
class HesapPlaniSatir:
    kod: str
    ad: str
    bakiye: float | None = None


@dataclass
class FirmaTablolari:
    bankalar: list[BankaHesabi]
    posler: list[PosHesabi]
    cariler: list[Cari]
    kurallar: list[Kural]
    ortaklar: list[Ortak] = field(default_factory=list)
    hesap_plani: list[HesapPlaniSatir] = field(default_factory=list)


def firma_yukle(firma_kodu: str) -> FirmaTablolari:
    d = FIRMA_DIZIN / firma_kodu
    return FirmaTablolari(
        bankalar=_bankalari_yukle(d / "01_banka_hesaplari.xlsx"),
        posler=_posleri_yukle(d / "02_pos_hesaplari.xlsx"),
        cariler=_carileri_yukle(d / "03_cari_eslesme.xlsx"),
        kurallar=_kurallari_yukle(d / "04_kural_sozlugu.xlsx"),
        ortaklar=_ortaklari_yukle(d / "05_ortaklar.xlsx"),
        hesap_plani=_hesap_plani_yukle(d / "06_hesap_plani.xlsx"),
    )


def _hesap_plani_yukle(yol: Path) -> list[HesapPlaniSatir]:
    """06_hesap_plani.xlsx — TAM hesap planı (varsa). Yoksa boş liste."""
    if not yol.exists():
        return []
    import openpyxl
    out: list[HesapPlaniSatir] = []
    try:
        wb = openpyxl.load_workbook(yol, data_only=True, read_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row or not row[0]:
                continue
            kod = str(row[0]).strip()
            ad = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            bakiye = None
            if len(row) > 2 and row[2] is not None:
                try:
                    bakiye = float(str(row[2]).replace(".", "").replace(",", ".")
                                   if isinstance(row[2], str) else row[2])
                except (ValueError, TypeError):
                    bakiye = None
            if kod and kod.upper() != "HESAP KODU":
                out.append(HesapPlaniSatir(kod, ad, bakiye))
        wb.close()
    except Exception:
        return []
    return out


def _ortaklari_yukle(yol: Path) -> list[Ortak]:
    if not yol.exists():
        return []
    wb = openpyxl.load_workbook(yol, data_only=True)
    ws = wb.active
    bas = _ilk_satir_basliklar(ws)
    # başlık 'Ortak Adı' ile başlıyor; _ilk_satir_basliklar HESAP KODU/SIRA arar → bulamaz, manuel bul
    if bas == 1:
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row and row[0]:
                hucre = str(row[0]).strip().upper()
                if hucre in ("ORTAK ADI", "ORTAK AD", "ORTAK ADI / SOYADI"):
                    bas = i
                    break
    out = []
    for row in ws.iter_rows(min_row=bas + 1, values_only=True):
        if not row or not row[0]:
            continue
        out.append(Ortak(
            ad=str(row[0]).strip(),
            kod_131=str(row[1] or "").strip(),
            kod_331=str(row[2] or "").strip(),
            anahtarlar=_split_keys(row[3] if len(row) > 3 else None),
        ))
    return out


def _split_keys(s) -> list[str]:
    if not s:
        return []
    return [normalize(p) for p in str(s).split("|") if p.strip()]


def _bankalari_yukle(yol: Path) -> list[BankaHesabi]:
    wb = openpyxl.load_workbook(yol, data_only=True)
    ws = wb.active
    bas = _ilk_satir_basliklar(ws)
    out = []
    for row in ws.iter_rows(min_row=bas + 1, values_only=True):
        if not row:
            continue
        hesap_kodu = str(row[0]).strip() if row[0] else ""
        iban = str(row[4] or "").strip() if len(row) > 4 else ""
        # Hesap kodu boş ama IBAN'ı olan satırlar da gelsin (veritabanı firmaları, hesap planı gelmemiş)
        if not hesap_kodu and not iban:
            continue
        out.append(BankaHesabi(
            hesap_kodu=hesap_kodu,
            banka=str(row[1] or "").strip(),
            hesap_adi=str(row[2] or "").strip(),
            hesap_no=str(row[3] or "").strip(),
            iban=iban,
            anahtarlar=_split_keys(row[5] if len(row) > 5 else None),
        ))
    return out


def _posleri_yukle(yol: Path) -> list[PosHesabi]:
    wb = openpyxl.load_workbook(yol, data_only=True)
    ws = wb.active
    bas = _ilk_satir_basliklar(ws)
    out = []
    for row in ws.iter_rows(min_row=bas + 1, values_only=True):
        if not row or not row[0]:
            continue
        out.append(PosHesabi(
            hesap_kodu=str(row[0]).strip(),
            pos_adi=str(row[1] or "").strip(),
            banka=str(row[2] or "").strip(),
            anahtarlar=_split_keys(row[3]),
        ))
    return out


def _carileri_yukle(yol: Path) -> list[Cari]:
    wb = openpyxl.load_workbook(yol, data_only=True)
    ws = wb.active
    bas = _ilk_satir_basliklar(ws)
    out = []
    for row in ws.iter_rows(min_row=bas + 1, values_only=True):
        if not row or not row[0]:
            continue
        aktif = str(row[4] or "E").strip().upper() != "H"
        out.append(Cari(
            hesap_kodu=str(row[0]).strip(),
            tip=str(row[1] or "").strip(),
            ad=str(row[2] or "").strip(),
            anahtarlar=_split_keys(row[3]),
            aktif=aktif,
        ))
    return out


def _kurallari_yukle(yol: Path) -> list[Kural]:
    wb = openpyxl.load_workbook(yol, data_only=True)
    ws = wb.active
    bas = _ilk_satir_basliklar(ws)
    # Başlık satırından kolon sırasını tespit et (eski/yeni şema uyumu)
    basliklar = [str(c or "").strip().upper() for c in next(ws.iter_rows(min_row=bas, max_row=bas, values_only=True))]
    def kol(adlar, vars):
        for a in adlar:
            for i, b in enumerate(basliklar):
                if a in b:
                    return i
        return vars
    i_sira = kol(["SIRA"], 0)
    i_kelime = kol(["ANAHTAR"], 1)
    i_kod = kol(["HESAP KODU"], 2)
    i_ad = kol(["HESAP AD", "HESAP İSM"], 3)
    i_isaret = kol(["İŞARET", "ISARET"], 4)
    i_guven = kol(["GÜVEN", "GUVEN"], -1)
    i_not = kol(["NOT", "AÇIKLAMA"], -1)
    i_aktif = kol(["AKTİF", "AKTIF"], 5)

    out = []
    for row in ws.iter_rows(min_row=bas + 1, values_only=True):
        if not row or i_sira >= len(row) or row[i_sira] is None:
            continue
        try:
            sira = int(row[i_sira])
        except (TypeError, ValueError):
            continue
        def g(i, vars=""):
            return row[i] if 0 <= i < len(row) and row[i] is not None else vars
        try:
            guven = int(g(i_guven, 100)) if i_guven >= 0 else 100
        except (TypeError, ValueError):
            guven = 100
        aktif = str(g(i_aktif, "E")).strip().upper() != "H"
        out.append(Kural(
            sira=sira,
            anahtarlar=_split_keys(g(i_kelime)),
            hesap_kodu=str(g(i_kod)).strip(),
            hesap_adi=str(g(i_ad)).strip(),
            tutar_isareti=str(g(i_isaret, "HER İKİSİ")).strip().upper(),
            guven=guven,
            not_=str(g(i_not)).strip(),
            aktif=aktif,
        ))
    out.sort(key=lambda k: k.sira)
    return out


# ============================================================================= #
# Hareket & Eşleşme
# ============================================================================= #
@dataclass
class Hareket:
    tarih: str       # 'gg/aa/yyyy' veya datetime → __str__
    aciklama: str
    tutar: float
    referans: str = ""

    @property
    def negatif(self) -> bool:
        return self.tutar < 0


@dataclass
class Eslesme:
    hareket: Hareket
    hesap_kodu: str = ""
    hesap_adi: str = ""
    kaynak: str = "MANUEL"     # KURAL | CARI | POS | MANUEL
    guven: int = 0             # 0-100
    not_: str = ""
    oneriler: list[tuple[str, str]] = field(default_factory=list)


# ============================================================================= #
# Hesap planı arama yardımcıları
# ============================================================================= #
def _hesap_bul(tablolar: "FirmaTablolari", prefix: str, anahtar: str = "") -> str:
    """Kural sözlüğü, cari ve banka hesapları arasında prefix ile başlayan
    ve anahtar kelimeyi içeren hesap kodunu döner. Bulamazsa ''."""
    n_anahtar = normalize(anahtar)
    # 1) TAM hesap planında ara (06_hesap_plani.xlsx) — yaprak hesabı tercih et.
    # Kural sözlüğü sonradan gelir; aksi halde eski şablonlardaki 780.02/335.02
    # gibi ezber kodlar gerçek hesap planının önüne geçer.
    eslesen = [h for h in tablolar.hesap_plani if h.kod.startswith(prefix)
               and (not n_anahtar or n_anahtar in normalize(h.ad))]
    if eslesen:
        return _yaprak_sec(eslesen, tablolar.hesap_plani)
    # 2) Cari eşleşme tablosunda ara
    for c in tablolar.cariler:
        if c.hesap_kodu.startswith(prefix):
            if not n_anahtar or n_anahtar in normalize(c.ad):
                return c.hesap_kodu
    # 3) Banka hesaplarında ara
    for b in tablolar.bankalar:
        if b.hesap_kodu.startswith(prefix):
            if not n_anahtar or n_anahtar in normalize(b.hesap_adi):
                return b.hesap_kodu
    # 4) En son kural sözlüğüne bak. Bu tablo kullanıcı tarafından elle
    # zenginleştirilebilir ama başlangıç şablonundaki sabit kodlar bağlayıcı değildir.
    for k in tablolar.kurallar:
        if k.hesap_kodu.startswith(prefix):
            if not n_anahtar or n_anahtar in normalize(k.hesap_adi):
                return k.hesap_kodu
    return ""


def _yaprak_sec(adaylar: list, tum_plan: list) -> str:
    """Aday hesaplar arasından YAPRAK (alt kırılımı olmayan) hesabı seç.
    Örn 780 ile 780.02 varsa 780.02'yi (yaprak) döner, parent 780'i değil."""
    kodlar = [h.kod for h in adaylar]
    yapraklar = [k for k in kodlar if not any(o != k and o.startswith(k + ".") for o in (h.kod for h in tum_plan))]
    if yapraklar:
        # En çok kademeli (en spesifik) yaprağı tercih et
        return max(yapraklar, key=lambda k: k.count("."))
    return kodlar[0]


def _hesap_bul_ilk(tablolar: "FirmaTablolari", prefix: str) -> str:
    """Prefix ile başlayan İLK hesap kodunu döner (alt hesap farketmez)."""
    # TAM hesap planı her zaman en güvenilir kaynak.
    eslesen = [h for h in tablolar.hesap_plani if h.kod.startswith(prefix)]
    if eslesen:
        return _yaprak_sec(eslesen, tablolar.hesap_plani)
    for c in tablolar.cariler:
        if c.hesap_kodu.startswith(prefix):
            return c.hesap_kodu
    for b in tablolar.bankalar:
        if b.hesap_kodu.startswith(prefix):
            return b.hesap_kodu
    for k in tablolar.kurallar:
        if k.hesap_kodu.startswith(prefix):
            return k.hesap_kodu
    return ""


def _ortak_131_veya_331(tablolar: "FirmaTablolari") -> str:
    """Firmada 131 mi 331 mi kullanılacak? ÖNCE HESAP PLANINDAKİ BAKİYEYE bakar:
    131.xx alt hesaplarının mı yoksa 331.xx'in mi bakiyesi var → o kullanılır."""
    # 1) Hesap planı (06) bakiyesi — en güvenilir
    bak_131 = sum(abs(h.bakiye) for h in tablolar.hesap_plani
                  if h.kod.startswith("131.") and h.bakiye)
    bak_331 = sum(abs(h.bakiye) for h in tablolar.hesap_plani
                  if h.kod.startswith("331.") and h.bakiye)
    if bak_131 or bak_331:
        return "131" if bak_131 >= bak_331 else "331"
    # 2) Ortaklar tablosu
    for o in tablolar.ortaklar:
        if o.kod_131:
            return "131"
        if o.kod_331:
            return "331"
    # 3) Cari tablosunda 131/331 ara
    has_131 = any(c.hesap_kodu.startswith("131") for c in tablolar.cariler)
    has_331 = any(c.hesap_kodu.startswith("331") for c in tablolar.cariler)
    if has_131 and not has_331:
        return "131"
    if has_331 and not has_131:
        return "331"
    return "131"  # varsayılan


# Şahıs firmaları — şirket kredi kartı YOK, KK ödemesi = 131/331
SAHIS_FIRMALARI = {
    "ERDEM_OZSEN", "HULYA_HATUN_AKPINAR", "GIZEM_GIDA", "SINAN_CAKMAK",
    "SAMBAZ_OSGB", "BURAK_PARCA",
}

# Akıllı kısaltma tanıma — banka açıklamalarındaki kısaltmalar → cari adı
KISALTMA_HARITASI = {
    "MULTNT": "MULTINET",
    "TRKCLL": "TURKCELL",
    "TRCELL": "TURKCELL",
    "TTLKOM": "TURK TELEKOM",
    "TTMOBIL": "TURK TELEKOM",
    "TTNET": "TTNET",  # TTNET ayrı firma (Türk Telekom DEĞİL) → kendi 320 hesabı
    "BEDASE": "BEDAS",
    "SUPRONL": "SUPERONLINE",
    "SUPERONLINE": "SUPERONLINE",
    "VODAFNE": "VODAFONE",
    "VODAFON": "VODAFONE",
    "TRKDJT": "TURKCELL DİJİTAL",
    "IGDASE": "IGDAS",
    "ISTGAZD": "IGDAS",
}


# ============================================================================= #
# Eşleştirici
# ============================================================================= #
class Eslestirici:
    def __init__(self, firma_kodu: str, banka_hesap_kodu: str | None = None):
        """banka_hesap_kodu: işlenmekte olan banka (102.xx) — POS hakedişi ve virman için kullanılır."""
        self.firma_kodu = firma_kodu
        self.tablolar = firma_yukle(firma_kodu)
        self.banka_hesap_kodu = banka_hesap_kodu
        self.sahis_firma = firma_kodu in SAHIS_FIRMALARI

        # Firma ünvanı (virman tespitinde kendi adını tanımak için)
        import json
        _GENEL_KELIMELER = {"SAGLIK", "TEKNIK", "SERVIS", "GIDA", "INSAAT", "BILISIM",
                            "YAZILIM", "ENERJI", "TURIZM", "PAZARLAMA", "DANISMANLIK",
                            "EGITIM", "NAKLIYAT", "OTOMOTIV", "MATBAA", "ELEKTRIK",
                            "GUVENLIK", "HIZ", "PAZ", "MED", "DIS", "TIP", "MEDIKAL"}
        _fb_yol = Path(__file__).resolve().parent.parent / "firmalar" / firma_kodu / "firma_bilgi.json"
        self.firma_unvan_marka: str = ""
        self.firma_unvan_tokenlar: set[str] = set()
        if _fb_yol.exists():
            try:
                _fb = json.loads(_fb_yol.read_text(encoding="utf-8"))
                _unvan = normalize(_fb.get("unvan", ""))
                _tum = [t for t in _unvan.split() if len(t) >= 3 and t not in DURDURMA_KELIMELERI]
                if _tum:
                    self.firma_unvan_marka = _tum[0]
                self.firma_unvan_tokenlar = {t for t in _tum if t not in _GENEL_KELIMELER}
            except Exception:
                pass

        # Ortaktan alacak mı borç mu? (131 vs 331 tek hesap)
        self.ortak_hesap = _ortak_131_veya_331(self.tablolar)

        # İşlenmekte olan bankanın adı
        self.islem_banka_adi = ""

        # POS hakedişi için bu bankaya ait POS kodu (108.xx) bulunsun
        self.bu_banka_pos_kodu = ""
        self.bu_banka_cek_kodu = ""
        self.bu_banka_cek_tahsil_kodu = ""
        if banka_hesap_kodu:
            banka_obj = next((b for b in self.tablolar.bankalar if b.hesap_kodu == banka_hesap_kodu), None)
            if banka_obj:
                self.islem_banka_adi = banka_obj.banka
                banka_anahtari = normalize(banka_obj.banka)
                pos = next((p for p in self.tablolar.posler
                            if normalize(p.banka).startswith(banka_anahtari[:6])), None)
                if pos:
                    self.bu_banka_pos_kodu = pos.hesap_kodu
                for ad_anahtar, cek_kodu in CEK_HESAP_HARITASI.items():
                    if banka_anahtari.startswith(ad_anahtar):
                        self.bu_banka_cek_kodu = cek_kodu
                        break
                self.bu_banka_cek_tahsil_kodu = ""
                for ad_anahtar, cek_kodu in CEK_TAHSIL_HARITASI.items():
                    if banka_anahtari.startswith(ad_anahtar):
                        self.bu_banka_cek_tahsil_kodu = cek_kodu
                        break

        # Cari token nadirlik haritası
        from collections import Counter as _Counter
        self.token_nadirlik: _Counter = _Counter()
        for c in self.tablolar.cariler:
            for tok in _anlamli_tokenlar(c.ad):
                self.token_nadirlik[tok] += 1

        # Hesap no → 102.xx ve IBAN → 102.xx sözlükleri (VİRMAN tespitinde kullanılır)
        kendi = self._kendi_son_parcasi(banka_hesap_kodu)
        self.hesap_no_haritasi: dict[str, tuple[str, str]] = {}
        self.iban_haritasi: dict[str, tuple[str, str]] = {}
        for b in self.tablolar.bankalar:
            son_parca = b.hesap_no.split("-")[-1].lstrip("0") if b.hesap_no else ""
            if son_parca and son_parca != kendi:
                self.hesap_no_haritasi[son_parca] = (b.hesap_kodu, b.banka)
            iban = (getattr(b, "iban", "") or "").replace(" ", "").upper()
            if iban and b.hesap_kodu != banka_hesap_kodu:
                self.iban_haritasi[iban] = (b.hesap_kodu, b.banka)

    def _kendi_son_parcasi(self, kod: str | None) -> str:
        if not kod:
            return ""
        b = next((x for x in self.tablolar.bankalar if x.hesap_kodu == kod), None)
        if not b or not b.hesap_no:
            return ""
        return b.hesap_no.split("-")[-1].lstrip("0")

    def _virman_tespit(self, aciklama: str) -> tuple[str, str] | None:
        """Açıklamada kendi başka hesabımıza ait IBAN ya da hesap no varsa o 102.xx döner."""
        bosluksuz = re.sub(r"\s+", "", aciklama).upper()
        for iban, hedef in self.iban_haritasi.items():
            if iban and iban in bosluksuz:
                return hedef
        for no in re.findall(r"\d{6,}", aciklama):
            no_temiz = no.lstrip("0")
            if no_temiz in self.hesap_no_haritasi:
                return self.hesap_no_haritasi[no_temiz]
        n_acik = normalize(aciklama)
        for no, hedef in self.hesap_no_haritasi.items():
            if len(no) >= 5 and no in n_acik:
                return hedef
        # Kendi firma ünvanı geçiyor → kendi hesabına virman (EFT/INT vb.)
        # Marka tokeni (ilk kelime) MUTLAKA eşleşmeli + en az 2 FARKLI ayırt edici token
        if self.firma_unvan_marka and self.firma_unvan_marka in n_acik:
            ek_tokenlar = self.firma_unvan_tokenlar - {self.firma_unvan_marka}
            if ek_tokenlar:
                h_tok = set(n_acik.split())
                eslesen = sum(1 for t in ek_tokenlar if t in h_tok or t in n_acik)
                if eslesen >= min(2, len(ek_tokenlar)):
                    return ("VIRMAN_KENDI", "KENDİ FİRMA (HESAP ARASI)")
        return None

    def _kk_kart_no_esle(self, aciklama: str) -> tuple[str, str] | None:
        """EKSTRE BORÇ TAHSİLATI açıklamasından kart ön ekini çıkar, 329'da ara."""
        # Kart no açıklamanın ortasında da olabilir (başta tarih vs.) → önce 4hane+-+uzunno ara
        m = re.search(r"(\d{4})-\d{6,}", aciklama) or re.match(r"(\d{4})[-\s]", aciklama.strip())
        if not m:
            return None
        kart_on_ek = m.group(1)
        # 1) TAM hesap planında (06) 329 hesabı — adında kart 4 hanesi geçen (EN GÜVENİLİR)
        for h in self.tablolar.hesap_plani:
            if h.kod.startswith("329") and kart_on_ek in h.ad:
                return (h.kod, h.ad)
        for c in self.tablolar.cariler:
            if c.hesap_kodu.startswith("329"):
                if kart_on_ek in c.ad or kart_on_ek in c.hesap_kodu:
                    return (c.hesap_kodu, c.ad)
        for k in self.tablolar.kurallar:
            if k.hesap_kodu.startswith("329"):
                if kart_on_ek in k.hesap_adi or kart_on_ek in k.hesap_kodu:
                    return (k.hesap_kodu, k.hesap_adi)
        return None

    def _mevduat_hesap_bul(self, aciklama: str) -> str:
        """MEVDUAT GERİ DÖNÜŞÜ — hesap no'dan vadeli hesap (102.xx) bul."""
        for no in re.findall(r"\d{5,}", aciklama):
            no_temiz = no.lstrip("0")
            for b in self.tablolar.bankalar:
                n_ad = normalize(b.hesap_adi)
                if ("VADELI" in n_ad or "VADE" in n_ad) and no_temiz and no_temiz in b.hesap_no:
                    return b.hesap_kodu
        # Vadeli hesap adında "VADELİ" geçeni bul
        for b in self.tablolar.bankalar:
            n_ad = normalize(b.hesap_adi)
            if "VADELI" in n_ad and b.hesap_kodu.startswith("102"):
                return b.hesap_kodu
        return ""

    def _fon_hesap_bul(self) -> str:
        """FON alım/satım → 102.xx altında fon hesabı bul."""
        for b in self.tablolar.bankalar:
            n_ad = normalize(b.hesap_adi)
            if "FON" in n_ad and b.hesap_kodu.startswith("102"):
                return b.hesap_kodu
        return _hesap_bul(self.tablolar, "118", "FON") or ""

    def _faiz_geliri_hesap_bul(self) -> str:
        """642 faiz geliri — bankaya göre alt hesap bul."""
        banka = normalize(self.islem_banka_adi)
        kod = _hesap_bul(self.tablolar, "642", self.islem_banka_adi)
        if kod:
            return kod
        return _hesap_bul_ilk(self.tablolar, "642") or "642"

    def _stopaj_hesap_bul(self) -> str:
        """193 stopaj — bankaya göre alt hesap bul."""
        kod = _hesap_bul(self.tablolar, "193", self.islem_banka_adi)
        if kod:
            return kod
        return _hesap_bul_ilk(self.tablolar, "193") or "193"

    def _masraf_hesap_bul(self) -> str:
        """780 masraf hesabı — firmanın hesap planından oku."""
        eslesen = [h for h in self.tablolar.hesap_plani if h.kod.startswith("780")]
        if eslesen:
            return _yaprak_sec(eslesen, self.tablolar.hesap_plani)
        for c in self.tablolar.cariler:
            if c.hesap_kodu.startswith("780"):
                return c.hesap_kodu
        # 06_hesap_plani yoksa eski kural sözlüğündeki 780.02'yi ezber kabul etme.
        return "780"

    def _personel_hesap_bul(self, ad: str = "") -> str:
        """335 personel hesabı — firmanın hesap planından oku."""
        if ad:
            kod = _hesap_bul(self.tablolar, "335", ad)
            if kod:
                return kod
        eslesen = [h for h in self.tablolar.hesap_plani if h.kod.startswith("335")]
        if eslesen:
            return _yaprak_sec(eslesen, self.tablolar.hesap_plani)
        for c in self.tablolar.cariler:
            if c.hesap_kodu.startswith("335"):
                return c.hesap_kodu
        # 06_hesap_plani yoksa eski kural sözlüğündeki 335.02'yi ezber kabul etme.
        return "335"

    # Banka markaları — ekstre banka adı ile hesap planı POS adını KESİN eşleştirmek için
    _BANKA_MARKALARI = ["VAKIFBANK", "VAKIF", "GARANTI", "ISBANK", "IS BANKASI", "AKBANK",
                        "QNB", "FINANSBANK", "HALKBANK", "ZIRAAT", "YAPIKREDI", "YAPI KREDI",
                        "TEB", "KUVEYTTURK", "DENIZBANK", "ENPARA", "ING", "SEKERBANK",
                        "ALBARAKA", "FIBABANKA", "ODEABANK", "HSBC", "BURGAN"]

    def _pos_hesap_bul(self) -> tuple[str, bool]:
        """POS tahsilat/hakediş hesabını firmanın hesap planından döner → (kod, kesin_mi).
        GÜVENLİK: Vakıfbank POS'unu Garanti'ye atmamak için, birden çok POS hesabı varsa
        bankayı KESİN eşleştirir; eşleşmezse boş döner (KONTROL ET) — asla yanlış bankaya atmaz."""
        # 1) 02_pos tablosunda bu bankanın POS kodu tanımlıysa (zaten bankaya bağlı) — kesin
        if self.bu_banka_pos_kodu:
            return self.bu_banka_pos_kodu, True
        banka_n = normalize(self.islem_banka_adi or "")
        # ekstrenin bankası hangi marka?
        ekstre_marka = next((m for m in self._BANKA_MARKALARI if m in banka_n), "")
        # firmanın hesap planındaki tüm POS hesapları (adında 'POS' geçen 127/108)
        pos_hesaplar = [h for h in self.tablolar.hesap_plani
                        if (h.kod.startswith("127.") or h.kod.startswith("108."))
                        and "POS" in normalize(h.ad)]
        if not pos_hesaplar:
            kod = _hesap_bul_ilk(self.tablolar, "127") or _hesap_bul_ilk(self.tablolar, "108") or ""
            return kod, False
        # bankası KESİN eşleşen POS hesapları
        if ekstre_marka:
            eslesen = [h for h in pos_hesaplar if ekstre_marka in normalize(h.ad)]
            if eslesen:
                return _yaprak_sec(eslesen, self.tablolar.hesap_plani), True
        # banka eşleşmedi: TEK POS hesabı varsa güvenli (tek banka), yoksa atma → işaretle
        if len(pos_hesaplar) == 1:
            return pos_hesaplar[0].kod, True
        return "", False

    @property
    def _chart_kod_set(self) -> set:
        if not hasattr(self, "_chart_kod_cache"):
            self._chart_kod_cache = {h.kod for h in self.tablolar.hesap_plani}
        return self._chart_kod_cache

    @property
    def _yaprak_kodlar(self) -> set:
        """Hesap planında alt kırılımı OLMAYAN (yaprak) kodlar."""
        if not hasattr(self, "_yaprak_cache"):
            kodlar = {h.kod for h in self.tablolar.hesap_plani}
            self._yaprak_cache = {k for k in kodlar
                                  if not any(o != k and o.startswith(k + ".") for o in kodlar)}
        return self._yaprak_cache

    @property
    def _chart_isim_idx(self):
        """[(kod, token_set)] — sadece yaprak hesaplar (isimle eşleştirme için)."""
        if not hasattr(self, "_chart_idx_cache"):
            self._chart_idx_cache = [(h.kod, set(_anlamli_tokenlar(h.ad)))
                                     for h in self.tablolar.hesap_plani
                                     if h.kod in self._yaprak_kodlar and _anlamli_tokenlar(h.ad)]
        return self._chart_idx_cache

    def _chart_isim_esle(self, aciklama: str, prefixler: tuple, min_ortak: int = 2) -> str:
        """Açıklamadaki isimle hesap planındaki YAPRAK hesabı eşleştir (en çok token örtüşen)."""
        a_tok = set(_anlamli_tokenlar(aciklama))
        if not a_tok:
            return ""
        best = ("", 0)
        for kod, h_tok in self._chart_isim_idx:
            if not any(kod.startswith(p) for p in prefixler):
                continue
            ortak = len(a_tok & h_tok)
            if ortak > best[1]:
                best = (kod, ortak)
        return best[0] if best[1] >= min_ortak else ""

    # Bare parent → otomatik ilk yaprak seçilebilecek gider/sonuç hesapları
    _GIDER_PREFIX = {"780", "770", "760", "335", "642", "679", "653", "659", "656", "646", "780"}

    def _kod_safety(self, e: "Eslesme") -> None:
        """Hesap planına karşı SON güvenlik düzeltmesi (sıra: bare→planda-yok→cari isim eşleşmesi)."""
        if not self.tablolar.hesap_plani:
            return
        kod = (e.hesap_kodu or "").strip()
        # A) bare parent (noktasız, planında alt hesabı olan) → isimle alt hesaba çek
        if kod and "." not in kod and kod not in ("999", "SIL", "SİL"):
            leaves = [h.kod for h in self.tablolar.hesap_plani if h.kod.startswith(kod + ".")]
            if leaves:
                yeni = self._chart_isim_esle(e.hareket.aciklama, (kod,), min_ortak=2)
                if not yeni:
                    if len(set(leaves)) == 1:
                        yeni = leaves[0]
                    elif kod in self._GIDER_PREFIX:
                        yeni = _yaprak_sec([h for h in self.tablolar.hesap_plani
                                            if h.kod.startswith(kod + ".")], self.tablolar.hesap_plani)
                if yeni:
                    e.hesap_kodu = yeni
                    e.not_ = (e.not_ or "") + f" [{kod}→alt hesap]"
                else:
                    e.hesap_kodu = "999"; e.kaynak = "GEÇİCİ"
                    e.not_ = (e.not_ or "") + f" [ana hesap {kod} — alt hesap eşleşmedi, KONTROL ET]"
        # B) kod planda YOK (102 banka, 999, special hariç) → düzelt veya 999 işaretle
        elif kod and kod not in ("999", "SIL", "SİL") and kod not in self._chart_kod_set and not kod.startswith("102"):
            pref = kod.split(".")[0]
            yeni = self._chart_isim_esle(e.hareket.aciklama, (pref,), min_ortak=2)
            if not yeni and pref in ("101", "103") and self.islem_banka_adi:
                yeni = _hesap_bul(self.tablolar, pref, self.islem_banka_adi.split()[0])
            if not yeni:
                leaves = [h.kod for h in self.tablolar.hesap_plani if h.kod.startswith(pref + ".")]
                if len(set(leaves)) == 1:
                    yeni = leaves[0]
            if yeni and yeni in self._chart_kod_set:
                e.not_ = (e.not_ or "") + f" [{kod} planda yok → {yeni}]"
                e.hesap_kodu = yeni
            else:
                e.not_ = (e.not_ or "") + f" [{kod} planda yok, KONTROL ET]"
                e.hesap_kodu = "999"; e.kaynak = "GEÇİCİ"
        # C) SON: hâlâ 999/boş ise isimle cari bul (120/320/329/335)
        if (e.hesap_kodu or "").strip() in ("", "999"):
            na = normalize(e.hareket.aciklama)
            # Açıkça VİRMAN → cariye atama (kendi hesapları arası), 999 kalsın
            if "VIRMAN" in na:
                e.hesap_kodu = "999"; e.kaynak = "GEÇİCİ"
                if "mevduat/virman" not in (e.not_ or ""):
                    e.not_ = (e.not_ or "") + " [VİRMAN — mevduat/hesaplar arası, KONTROL ET]"
                return
            yeni = self._chart_isim_esle(e.hareket.aciklama, ("120", "320", "329", "335"), min_ortak=2)
            if yeni:
                # GÜVENLİK: eşleşen hesabın ADI firmanın KENDİ ünvanı ise (virman riski) → atama
                yeni_ad = next((normalize(h.ad) for h in self.tablolar.hesap_plani if h.kod == yeni), "")
                if self.firma_unvan_marka and self.firma_unvan_marka in yeni_ad:
                    e.hesap_kodu = "999"; e.kaynak = "GEÇİCİ"
                    e.not_ = (e.not_ or "") + " [kendi ünvanı eşleşti — virman olabilir, KONTROL ET]"
                    return
                e.hesap_kodu = yeni
                e.kaynak = "AKILLI"
                e.not_ = (e.not_ or "") + f" [hesap planı isim eşleşmesi → {yeni}]"

    def _arac_gider_hesap_bul(self) -> str:
        """760/770 araç giderleri — firmanın hesap planından oku."""
        for prefix in ("760", "770"):
            for k in self.tablolar.kurallar:
                n_ad = normalize(k.hesap_adi)
                if k.hesap_kodu.startswith(prefix) and ("ARAC" in n_ad or "TASIT" in n_ad or "HGS" in n_ad):
                    return k.hesap_kodu
        return _hesap_bul_ilk(self.tablolar, "760") or _hesap_bul_ilk(self.tablolar, "770") or "770"

    def _genel_yonetim_hesap_bul(self, anahtar: str = "") -> str:
        """770 genel yönetim giderleri — anahtar varsa eşleşen alt hesap."""
        if anahtar:
            kod = _hesap_bul(self.tablolar, "770", anahtar)
            if kod:
                return kod
        return _hesap_bul_ilk(self.tablolar, "770") or "770"

    def _ortak_hesap_kodu(self, ortak_obj=None, negatif: bool = False) -> str:
        """131/331 ortağın hesap kodunu döner. TEK HESAP KURALI: hangisinde bakiye
        varsa HEP o kullanılır (self.ortak_hesap ile belirlenir)."""
        if ortak_obj:
            # Tek hesap kuralı: self.ortak_hesap "131" ise hep 131, "331" ise hep 331
            if self.ortak_hesap == "131" and ortak_obj.kod_131:
                return ortak_obj.kod_131
            if self.ortak_hesap == "331" and ortak_obj.kod_331:
                return ortak_obj.kod_331
            return ortak_obj.kod_131 or ortak_obj.kod_331
        if self.tablolar.ortaklar:
            o = self.tablolar.ortaklar[0]
            if self.ortak_hesap == "131" and o.kod_131:
                return o.kod_131
            if self.ortak_hesap == "331" and o.kod_331:
                return o.kod_331
            return o.kod_131 or o.kod_331
        return _hesap_bul_ilk(self.tablolar, self.ortak_hesap) or self.ortak_hesap

    def _ortak_bul(self, aciklama: str):
        """Açıklamada ortak adı geçiyorsa Ortak nesnesini döner."""
        n_acik = normalize(aciklama)
        h_tok = _anlamli_tokenlar(aciklama)
        for o in self.tablolar.ortaklar:
            o_tok = _anlamli_tokenlar(o.ad)
            if not o_tok:
                continue
            if o_tok.issubset(h_tok) or any(kw and kw in n_acik for kw in o.anahtarlar):
                return o
            if len(o_tok) >= 2:
                eslesme = sum(1 for t in o_tok if t in h_tok or any(t in ht or ht in t for ht in h_tok if len(ht) >= 4 and abs(len(t)-len(ht)) <= 1))
                if eslesme >= len(o_tok) - 1 and eslesme >= 2:
                    soyad_tok = list(o_tok)[-1] if o_tok else ""
                    if soyad_tok and soyad_tok in n_acik:
                        return o
        return None

    def _kural_kodunu_coz(self, kural: Kural, hareket: Hareket) -> tuple[str, str]:
        """Kural tablosundaki eski sabitleri firma hesap planına göre dinamik çözer."""
        kod = (kural.hesap_kodu or "").strip()
        ad = kural.hesap_adi
        if kod in {"780.??", "780.02"} or kod.startswith("780."):
            return self._masraf_hesap_bul(), "FİNANSMAN GİDERİ"
        if kod in {"335.??", "335.02"} or kod.startswith("335."):
            return self._personel_hesap_bul(), "PERSONEL ÖDEMESİ"
        if kod in {"760.??", "760.02"} or kod.startswith("760."):
            return self._arac_gider_hesap_bul(), "ARAÇ GİDERLERİ"
        if kod in {"770.??", "770"} or kod.startswith("770."):
            return self._genel_yonetim_hesap_bul(), "GENEL YÖNETİM GİDERİ"
        if kod in {"193.??", "193"}:
            return self._stopaj_hesap_bul(), "STOPAJ"
        if kod in {"642.??", "642"}:
            return self._faiz_geliri_hesap_bul(), "FAİZ GELİRİ"
        return kod, ad

    # --------------------------------------------------------------------- #
    def esle(self, hareket: Hareket) -> Eslesme:
        n_acik = normalize(hareket.aciklama)
        acik_raw = hareket.aciklama.strip()

        # ── 0) VİRMAN — hesaplar arası transfer ──
        # Şahıs firmalarında ihtiyaç/paylaşım/yardım = ortak, virman değil
        virman = self._virman_tespit(hareket.aciklama)
        if virman:
            if self.sahis_firma and any(kw in n_acik for kw in ("IHTIYAC", "PAYLASIM", "YARDIM")):
                kod = self._ortak_hesap_kodu(negatif=hareket.negatif)
                return Eslesme(hareket, kod, f"{self.ortak_hesap} ORTAK (İHTİYAÇ)", "AKILLI", 85,
                              "Şahıs firma + ihtiyaç/paylaşım → 131/331 (virman değil)")
            kod, banka_ad = virman
            if kod == "VIRMAN_KENDI":
                # Kendi firma virmanı olabilir AMA Vakıfbank gibi bankalarda
                # açıklamada her zaman firma adı geçer. Önce ortak/kural/cari kontrol et.
                # 1) ORTAK kontrolü (131/331)
                ortak = self._ortak_bul(hareket.aciklama)
                if ortak:
                    o_kod = self._ortak_hesap_kodu(ortak, negatif=hareket.negatif)
                    return Eslesme(hareket, o_kod, f"{self.ortak_hesap} ORTAK - {ortak.ad}", "AKILLI", 90,
                                  f"Ortak eşleşti (firma adı da var): {ortak.ad[:25]}")
                # 2) KURAL kontrolü
                for k in self.tablolar.kurallar:
                    for aw in k.anahtarlar:
                        aw_n = normalize(aw)
                        if aw_n and len(aw_n) >= 4 and aw_n in n_acik:
                            kod, ad = self._kural_kodunu_coz(k, hareket)
                            return Eslesme(hareket, kod, ad, "KURAL", 100,
                                          f"Kural eşleşti (firma adı da var): {aw}")
                # 3) CARİ kontrolü — tam substring
                for c in self.tablolar.cariler:
                    n_ad = normalize(c.ad)
                    if n_ad and len(n_ad) >= 4 and n_ad in n_acik:
                        return Eslesme(hareket, c.hesap_kodu, c.ad, "CARI", 85,
                                      f"Cari eşleşti (firma adı da var): {c.ad[:25]}")
                # 4) CARİ kontrolü — token bazlı (kısaltma farkları: SAN/SANAYI)
                _GENEL = {"LTD", "STI", "SAN", "TIC", "AS", "VE", "DIS", "IC", "OTO", "INS",
                          "GIDA", "BILISIM", "SAGLIK", "GRUP", "HOLDING", "SIRKETI", "LIMITED"}
                h_tok = set(n_acik.split())
                for c in self.tablolar.cariler:
                    n_ad = normalize(c.ad)
                    if not n_ad or len(n_ad) < 4:
                        continue
                    c_tok = [t for t in n_ad.split() if t not in _GENEL and len(t) >= 3]
                    if len(c_tok) >= 2 and all(t in h_tok for t in c_tok[:2]):
                        return Eslesme(hareket, c.hesap_kodu, c.ad, "CARI", 85,
                                      f"Cari eşleşti (token): {c.ad[:25]}")
                return Eslesme(hareket, "999", "HESAPLAR ARASI VİRMAN (KENDİ FİRMA)", "VIRMAN", 50,
                              "Kendi firma ünvanı eşleşti — karşı hesap no bulunamadı, DİĞER BANKA İLE KONTROL ET")
            return Eslesme(hareket, kod, f"VİRMAN → {banka_ad}", "VIRMAN", 100,
                          f"Hesaplar arası transfer: {banka_ad}")

        # HESAPLAR ARASI / HESAPLARIM ARASI / VİRMAN kelimesi
        # KK ödemesi "KREDİ KARTI OTOMATİK VİRMAN" → 329, virman değil
        hesaplar_arasi = ("HESAPLAR ARASI" in n_acik or "HESAPLARIM ARASI" in n_acik)
        kredi_karti_virman = ("KREDI KART" in n_acik or "KRE KART" in n_acik or
                              "KKH VIRMAN" in n_acik or "KKH" in n_acik)
        if (hesaplar_arasi or ("VIRMAN" in n_acik and not kredi_karti_virman)):
            # VİRMAN açıklamasında ORTAK ismi varsa → ortak eşleştir
            ortak = self._ortak_bul(hareket.aciklama)
            if ortak:
                o_kod = self._ortak_hesap_kodu(ortak, negatif=hareket.negatif)
                return Eslesme(hareket, o_kod, f"{self.ortak_hesap} ORTAK - {ortak.ad}", "AKILLI", 90,
                              f"VİRMAN açıklamasında ortak eşleşti → {o_kod}")
            # VİRMAN açıklamasında CARİ ismi varsa → cari eşleştir, virman değil
            for c in self.tablolar.cariler:
                n_ad = normalize(c.ad)
                if n_ad and len(n_ad) >= 4 and n_ad in n_acik:
                    return Eslesme(hareket, c.hesap_kodu, c.ad, "CARI", 85,
                                  f"VİRMAN açıklamasında cari isim eşleşti → {c.hesap_kodu}")
            return Eslesme(hareket, "999", "HESAPLAR ARASI VİRMAN", "VIRMAN", 50,
                          "Hesaplar arası virman — karşı hesap bulunamadı, DİĞER BANKA EXCELLERİ İLE KONTROL ET")

        # VERGİ ÖDEMESİ (VERGI-0032 G.GEÇİCİ / VERGI-0015 KDV gibi) → 999 (ortağa ATMA; vergi hesabına elle)
        if hareket.negatif and re.search(r"VERGI[\s-]*\d", n_acik) and "TRANSFER" not in n_acik:
            return Eslesme(hareket, "999", "VERGİ ÖDEMESİ", "AKILLI", 50,
                          "Vergi ödemesi → 999 (ortağa atma; doğru vergi hesabına elle ata)")

        # VERGİ TRANSFERLERİ → şahıs firmalarında = ortak 131/331, şirketlerde virman
        if ("VERGI TRANSFERLERI" in n_acik or "VERGI TRANSFER" in n_acik) and \
           "KESINTI VE EKLER" not in n_acik:
            if self.sahis_firma:
                # Şahıs firma vergi transferi = genelde kendi banka hesapları arası (özelden devlet bankasına tahsil).
                # kaynak=VIRMAN → virman_mutabakat karşı 102'yi bulur; eşleşmezse 999/KONTROL.
                return Eslesme(hareket, "999", "VERGİ TRANSFERİ (hesaplar arası?)", "VIRMAN", 50,
                              "Şahıs vergi transferi → hesaplar arası virman aranıyor, eşleşmezse KONTROL ET")
            virman = self._virman_tespit(hareket.aciklama)
            if virman:
                kod, banka_ad = virman
                return Eslesme(hareket, kod, f"VERGİ TRANSFERİ → {banka_ad}", "VIRMAN", 85,
                              f"Vergi transferi = hesaplar arası virman")
            return Eslesme(hareket, "999", "VERGİ TRANSFERİ", "AKILLI", 50,
                          "Vergi transferi — virman olabilir, DİĞER BANKA EXCELLERİ İLE KONTROL ET")

        # ── 0.5) KESİNTİ VE EKLERİ = HER ZAMAN 780 MASRAF ──
        # (maaş/avans/prim kelimesi olsa bile — banka masrafı)
        if "KESINTI VE EKLER" in n_acik or "KESINTI VE EKLER" in n_acik:
            kod = self._masraf_hesap_bul()
            return Eslesme(hareket, kod, "BANKA MASRAFI (KESİNTİ)", "AKILLI", 100,
                          "KESİNTİ VE EKLERİ = banka masrafı → 780")

        # ── 0.6) BRÜT FAİZ / FAİZ GELİRİ (POZİTİF) → 642 ──
        if not hareket.negatif and ("FAIZ" in n_acik or "FAİZ" in n_acik) and "BRUT" in n_acik:
            kod = self._faiz_geliri_hesap_bul()
            return Eslesme(hareket, kod, "FAİZ GELİRİ", "AKILLI", 95,
                          f"Brüt faiz geliri → 642 ({self.islem_banka_adi})")

        # ── 0.7) GELİR VERGİSİ / STOPAJ (NEGATİF) → 193 ──
        if hareket.negatif and ("STOPAJ" in n_acik or ("GELIR VERGISI" in n_acik and "FAIZ" not in n_acik)):
            kod = self._stopaj_hesap_bul()
            return Eslesme(hareket, kod, "STOPAJ KESİNTİSİ", "AKILLI", 95,
                          f"Stopaj → 193 ({self.islem_banka_adi})")

        # VERGİ + hesap no → stopaj olabilir (ör: "00211/6394052 VERGİ")
        if hareket.negatif and "VERGI" in n_acik and re.search(r"\d{4,}\s+\d+\s+VERGI", n_acik):
            kod = self._stopaj_hesap_bul()
            return Eslesme(hareket, kod, "STOPAJ (VERGİ KESİNTİSİ)", "AKILLI", 85,
                          f"Faiz stopajı → 193 ({self.islem_banka_adi})")

        # POZİTİF FAİZ (BRÜT olmadan da) → 642
        if not hareket.negatif and ("FAIZ" in n_acik) and "KREDI" not in n_acik:
            kod = self._faiz_geliri_hesap_bul()
            return Eslesme(hareket, kod, "FAİZ GELİRİ", "AKILLI", 90,
                          f"Faiz geliri → 642 ({self.islem_banka_adi})")

        # ── 0.8) MEVDUAT GERİ DÖNÜŞÜ → 102.xx vadeli hesap ──
        if "MEVDUAT GERI DONUSU" in n_acik or "MEVDUAT GERI" in n_acik:
            kod = self._mevduat_hesap_bul(hareket.aciklama)
            if kod:
                return Eslesme(hareket, kod, "VADELİ HESAP (MEVDUAT)", "AKILLI", 95,
                              "Mevduat geri dönüşü → 102 vadeli hesap")
            return Eslesme(hareket, "999", "MEVDUAT GERİ DÖNÜŞÜ", "AKILLI", 50,
                          "Mevduat geri dönüşü → vadeli hesap bulunamadı, KONTROL ET")

        # ── 0.9) VADELİ HSP AÇILIŞI → 102.xx vadeli hesap ──
        if "VADELI HSP" in n_acik or "VADELI HESAP" in n_acik:
            kod = self._mevduat_hesap_bul(hareket.aciklama)
            if kod:
                return Eslesme(hareket, kod, "VADELİ HESAP AÇILIŞI", "AKILLI", 90,
                              "Vadeli hesap açılışı → 102 vadeli hesap")

        # ── 1.0) FON ALIM/SATIM → 102.xx fon hesabı ──
        if "FON SATIS" in n_acik or "FON ALIM" in n_acik or "FON ALISI" in n_acik or "PARA PIYASA" in n_acik:
            kod = self._fon_hesap_bul()
            if kod:
                return Eslesme(hareket, kod, "FON HESABI", "AKILLI", 95,
                              "Fon alım/satım → 102 fon hesabı (118 karşı hesap)")

        # ── 1.1) EKSTRE BORÇ TAHSİLATI → 329 KK ödemesi ──
        if "EKSTRE BORC TAHSILATI" in n_acik or "EKSTRE BORÇ TAHSİLATI" in n_acik:
            kk = self._kk_kart_no_esle(acik_raw)
            if kk:
                return Eslesme(hareket, kk[0], kk[1], "AKILLI", 95,
                              f"KK ödemesi → 329 (kart: {kk[1][:30]})")
            kod = _hesap_bul_ilk(self.tablolar, "329") or "999"
            kk_kodlar = set(c.hesap_kodu for c in self.tablolar.cariler if c.hesap_kodu.startswith("329"))
            kk_kodlar |= set(k.hesap_kodu for k in self.tablolar.kurallar if k.hesap_kodu.startswith("329"))
            guven = 85 if len(kk_kodlar) <= 1 else 60
            return Eslesme(hareket, kod, "KK ÖDEMESİ", "AKILLI", guven,
                          "Ekstre borç tahsilatı → 329" + (" (tek kart)" if len(kk_kodlar) <= 1 else " — kart eşleşmedi, KONTROL ET"))

        # ── 1.2) K.Kartı Ödeme / KRE.KART / KKH VİRMAN / MUH.HESABINA ──
        if ("KARTI ODEME" in n_acik or "K KARTI ODEME" in n_acik or "KART ODEME" in n_acik or
            "KREDI KARTI" in n_acik or "KRE KART" in n_acik or "KKH VIRMAN" in n_acik or
            ("MUH HESABINA" in n_acik and "KRE" in n_acik)):
            # Şahıs firmalarında şirket kartı YOK → 131/331
            if self.sahis_firma:
                kod = self._ortak_hesap_kodu(negatif=hareket.negatif)
                return Eslesme(hareket, kod, f"{self.ortak_hesap} ORTAK (KK YOK)", "AKILLI", 90,
                              "Şahıs firmasında şirket kartı yok → 131/331")
            # Kart son 4 hane bul (masked: 4273 **** **** 8010 veya 16 haneli: 4988520000030176)
            m = re.search(r"(\d{4})\s*\*+\s*\*+\s*(\d{4})", acik_raw)
            kart_ilk4 = ""
            kart_son4 = ""
            if m:
                kart_ilk4 = m.group(1)
                kart_son4 = m.group(2)
            else:
                m16 = re.search(r"(\d{16})", acik_raw)
                if m16:
                    kart_ilk4 = m16.group(1)[:4]
                    kart_son4 = m16.group(1)[-4:]
            if kart_ilk4 or kart_son4:
                for arama in [kart_son4, kart_ilk4]:
                    if not arama:
                        continue
                    for c in self.tablolar.cariler:
                        if c.hesap_kodu.startswith("329") and arama in c.ad:
                            return Eslesme(hareket, c.hesap_kodu, c.ad, "AKILLI", 95,
                                          f"KK ödemesi → 329 (kart: {arama})")
                    for k in self.tablolar.kurallar:
                        if k.hesap_kodu.startswith("329") and arama in k.hesap_adi:
                            return Eslesme(hareket, k.hesap_kodu, k.hesap_adi, "AKILLI", 95,
                                          f"KK ödemesi → 329 (kart: {arama})")
            kod = _hesap_bul_ilk(self.tablolar, "329") or "999"
            kk_kodlar = set(c.hesap_kodu for c in self.tablolar.cariler if c.hesap_kodu.startswith("329"))
            kk_kodlar |= set(k.hesap_kodu for k in self.tablolar.kurallar if k.hesap_kodu.startswith("329"))
            guven = 85 if len(kk_kodlar) <= 1 else 60
            return Eslesme(hareket, kod, "KK ÖDEMESİ", "AKILLI", guven,
                          "KK ödemesi → 329" + (" (tek kart)" if len(kk_kodlar) <= 1 else " — kart bulunamadı, KONTROL ET"))

        # ── 1.3) A.BUSINESS OTM.ÖDEME = Akbank KK ödemesi ──
        if "BUSINESS OTM" in n_acik or "A BUSINESS OTM" in n_acik:
            kod = _hesap_bul(self.tablolar, "329", "AKBANK") or _hesap_bul_ilk(self.tablolar, "329")
            if kod:
                return Eslesme(hareket, kod, "AKBANK KK ÖDEMESİ", "AKILLI", 90,
                              "A.BUSINESS OTM.ÖDEME = Akbank KK → 329")
            return Eslesme(hareket, "999", "AKBANK KK ÖDEMESİ", "AKILLI", 50,
                          "A.BUSINESS OTM.ÖDEME — 329 hesabı bulunamadı")

        # ── 1.4) DEVLET DESTEĞİ → 679 ──
        if "DEVLET DESTEK" in n_acik or "DEVLET DEST" in n_acik:
            kod = _hesap_bul_ilk(self.tablolar, "679") or "679"
            return Eslesme(hareket, kod, "DEVLET DESTEĞİ GELİRİ", "AKILLI", 95,
                          "Devlet desteği → 679")

        # ── 1.5) KMH ANAPARA → 300 (banka adına göre alt hesap) ──
        if "KMH" in n_acik and ("ANAPARA" in n_acik or "ANA PARA" in n_acik):
            kod = _hesap_bul(self.tablolar, "300", self.islem_banka_adi) if self.islem_banka_adi else ""
            if not kod:
                kod = _hesap_bul_ilk(self.tablolar, "300") or "300"
            return Eslesme(hareket, kod, "KREDİLİ MEVDUAT HESABI", "AKILLI", 90,
                          f"KMH anapara borcu → {kod}")

        # ── 1.6) DEPOZİTO → 126 ──
        if "DEPOZIT" in n_acik or "DEPOZİT" in n_acik:
            kod = _hesap_bul_ilk(self.tablolar, "126") or "126"
            return Eslesme(hareket, kod, "DEPOZİTO", "AKILLI", 90,
                          "Depozito → 126")

        # ── 1.7) SENET ÖDEMESİ ──
        # Gelen (pozitif) → 121 alacak senetleri / 120 alıcılar
        # Giden (negatif) → 321 borç senetleri / 320 satıcılar
        if "SENET" in n_acik and ("ODEME" in n_acik or "VADELI" in n_acik or "ODE" in n_acik):
            if hareket.negatif:
                arama_onekleri = ["321", "320"]
                varsayilan = "321"
            else:
                arama_onekleri = ["121", "120"]
                varsayilan = "121"
            for onek in arama_onekleri:
                for c in self.tablolar.cariler:
                    if c.hesap_kodu.startswith(onek):
                        n_ad = normalize(c.ad)
                        if n_ad and len(n_ad) >= 4 and n_ad in n_acik:
                            return Eslesme(hareket, c.hesap_kodu, c.ad, "AKILLI", 90,
                                          f"Senet ödemesi → {c.hesap_kodu} ({c.ad[:25]})")
            kod = _hesap_bul_ilk(self.tablolar, varsayilan) or varsayilan
            return Eslesme(hareket, kod, "SENET ÖDEMESİ", "AKILLI", 50,
                          f"Senet ödemesi → {varsayilan} (ünvan eşleşmedi, KONTROL ET)")

        # ── 1.8) HGS / ARAÇ GİDERLERİ → 760/770 ──
        if "HGS" in n_acik or "OGS" in n_acik:
            kod = self._arac_gider_hesap_bul()
            return Eslesme(hareket, kod, "ARAÇ GİDERİ (HGS)", "AKILLI", 95,
                          "HGS/OGS → araç gideri")

        # ── 1.9) AİDAT ÖDEMESİ → 770 (alt hesap ara) ──
        if "AIDAT" in n_acik and hareket.negatif:
            kod = self._genel_yonetim_hesap_bul("AIDAT")
            return Eslesme(hareket, kod, "AİDAT (GENEL YÖNETİM)", "AKILLI", 90,
                          f"Aidat ödemesi → {kod}")

        # ── 2.0) KİRA ÖDEMESİ → 329 isimle eşleştir ──
        if ("KIRA" in n_acik or "KİRA" in n_acik) and hareket.negatif:
            # İsim bul ve 329'da ara
            for c in self.tablolar.cariler:
                if c.hesap_kodu.startswith("329"):
                    n_ad = normalize(c.ad)
                    if n_ad and len(n_ad) >= 4 and n_ad in n_acik:
                        return Eslesme(hareket, c.hesap_kodu, c.ad, "AKILLI", 95,
                                      f"Kira ödemesi → 329 ({c.ad[:25]})")
            # 329 CARİLERDE eşleşme yoksa — kural sözlüğündeki KK kuralını alma!
            # Sadece carilerden ilk 329 alt hesabını bul
            cari_329 = next((c.hesap_kodu for c in self.tablolar.cariler
                            if c.hesap_kodu.startswith("329")), "")
            kod = cari_329 or "329"
            return Eslesme(hareket, kod, "KİRA ÖDEMESİ", "AKILLI", 85,
                          "Kira ödemesi → 329 — isim eşleşmedi, alt hesap kontrol et")

        # ── 2.05) DAMGA VERGİSİ → 770 ──
        if "DAMGA" in n_acik and "VERGI" in n_acik:
            kod = self._genel_yonetim_hesap_bul()
            return Eslesme(hareket, kod, "DAMGA VERGİSİ (770)", "AKILLI", 90,
                          "Damga vergisi → 770 genel yönetim")

        # ── 2.06) SGK ÖDEMESİ (non-şahıs firmalar) → 361 ──
        if not self.sahis_firma and hareket.negatif and ("SGK" in n_acik or "SSK" in n_acik) and \
           ("EMS" in n_acik or "BAG" in n_acik or "ODE" in n_acik or "TAHSILAT" in n_acik):
            kod = _hesap_bul_ilk(self.tablolar, "361") or "361"
            return Eslesme(hareket, kod, "SGK ÖDEMESİ", "AKILLI", 85,
                          "SGK/SSK ödemesi → 361")

        # ── 2.07) MAAŞ ÖDEMESİ (genel — ortak olmasına gerek yok) → 335 ──
        if hareket.negatif and ("MAAS ODEMESI" in n_acik or "MAAS ODE" in n_acik or
                                "UCRET ODEMESI" in n_acik or "YOL UCRETI" in n_acik or
                                "YEMEK UCRETI" in n_acik or "BAYRAM ODEMESI" in n_acik or
                                "MESAI ODEMESI" in n_acik or "AVANS ODEMESI" in n_acik):
            kod = self._personel_hesap_bul()
            return Eslesme(hareket, kod, "335 PERSONEL ÖDEMESİ", "AKILLI", 85,
                          "Maaş/ücret/avans ödemesi → 335 personel")

        # ── 2.08) İTHALAT (PEŞİN İTH / İTH İŞL MAL BED) → 320 cari ──
        if "ITH" in n_acik and ("PESIN" in n_acik or "MAL BED" in n_acik or "ITHALAT" in n_acik):
            for c in self.tablolar.cariler:
                if c.hesap_kodu.startswith("320"):
                    n_ad = normalize(c.ad)
                    if n_ad and len(n_ad) >= 4 and n_ad in n_acik:
                        return Eslesme(hareket, c.hesap_kodu, c.ad, "AKILLI", 85,
                                      f"İthalat ödemesi → 320 ({c.ad[:25]})")
            return Eslesme(hareket, "999", "İTHALAT ÖDEMESİ", "AKILLI", 40,
                          "İthalat ödemesi — caride bulunamadı, KONTROL ET")

        # ── 2.09) PROVIZYON (KASADAN ÖDENE) → 100 kasa ──
        if "PROVIZYON" in n_acik and ("KASA" in n_acik or "KASADAN" in n_acik):
            kod = _hesap_bul_ilk(self.tablolar, "100") or "100"
            return Eslesme(hareket, kod, "KASA (PROVİZYON)", "AKILLI", 80,
                          "Provizyon kasadan → 100 kasa")

        # ── 2.095) İşBank EFT/HAVALE KOMİSYONU "ÜCRET H26xxxxx ... TRY ÜZ." → 780 ──
        if hareket.negatif and "UCRET H" in n_acik and "TRY UZ" in n_acik:
            kod = self._masraf_hesap_bul()
            return Eslesme(hareket, kod, "BANKA MASRAFI (İŞBANK KOMİSYON)", "AKILLI", 95,
                          "İşBank EFT/Havale ücreti → 780 masraf")

        # ── 2.096) TAKSİTLİ KREDİ → 300 (banka adına göre alt hesap) ──
        if hareket.negatif and "TAKSITLI" in n_acik and re.search(r"\d{6,}", acik_raw):
            kod = _hesap_bul(self.tablolar, "300", self.islem_banka_adi) if self.islem_banka_adi else ""
            if not kod:
                kod = _hesap_bul_ilk(self.tablolar, "300") or "300"
            return Eslesme(hareket, kod, "TAKSİTLİ KREDİ", "AKILLI", 85,
                          f"Taksitli kredi ödemesi → {kod}")

        # ── 2.097) TRKDJT (Turkcell Dijital fatura) → 320 satıcı ──
        if "TRKDJT" in n_acik:
            kod = _hesap_bul(self.tablolar, "320", "TURKCELL") or _hesap_bul_ilk(self.tablolar, "320") or "320"
            return Eslesme(hareket, kod, "TRKDJT (TURKCELL DİJİTAL)", "AKILLI", 90,
                          "Turkcell Dijital fatura ödemesi → 320 satıcı")

        # ── 2.098) POS hakediş/tahsilat "nolu isyerinin...pos ve batchinin" → firma POS hesabı (127/108, hesap planından, BANKAYA GÖRE) ──
        if not hareket.negatif and "NOLU ISYERININ" in n_acik and "POS" in n_acik:
            kod, kesin = self._pos_hesap_bul()
            if kod and kesin:
                return Eslesme(hareket, kod, "POS HAKEDİŞ/TAHSİLAT", "AKILLI", 95,
                              f"POS satış hakedişi → {kod} ({self.islem_banka_adi})")
            return Eslesme(hareket, kod or "999", "POS HAKEDİŞ (BANKA EŞLEŞMEDİ)", "AKILLI", 40,
                          f"POS satışı — '{self.islem_banka_adi}' bankasının POS hesabı kesin bulunamadı, KONTROL ET")

        # ── 2.099) YATIRIM HESABINDAN AKTARMA (şirket firmalar) → virman yatırım hesabı ──
        if not self.sahis_firma and "YATIRIM" in n_acik and ("AKTARMA" in n_acik or "AKTARIM" in n_acik):
            yatirim = next((b for b in self.tablolar.bankalar
                           if "YATIRIM" in normalize(b.hesap_adi or "") and b.hesap_kodu.startswith("102")),
                          None)
            if yatirim:
                return Eslesme(hareket, yatirim.hesap_kodu, f"VİRMAN → {yatirim.banka} YATIRIM",
                              "VIRMAN", 90, f"Yatırım hesabı aktarma → {yatirim.hesap_kodu}")
            return Eslesme(hareket, "999", "YATIRIM HESABI AKTARMA", "VIRMAN", 50,
                          "Yatırım hesabından aktarma — yatırım hesabı bulunamadı, KONTROL ET")

        # ── 2.0995) HVL-INT / HVL-CEP personel ödemeleri (Yol Parası, Mesai vb.) → 335 ──
        if hareket.negatif and ("HVL INT" in n_acik or "HVL CEP" in n_acik):
            if any(kw in n_acik for kw in ("YOL PARASI", "MESAI", "TEZMAN", "YOLLUK")):
                kod = self._personel_hesap_bul()
                return Eslesme(hareket, kod, "335 PERSONEL (YOL/MESAİ)", "AKILLI", 85,
                              "HVL personel yol parası/mesai → 335")

        # ── 2.1) ARTI PARA VERGİ / MKK → 780 masraf ──
        if "ARTI PARA VERGI" in n_acik or "MKK" in n_acik:
            kod = self._masraf_hesap_bul()
            return Eslesme(hareket, kod, "BANKA MASRAFI", "AKILLI", 90,
                          "Artı Para Vergi / MKK → 780 masraf")

        # ── 2.2) POS BAKIM/YAZILIM ÜCRETİ → 780 masraf ──
        if hareket.negatif and ("POS YAZILIM" in n_acik or "POS BAKIM" in n_acik or
                                ("POS" in n_acik and ("DNNM" in n_acik or "DONM" in n_acik))):
            kod = self._masraf_hesap_bul()
            return Eslesme(hareket, kod, "POS BAKIM/YAZILIM ÜCRETİ", "AKILLI", 95,
                          "POS bakım ücreti → 780 masraf")

        # ── 2.3) QR P.ÇEKME / BANKAMATİK / ATM → 100 KASA ──
        # AMA isim yazıyorsa 131/331!
        if ("QR" in n_acik or "BANKAMATIK" in n_acik) and ("CEKME" in n_acik or "PARA CEKME" in n_acik):
            kod = _hesap_bul_ilk(self.tablolar, "100") or "100"
            return Eslesme(hareket, kod, "TL KASA (ATM/QR)", "AKILLI", 90,
                          "QR para çekme → 100 kasa")

        # PARA YATIRMA — isim yoksa 100, isim varsa 131/331
        if "PARA YATIRMA" in n_acik or "NAKIT YATIRMA" in n_acik:
            # İsim var mı kontrol et
            ortak = self._ortak_bul(hareket.aciklama)
            if ortak:
                kod = self._ortak_hesap_kodu(ortak, negatif=hareket.negatif)
                return Eslesme(hareket, kod, f"{self.ortak_hesap} ORTAK - {ortak.ad}", "AKILLI", 85,
                              f"Para yatırma + ortak adı → {self.ortak_hesap}")
            kod = _hesap_bul_ilk(self.tablolar, "100") or "100"
            return Eslesme(hareket, kod, "TL KASA", "AKILLI", 90,
                          "Para yatırma → 100 kasa")

        # ── 2.4) DÖVİZ SATIM/ALIM (TL hesabında) → TL hesabından SİL ──
        if "DOVIZ SATIM" in n_acik or "DOVIZ SATIS" in n_acik or "DOVIZ ALIM" in n_acik:
            return Eslesme(hareket, "SIL", "DÖVİZ İŞLEMİ (TL HESABINDAN SİL)", "AKILLI", 100,
                          "Döviz alım/satım — TL hesabından sil, döviz hesabında kalsın")

        # ── 2.5) TANER BATTAL (muhasebeci) → ÖNCE 329 (borç), yoksa 195 — ASLA 120/320 DEĞİL ──
        if "TANER BATTAL" in n_acik:
            kod = _hesap_bul(self.tablolar, "329", "BATTAL") or _hesap_bul(self.tablolar, "195", "BATTAL")
            if not kod:
                kod = _hesap_bul_ilk(self.tablolar, "329") or _hesap_bul_ilk(self.tablolar, "195") or "329"
            return Eslesme(hareket, kod, "MUHASEBE ÖDEMESİ (TANER BATTAL)", "AKILLI", 95,
                          "Taner Battal → 329 muhasebeci ödemesi (yoksa 195)")

        # ── 2.6) Maskeli isim (FU**** ER****) → 999 ──
        if re.search(r"[A-Z]{2}\*{3,}", n_acik):
            return Eslesme(hareket, "999", "GEÇİCİ HESAP", "AKILLI", 30,
                          "Maskeli isim → 999 (tam ad belirtilmedi)")

        # ── 2.7) ÇEK ÖDEMESİ / ÇEK TAHSİLATI → 103/101 ──
        if ("CEK" in n_acik or "ÇEK" in n_acik) and ("TAKAS" in n_acik or "TKS" in n_acik or "ODEME" in n_acik
                                                        or "CIKIS" in n_acik or "IADE" in n_acik or "TAHSIL" in n_acik
                                                        or "GIRIS" in n_acik or "BORDRO" in n_acik):
            if hareket.negatif:
                kod = self.bu_banka_cek_kodu or \
                      _hesap_bul(self.tablolar, "103", self.islem_banka_adi) or \
                      _hesap_bul_ilk(self.tablolar, "103") or "103"
                return Eslesme(hareket, kod, "VERİLEN ÇEKLER", "AKILLI", 90,
                              f"Çek ödemesi → {kod}")
            else:
                kod = self.bu_banka_cek_tahsil_kodu or \
                      _hesap_bul(self.tablolar, "101", self.islem_banka_adi) or \
                      _hesap_bul_ilk(self.tablolar, "101") or "101"
                return Eslesme(hareket, kod, "ALINAN ÇEKLER", "AKILLI", 90,
                              f"Çek tahsilatı → {kod}")

        # ── 2.8) Akıllı kısaltma tanıma (MULTNT, TRKCLL vs.) → cari arat ──
        for kisaltma, tam_ad in KISALTMA_HARITASI.items():
            if kisaltma in n_acik:
                n_tam = normalize(tam_ad)
                for c in self.tablolar.cariler:
                    if n_tam in normalize(c.ad):
                        return Eslesme(hareket, c.hesap_kodu, c.ad, "AKILLI", 85,
                                      f"Kısaltma {kisaltma} → {c.ad[:30]}")
                # Caride yoksa TAM hesap planında (06) 320/120 ara
                for h in self.tablolar.hesap_plani:
                    if (h.kod.startswith("320") or h.kod.startswith("120")) and n_tam in normalize(h.ad):
                        return Eslesme(hareket, h.kod, h.ad, "AKILLI", 85,
                                      f"Kısaltma {kisaltma} → {h.ad[:30]}")
                # Bulunamadı → 999
                return Eslesme(hareket, "999", f"CARİ ({tam_ad})", "AKILLI", 40,
                              f"Kısaltma {kisaltma} = {tam_ad} ama bulunamadı")

        # ── 2.85) HUZUR HAKKI / İHBAR TAZMİNATI / PRİM ÖDEMESİ (ad soyad ile) → 335 ──
        # Ortaklar tablosunda olup olmadığına bakılmaksızın, açıklamada bu kelimeler + isim varsa 335
        personel_kelimeleri = ("HUZUR HAKKI" in n_acik or "IHBAR TAZMINATI" in n_acik or
                               "KIDEM TAZMINATI" in n_acik)
        if personel_kelimeleri and hareket.negatif:
            kod = self._personel_hesap_bul()
            return Eslesme(hareket, kod, "335 PERSONEL ÖDEMESİ", "AKILLI", 90,
                          "Huzur hakkı/ihbar/kıdem tazminatı → 335 personel")

        # Ad soyad + PRİM ÖDEMESİ (AMA KESİNTİ VE EKLERİ-Prim değil, o yukarıda 780 olarak yakalandı)
        if ("PRIM ODEMESI" in n_acik or "PRİM ÖDEMESİ" in n_acik) and hareket.negatif:
            kod = self._personel_hesap_bul()
            return Eslesme(hareket, kod, "335 PERSONEL (PRİM)", "AKILLI", 85,
                          "Prim ödemesi → 335 personel (KESİNTİ VE EKLERİ-Prim = 780, bu farklı)")

        # ── 2.9) Şahıs firmalarında özel kurallar ──
        if self.sahis_firma:
            # SGK/BAĞKUR ödemesi → şahıs firmalarında HER ZAMAN 131/331 (şahsi ödeme)
            if ("SGK" in n_acik or "BAGKUR" in n_acik or "BAG KUR" in n_acik) and hareket.negatif:
                kod = self._ortak_hesap_kodu(negatif=True)
                return Eslesme(hareket, kod, f"{self.ortak_hesap} ORTAK (SGK ŞAHSİ)", "AKILLI", 85,
                              "Şahıs firmada SGK/Bağkur → şahsi ödeme 131/331")
            # ALTIN/GÜMÜŞ/KIYMETLİ MADEN → 131/331
            if "ALTIN" in n_acik or "GUMUS" in n_acik or "KIYMETLI MADEN" in n_acik or "MEVDUAT" in n_acik and "ALTIN" in n_acik:
                kod = self._ortak_hesap_kodu(negatif=hareket.negatif)
                return Eslesme(hareket, kod, f"{self.ortak_hesap} ORTAK (KIYMETLİ MADEN)", "AKILLI", 85,
                              "Altın/gümüş alımı → 131/331 (şahsi)")
            # YATIRIM HESABI → 131/331
            if "YATIRIM" in n_acik and ("AKTARMA" in n_acik or "HESAB" in n_acik):
                kod = self._ortak_hesap_kodu(negatif=hareket.negatif)
                return Eslesme(hareket, kod, f"{self.ortak_hesap} ORTAK (YATIRIM)", "AKILLI", 85,
                              "Yatırım hesabı → 131/331 (şahıs firmasında)")
            # AKTARIM → 131/331 (120.02 DEĞİL!)
            if "AKTARIM" in n_acik:
                kod = self._ortak_hesap_kodu(negatif=hareket.negatif)
                return Eslesme(hareket, kod, f"{self.ortak_hesap} ORTAK (AKTARIM)", "AKILLI", 80,
                              "Aktarım → 131/331")
            # KKH VİRMAN / MUH.HESABINA → 131/331 (kart takip etmiyoruz)
            if "KKH VIRMAN" in n_acik or "MUH HESABINA" in n_acik:
                kod = self._ortak_hesap_kodu(negatif=hareket.negatif)
                return Eslesme(hareket, kod, f"{self.ortak_hesap} ORTAK", "AKILLI", 85,
                              "KKH virman / muhasebe hesabına → 131/331")

        # ── 2.95) POS TAHSİLATI (NET SATIŞ TUTARI) → 127 / negatif → 780 masraf ──
        if "NET SATIS TUTARI" in n_acik or "POS TAHSILAT" in n_acik:
            if hareket.negatif:
                kod = self._masraf_hesap_bul()
                return Eslesme(hareket, kod, "POS KOMİSYON KESİNTİSİ", "AKILLI", 90,
                              "POS net satış tutarı (negatif) → 780 masraf/komisyon")
            kod = _hesap_bul(self.tablolar, "127", self.islem_banka_adi) or _hesap_bul_ilk(self.tablolar, "127") or "127"
            return Eslesme(hareket, kod, "POS TAHSİLATI", "AKILLI", 90,
                          f"POS tahsilatı → 127 ({self.islem_banka_adi})")

        # ── 2.96) FATURA ÖDEMELERİ (AESAŞ/BEDAŞ/İGDAŞ/İSKİ/VODAFONE vb.) ──
        # ÖNCE 320 carilerinde ara, yoksa 780/770, en son fallback 780
        _FATURA_ESLEME = {
            "AESAS": ["AESAS", "ENERJI", "ENERJISA"], "AYESAS": ["AYESAS", "ENERJI", "ENERJISA"],
            "BEDAS": ["BEDAS", "ENERJI", "ENERJISA"], "ENERJISA": ["ENERJI", "ENERJISA"],
            "IGDAS": ["IGDAS"], "ISTGAZD": ["IGDAS"],
            "ISKI": ["ISKI"], "ISKIFO": ["ISKI"],
            "VODAFO": ["VODAFON", "VODAFONE"], "VODAFON": ["VODAFON", "VODAFONE"],
            "TURK TELEKOM": ["TURK TELEKOM", "TELEKOM"], "TURKCELL": ["TURKCELL"],
            "SUPERONLINE": ["SUPERONLINE"], "TRKDJT": ["TURKCELL"],
        }
        fatura_bulunan = None
        for kw in _FATURA_ESLEME:
            if kw in n_acik:
                fatura_bulunan = _FATURA_ESLEME[kw]
                break
        if hareket.negatif and fatura_bulunan:
            # 1) 320 carilerinde ara (kurum adı eşleşmesi)
            for c in self.tablolar.cariler:
                if not c.hesap_kodu.startswith("320"):
                    continue
                n_ad = normalize(c.ad)
                if not n_ad or len(n_ad) < 3:
                    continue
                # Cari adı açıklamada geçiyor mu?
                if n_ad in n_acik:
                    return Eslesme(hareket, c.hesap_kodu, c.ad, "AKILLI", 92,
                                  f"Fatura ödemesi → {c.hesap_kodu} ({c.ad[:30]})")
                # Fatura anahtar kelimeleri cari adında geçiyor mu?
                for esleme_kw in fatura_bulunan:
                    if esleme_kw in n_ad:
                        return Eslesme(hareket, c.hesap_kodu, c.ad, "AKILLI", 90,
                                      f"Fatura ödemesi → {c.hesap_kodu} ({c.ad[:30]})")
            # 2) 780/770 carilerinde ara
            for c in self.tablolar.cariler:
                if c.hesap_kodu.startswith("780") or c.hesap_kodu.startswith("770"):
                    n_ad = normalize(c.ad)
                    if n_ad and len(n_ad) >= 3 and n_ad in n_acik:
                        return Eslesme(hareket, c.hesap_kodu, c.ad, "AKILLI", 90,
                                      f"Fatura ödemesi → {c.hesap_kodu}")
            # 3) Fallback: 780 masraf
            kod = self._masraf_hesap_bul()
            return Eslesme(hareket, kod, "FATURA ÖDEMESİ (780)", "AKILLI", 75,
                          "Fatura ödemesi — 320 carilerinde bulunamadı → 780 masraf")

        # ── 2.97) SGK FT.NO (eczane fatura tahsilatı) → 120 ──
        if not hareket.negatif and "SGK" in n_acik and "FT" in n_acik:
            kod = _hesap_bul_ilk(self.tablolar, "120") or "120"
            return Eslesme(hareket, kod, "SGK FATURA TAHSİLATI", "AKILLI", 85,
                          "SGK fatura no → 120 müşteri alacağı")

        # ── 2.98) BORÇ VERME → 131 ──
        if not hareket.negatif and "BORC VERME" in n_acik:
            kod = _hesap_bul_ilk(self.tablolar, "131") or "131"
            return Eslesme(hareket, kod, "BORÇ VERME (131)", "AKILLI", 80,
                          "Borç verme → 131 ortaklardan alacaklar")

        # ── 3) Kural sözlüğü ──
        for k in self.tablolar.kurallar:
            if not k.aktif:
                continue
            if k.tutar_isareti == "NEGATIF" and not hareket.negatif:
                continue
            if k.tutar_isareti == "POZITIF" and hareket.negatif:
                continue
            for kw in k.anahtarlar:
                if kw and kw in n_acik:
                    kod, ad = self._kural_kodunu_coz(k, hareket)
                    if kod == "108.??":
                        kod, kesin = self._pos_hesap_bul()
                        if not kod:
                            return Eslesme(hareket, "999", "POS HAKEDİŞ (BANKA EŞLEŞMEDİ)", "AKILLI", 40,
                                           f"POS — '{self.islem_banka_adi}' POS hesabı kesin bulunamadı, KONTROL ET")
                        pos = next((p for p in self.tablolar.posler if p.hesap_kodu == kod), None)
                        ad = pos.pos_adi if pos else "POS HAKEDİŞ/TAHSİLAT"
                    elif kod == "103.??":
                        kod = self.bu_banka_cek_kodu or "103"
                        ad = "VERİLEN ÇEKLER (103)"
                    elif kod == "101.??":
                        kod = self.bu_banka_cek_tahsil_kodu or "101"
                        ad = "ALINAN ÇEKLER (101)"
                    not_metni = k.not_ if k.not_ else f"Kural #{k.sira}: {kw}"
                    kaynak = "KURAL" if kod else "MANUEL"
                    return Eslesme(hareket, kod, ad, kaynak, k.guven, not_metni)

        # ── 3.5) ORTAK/yetkili hareketi → 131/331 veya 335 ──
        if self.tablolar.ortaklar:
            ortak = self._ortak_bul(hareket.aciklama)
            if ortak:
                # HUZUR HAKKI / MAAŞ / PRİM / İHBAR / KIDEM → 335
                personel_kelime = ("HUZUR HAKKI" in n_acik or "MAAS" in n_acik or "MAAŞ" in n_acik or
                                   "PRIM" in n_acik or "PRİM" in n_acik or "IHBAR" in n_acik or
                                   "İHBAR" in n_acik or "KIDEM" in n_acik or "KÍDEM" in n_acik or
                                   "UCRET ODEME" in n_acik or "ÜCRET ÖDEME" in n_acik)
                if personel_kelime:
                    kod = self._personel_hesap_bul(ortak.ad)
                    return Eslesme(hareket, kod, f"335 PERSONEL - {ortak.ad}", "AKILLI", 90,
                                  f"Ortak + maaş/huzur hakkı → 335 personel")
                # Normal ortak hareketi → 131/331
                kod = self._ortak_hesap_kodu(ortak, negatif=hareket.negatif)
                tip = "131 ORTAKTAN ALACAK" if hareket.negatif else "331 ORTAĞA BORÇ"
                return Eslesme(hareket, kod, f"{tip} - {ortak.ad}", "ORTAK", 80,
                              f"Ortak hareketi → {self.ortak_hesap}")

        # İHTİYAÇ / PAYLAŞIM / YARDIM + isim → 131/331
        if any(kw in n_acik for kw in ("IHTIYAC", "PAYLASIM", "YARDIM")):
            kod = self._ortak_hesap_kodu(negatif=hareket.negatif)
            return Eslesme(hareket, kod, f"{self.ortak_hesap} ORTAK (İHTİYAÇ/PAYLAŞIM)", "AKILLI", 75,
                          "İhtiyaç/paylaşım/yardım → 131/331 ortak")

        # ── 3.7) FAST ÜCRET İADESİ (küçük tutar) → 780 masraf ──
        if "UCRET IADESI" in n_acik or "ÜCRET İADESİ" in n_acik:
            if abs(hareket.tutar) < 100:
                kod = self._masraf_hesap_bul()
                return Eslesme(hareket, kod, "MASRAF İADESİ", "AKILLI", 85,
                              "Ücret iadesi (küçük tutar) → 780 masraf")

        # ── 4) Cari eşleşme — token eşleştir ──
        hareket_tokenlari = _anlamli_tokenlar(hareket.aciklama)

        guclu = []
        zayif = []
        for c in self.tablolar.cariler:
            if not c.aktif:
                continue
            tip_uygun = (c.tip.startswith("120") and not hareket.negatif) or \
                        (c.tip.startswith("320") and hareket.negatif)
            n_ad = normalize(c.ad)

            if n_ad and len(n_ad) >= 5 and n_ad in n_acik:
                guclu.append((95 + (5 if tip_uygun else 0), c))
                continue
            if c.anahtarlar and any(kw and kw in n_acik for kw in c.anahtarlar):
                guclu.append((90 + (5 if tip_uygun else 0), c))
                continue
            cari_tokenlari = _anlamli_tokenlar(c.ad)
            if not (cari_tokenlari and hareket_tokenlari):
                continue
            ortak = cari_tokenlari & hareket_tokenlari
            if not ortak:
                continue
            oran = len(ortak) / len(cari_tokenlari)
            if len(ortak) >= 2:
                skor = 80 if oran >= 0.6 else 72
                guclu.append((skor + (5 if tip_uygun else 0), c))
            else:
                tek = next(iter(ortak))
                if len(tek) >= 6 and self.token_nadirlik.get(tek, 0) == 1:
                    guclu.append((72 + (5 if tip_uygun else 0), c))
                elif len(tek) >= 5 and oran >= 0.5:
                    zayif.append((40 + (5 if tip_uygun else 0), c))

        if guclu:
            guclu.sort(key=lambda x: -x[0])
            skor, c = guclu[0]
            ayni_skor = [a for s, a in guclu if s == skor]
            oneri = [(a.hesap_kodu, a.ad) for _, a in guclu[1:4]]
            if len(ayni_skor) > 1:
                return Eslesme(hareket, c.hesap_kodu, c.ad, "CARI", min(skor - 25, 55),
                               f"Cari ({c.tip}) — {len(ayni_skor)} eşit aday, KONTROL ET",
                               [(a.hesap_kodu, a.ad) for a in ayni_skor[1:4]] + oneri)
            return Eslesme(hareket, c.hesap_kodu, c.ad, "CARI", min(skor, 100), f"Cari eşleşti: {c.tip}", oneri)

        # ── 5) Eşleşmeyen → 999 GEÇİCİ HESAP ──
        # Eski davranış: pozitif → 120.02 idi. YANLIŞ! Artık her şey 999.
        if zayif:
            zayif.sort(key=lambda x: -x[0])
            _, c = zayif[0]
            oneri = [(a.hesap_kodu, a.ad) for _, a in zayif[:3]]
            return Eslesme(hareket, "999", "GEÇİCİ HESAP", "GEÇİCİ", 30,
                           f"Eşleşmedi → 999 (olası: {c.hesap_kodu} {c.ad[:25]}, KONTROL ET)", oneri)

        return Eslesme(hareket, "999", "GEÇİCİ HESAP", "GEÇİCİ", 30,
                       "Eşleşmedi → 999 geçici hesap (KONTROL ET)")

    # --------------------------------------------------------------------- #
    def toplu(self, hareketler: Iterable[Hareket]) -> list[Eslesme]:
        sonuc = [self.esle(h) for h in hareketler]
        for e in sonuc:
            self._kod_safety(e)
        return sonuc


# ============================================================================= #
# Bakiye kontrolü
# ============================================================================= #
def bakiye_kontrol(hareketler: list[Hareket], acilis: float, kapanis: float, tolerans: float = 0.05) -> dict:
    toplam = sum(h.tutar for h in hareketler)
    hesaplanan_kapanis = acilis + toplam
    fark = kapanis - hesaplanan_kapanis
    return {
        "acilis": acilis,
        "kapanis_beklenen": kapanis,
        "hareket_toplami": toplam,
        "hesaplanan_kapanis": hesaplanan_kapanis,
        "fark": fark,
        "tamam": abs(fark) <= tolerans,
    }
