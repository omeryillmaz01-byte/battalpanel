# -*- coding: utf-8 -*-
"""
Firma kurulumu — Hesap planı Excel'inden firma için referans tabloları üretir.

Üretilen dosyalar (firmalar/<FIRMA_KODU>/):
  01_banka_hesaplari.xlsx → 102.xx kodları + hesap no/IBAN
  02_pos_hesaplari.xlsx   → 108.xx kodları
  03_cari_eslesme.xlsx    → 120.xx ve 320.xx cari listesi (manuel doldurulacak ek sütun)
  04_kural_sozlugu.xlsx   → Anahtar kelime → hesap kodu (başlangıç kuralları)

Kullanım:
  python firma_kur.py "ISIK_PETROL" "C:/.../ISIK PETROL HESAP PLANI.xlsx"
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent.parent
FIRMA_DIZIN = BASE / "firmalar"

# Görsel sabitler
BASLIK_DOLGU = PatternFill("solid", fgColor="305496")
BASLIK_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_BASLIK_DOLGU = PatternFill("solid", fgColor="FFD966")
KENAR = Border(*(Side(style="thin", color="BFBFBF"),) * 4)
ORTA = Alignment(horizontal="center", vertical="center", wrap_text=True)
SOL = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _hesap_planini_oku(yol: Path) -> list[tuple[str, str, float | None]]:
    wb = openpyxl.load_workbook(yol, data_only=True)
    ws = wb.active
    satirlar: list[tuple[str, str, float | None]] = []
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        kod = str(row[0]).strip()
        ad = str(row[1]).strip() if row[1] is not None else ""
        bakiye = row[2] if len(row) > 2 else None
        if not kod or kod == "HESAP KODU":
            continue
        satirlar.append((kod, ad, bakiye))
    return satirlar


def _hesap_no_cikar(ad: str) -> str:
    """'( 28620 )' veya '( 0271 - 028620 )' gibi parantez içeriklerini çıkarır."""
    m = re.search(r"\(([^)]+)\)", ad)
    return m.group(1).strip() if m else ""


def _temizle_banka_adi(ad: str) -> str:
    """'AKBANK VADESİZ TL HS. ( 28620 )' → 'AKBANK VADESİZ TL HS.'"""
    return re.sub(r"\s*\([^)]*\)\s*$", "", ad).strip()


def _gercek_pos_mu(ad: str) -> bool:
    """GERÇEK POS terminal hesabı mı? 'Posta Pulları' (POSTA) yanlış eşleşmesini engeller."""
    a = ad.upper()
    if "POSTA" in a:        # Posta Pulları, Posta Çeki vb. → POS DEĞİL
        return False
    return ("POS HS" in a or "POS HESAB" in a or "KART POS" in a or
            " POS " in f" {a} " or a.endswith(" POS") or "POS TERMİNAL" in a or "POS TERMINAL" in a)


def _stil_basliklar(ws, satir: int, kolon_basliklar: list[str]) -> None:
    for j, baslik in enumerate(kolon_basliklar, 1):
        h = ws.cell(row=satir, column=j, value=baslik)
        h.fill = BASLIK_DOLGU
        h.font = BASLIK_FONT
        h.alignment = ORTA
        h.border = KENAR
    ws.row_dimensions[satir].height = 30


def _otomatik_genislik(ws, max_kolon: int) -> None:
    for j in range(1, max_kolon + 1):
        max_uzunluk = 8
        for row in ws.iter_rows(min_col=j, max_col=j, values_only=True):
            v = row[0]
            if v is None:
                continue
            l = len(str(v))
            if l > max_uzunluk:
                max_uzunluk = l
        ws.column_dimensions[get_column_letter(j)].width = min(max_uzunluk + 3, 60)


# ----------------------------------------------------------------------------- #
# 01 — BANKA HESAPLARI
# ----------------------------------------------------------------------------- #
def banka_hesaplari_uret(satirlar, cikti: Path) -> int:
    """102.xx altındaki yaprak hesapları (hesap no'su olanları) çıkarır."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Banka Hesaplari"

    aciklama = [
        "BU TABLO: Banka ekstresi yüklendiğinde sistemin doğru 102.xx koduna eşleştirmesi için kullanılır.",
        "→ 'Banka Hesap No' kolonu hesap planından otomatik dolduruldu.",
        "→ 'IBAN' kolonunu sen dolduracaksın (boş kalabilir, ama dolu olursa eşleştirme %100 kesinlikle çalışır).",
        "→ 'Anahtar Kelimeler' kolonu CSV açıklamasında geçen alternatif isimleri yaz (ör. 'AKBNK', 'AK BANK', 'AKBANK').",
    ]
    for i, t in enumerate(aciklama, 1):
        c = ws.cell(row=i, column=1, value=t)
        c.font = Font(italic=True, color="595959", size=10)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)

    basliklar = ["Hesap Kodu", "Banka", "Hesap Adı", "Hesap No", "IBAN", "Anahtar Kelimeler"]
    _stil_basliklar(ws, len(aciklama) + 2, basliklar)

    satir = len(aciklama) + 3
    sayac = 0
    for kod, ad, bakiye in satirlar:
        # 102.xx — sadece banka adı/hesap no taşıyan satırlar (yaprak)
        if not kod.startswith("102."):
            continue
        # parantezli ya da bakiyesi olan = gerçek hesap; saf 'AK BANK' gibi alt başlıkları atla
        hesap_no = _hesap_no_cikar(ad)
        if not hesap_no and bakiye is None:
            continue
        if not hesap_no:
            # Bakiyesi var ama parantez yok → yine de ekleyelim
            hesap_no = ""
        banka_adi = _temizle_banka_adi(ad)
        # banka kısa adını çıkar (ilk kelime grubu)
        m = re.match(r"([A-ZÇĞİÖŞÜ ]+?)(?:\s+VADESİZ|\s+VADELİ|\s+FON|\s+POS|\s*$)", banka_adi)
        banka_kisa = m.group(1).strip() if m else banka_adi.split()[0]

        ws.cell(row=satir, column=1, value=kod).alignment = ORTA
        ws.cell(row=satir, column=2, value=banka_kisa).alignment = SOL
        ws.cell(row=satir, column=3, value=banka_adi).alignment = SOL
        ws.cell(row=satir, column=4, value=hesap_no).alignment = SOL
        ws.cell(row=satir, column=5, value="").alignment = SOL  # IBAN
        ws.cell(row=satir, column=6, value="").alignment = SOL  # Anahtar kelimeler
        for j in range(1, 7):
            ws.cell(row=satir, column=j).border = KENAR
        satir += 1
        sayac += 1

    _otomatik_genislik(ws, 6)
    ws.freeze_panes = f"A{len(aciklama) + 3}"
    wb.save(cikti)
    return sayac


# ----------------------------------------------------------------------------- #
# 02 — POS HESAPLARI
# ----------------------------------------------------------------------------- #
def pos_hesaplari_uret(satirlar, cikti: Path) -> int:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "POS Hesaplari"

    aciklama = [
        "BU TABLO: POS tahsilatlarını doğru 108.xx koduna eşleştirmek için kullanılır.",
        "→ 'Anahtar Kelimeler': CSV'de POS hareketi tanımlanırken geçen ifadeler (ör. 'POS', 'KREDI KART', 'AKBANK POS').",
    ]
    for i, t in enumerate(aciklama, 1):
        c = ws.cell(row=i, column=1, value=t)
        c.font = Font(italic=True, color="595959", size=10)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)

    basliklar = ["Hesap Kodu", "POS Adı", "Banka", "Anahtar Kelimeler"]
    _stil_basliklar(ws, len(aciklama) + 2, basliklar)

    satir = len(aciklama) + 3
    sayac = 0
    for kod, ad, bakiye in satirlar:
        if not kod.startswith("108."):
            continue
        if not _gercek_pos_mu(ad):     # 'Posta Pulları' gibi sahte POS'ları atla
            continue
        banka = ad.replace("POS HS.", "").replace("POS", "").strip()
        ws.cell(row=satir, column=1, value=kod).alignment = ORTA
        ws.cell(row=satir, column=2, value=ad).alignment = SOL
        ws.cell(row=satir, column=3, value=banka).alignment = SOL
        ws.cell(row=satir, column=4, value="").alignment = SOL
        for j in range(1, 5):
            ws.cell(row=satir, column=j).border = KENAR
        satir += 1
        sayac += 1

    _otomatik_genislik(ws, 4)
    ws.freeze_panes = f"A{len(aciklama) + 3}"
    wb.save(cikti)
    return sayac


# ----------------------------------------------------------------------------- #
# 03 — CARİ EŞLEŞME (120 + 320)
# ----------------------------------------------------------------------------- #
def cari_eslesme_uret(satirlar, cikti: Path) -> int:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cari Eslesme"

    aciklama = [
        "BU TABLO: Banka açıklamasındaki cari adı 120.xx (alıcı) ya da 320.xx (satıcı) koduna eşleştirilir.",
        "→ Aynı cari hem alıcı hem satıcı olabilir; sistem önce 'Hareket Tipi' ipucuyla doğru olanı seçer.",
        "→ 'Anahtar Kelimeler': CSV açıklamasında geçebilecek alternatif yazımlar (LTD ŞTİ, A.Ş., kısaltma vs.).",
    ]
    for i, t in enumerate(aciklama, 1):
        c = ws.cell(row=i, column=1, value=t)
        c.font = Font(italic=True, color="595959", size=10)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)

    basliklar = ["Hesap Kodu", "Tip (120/320)", "Cari Adı", "Anahtar Kelimeler", "Aktif (E/H)"]
    _stil_basliklar(ws, len(aciklama) + 2, basliklar)

    satir = len(aciklama) + 3
    sayac = 0
    for kod, ad, bakiye in satirlar:
        if not (kod.startswith("120.") or kod.startswith("320.")):
            continue
        # Cari YAPRAK hesapları — hesap planında İKİ kod yapısı var:
        #   (a) 4 kademeli: 120.01.001.0001  (örn. AYGAZ)
        #   (b) 2 kademeli kısa kod: 320.05, 320.200, 320.573  (gerçek aktif tedarikçilerin çoğu burada!)
        nokta = kod.count(".")
        dort_kademe_leaf = nokta >= 3
        iki_kademe_leaf = bool(re.match(r"^(120|320)\.\d{1,3}$", kod))
        if not (dort_kademe_leaf or iki_kademe_leaf):
            continue  # ara başlıkları (120.01.001 grup başlığı vb.) atla
        # Yapısal başlıkları atla (örn. 'MERKEZ İŞYERİ ALICILAR HESABI ( A )', '120.01')
        if "ALICILAR HESABI" in ad or "SATICILAR HESABI" in ad or "İŞYERİ" in ad:
            continue
        if not ad:
            continue
        tip = "120 ALICI" if kod.startswith("120.") else "320 SATICI"
        ws.cell(row=satir, column=1, value=kod).alignment = ORTA
        ws.cell(row=satir, column=2, value=tip).alignment = ORTA
        ws.cell(row=satir, column=3, value=ad).alignment = SOL
        ws.cell(row=satir, column=4, value="").alignment = SOL
        ws.cell(row=satir, column=5, value="E").alignment = ORTA
        for j in range(1, 6):
            ws.cell(row=satir, column=j).border = KENAR
        satir += 1
        sayac += 1

    _otomatik_genislik(ws, 5)
    ws.freeze_panes = f"A{len(aciklama) + 3}"
    wb.save(cikti)
    return sayac


# ----------------------------------------------------------------------------- #
# 05 — ORTAKLAR (131 Ortaklardan Alacaklar / 331 Ortaklara Borçlar)
# ----------------------------------------------------------------------------- #
def ortaklar_uret(satirlar, cikti: Path) -> int:
    """131.xx ve 331.xx hesaplarından ortak adı → (131 kodu, 331 kodu) tablosu."""
    alacak = {}   # normalize ad → (kod, gerçek ad)
    borc = {}
    import unicodedata
    def nrm(s):
        s = str(s).upper().translate(str.maketrans("İIıŞşĞğÜüÖöÇç", "IIISSGGUUOOCC"))
        return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode().strip()
    for kod, ad, _ in satirlar:
        k = str(kod)
        if not ad or "ORTAK" in str(ad).upper() and "HESAB" in str(ad).upper():
            continue
        if k.startswith("131.") and k.count(".") >= 1 and "ALACAK" not in str(ad).upper():
            alacak[nrm(ad)] = (k, str(ad).strip())
        elif k.startswith("331.") and k.count(".") >= 1 and "BORÇ" not in str(ad).upper() and "BORC" not in str(ad).upper():
            borc[nrm(ad)] = (k, str(ad).strip())

    adlar = set(alacak) | set(borc)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ortaklar"
    aciklama = [
        "BU TABLO: Ortak/yetkili para hareketleri → 131 (çıkış/alacak) ya da 331 (giriş/borç).",
        "→ Ortak adına para ÇIKARSA (-) → 131; para GİRERSE (+) → 331.",
        "→ 'Anahtar Kelimeler': açıklamada geçebilecek alternatif yazımlar.",
    ]
    for i, t in enumerate(aciklama, 1):
        c = ws.cell(row=i, column=1, value=t); c.font = Font(italic=True, color="595959", size=10)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
    _stil_basliklar(ws, len(aciklama) + 2, ["Ortak Adı", "131 Kodu (çıkış)", "331 Kodu (giriş)", "Anahtar Kelimeler"])
    satir = len(aciklama) + 3
    for nad in sorted(adlar):
        ad131, ad331 = alacak.get(nad), borc.get(nad)
        gercek = (ad131 or ad331)[1]
        ws.cell(row=satir, column=1, value=gercek).alignment = SOL
        ws.cell(row=satir, column=2, value=ad131[0] if ad131 else "").alignment = ORTA
        ws.cell(row=satir, column=3, value=ad331[0] if ad331 else "").alignment = ORTA
        ws.cell(row=satir, column=4, value="").alignment = SOL
        for j in range(1, 5):
            ws.cell(row=satir, column=j).border = KENAR
        satir += 1
    _otomatik_genislik(ws, 4)
    wb.save(cikti)
    return len(adlar)


# ----------------------------------------------------------------------------- #
# 04 — KURAL SÖZLÜĞÜ (anahtar kelime → hesap kodu)
# ----------------------------------------------------------------------------- #
def kural_sozlugu_uret(cikti: Path, pos_kodu: str = "108.??",
                       pos_ad: str = "POS TAHSİLATI (banka adına göre)") -> int:
    """Başlangıç kural seti — kullanıcı sonradan ekleyebilir.
    pos_kodu: POS tahsilatının yazılacağı hesap. IŞIK→'108.??' (akaryakıt), diğer firmalar→'127' vb."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kurallar"

    aciklama = [
        "BU TABLO: Banka açıklamasında geçen anahtar kelimelere göre OTOMATİK hesap kodu önerir.",
        "→ Sıra önemlidir: küçük sıra önce denenir. Daha ÖZEL kuralları üste (küçük sıraya) koy.",
        "→ 'Anahtar Kelimeler' aralarına | (boru) koyarak alternatif yazımları ekle: 'KOMISYON|KOMİSYON'.",
        "→ 'Tutar İşareti': NEGATIF = sadece gider (-) satırda eşleş. POZITIF = sadece (+). HER İKİSİ = sınır yok.",
        "→ 'Güven' 0-100: 100=kesin (yeşil), 60-79=kontrol önerilir (sarı), <60=mutlaka incele.",
        "→ 'Not': çıktıda o satırın yanında görünür; mali müşavir için uyarı/açıklama yaz.",
        "→ Hesap Kodu BOŞ bırakılırsa: kural eşleşir ama kod atanmaz (manuel doldur) — sadece NOT gösterilir.",
        "→ Finansman giderleri (komisyon/BSMV/masraf/EFT/POS ücreti): 780.?? = firmanın hesap planındaki uygun 780 alt hesabı.",
        "→ '108.??' özel kod: POS hakedişi — işlenen bankanın POS hesabına (108.xx) otomatik yazılır.",
    ]
    for i, t in enumerate(aciklama, 1):
        c = ws.cell(row=i, column=1, value=t)
        c.font = Font(italic=True, color="595959", size=10)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)

    basliklar = ["Sıra", "Anahtar Kelimeler", "Hesap Kodu", "Hesap Adı", "Tutar İşareti", "Güven", "Not", "Aktif (E/H)"]
    _stil_basliklar(ws, len(aciklama) + 2, basliklar)

    # (kelime, kod, hesap_adı, işaret, güven, not)
    baslangic = [
        # --- Finansman giderleri → 780.?? (hesap planından dinamik) ---
        ("POS HİZ|HİZ.ÜC|HIZ.UC|POS HİZMET|POS HIZMET|POS MASRAF|POS CİHAZI|POS UCRET|ÖKC|OKC|POS AIDAT|POS AİDAT",
            "780.??", "FİNANSMAN GİDERİ", "NEGATIF", 100, "POS hizmet/cihaz/aidat ücreti"),
        ("ÜYE İŞYERİ ÜCRET|UYE ISYERI UCRET|KATKI PAYI|KATKIPAYI|PUAN KOM|E P PUAN",
            "780.??", "FİNANSMAN GİDERİ", "NEGATIF", 100, "POS üye işyeri ücreti / katkı payı"),
        ("KREDI TAHSIS|KREDİ TAHSİS|KULLANDIRIM UCRET|KULLANDIRIM ÜCRET|DOSYA MASRAF|TAHSIS UCRET",
            "780.??", "FİNANSMAN GİDERİ", "NEGATIF", 100, "Kredi tahsis/kullandırım/dosya masrafı"),
        ("TEMINAT MEKTUBU|TEMİNAT MEKTUBU|TEM.MEK|TEM MEK|MEKTUBU KOM|MEKTUBU TEMDIT|MEKTUBU TEMDİT",
            "780.??", "FİNANSMAN GİDERİ", "NEGATIF", 100, "Teminat mektubu komisyonu"),
        ("REHİN ÜCRET|REHIN UCRET|EKSPERTİZ|EKSPERTIZ",
            "780.??", "FİNANSMAN GİDERİ", "NEGATIF", 100, "Rehin/ekspertiz ücreti"),
        ("GAYRİNAKDİ|GAYRINAKDI|NAKDI KRD|NAKDİ KRD|DÖNEM ÜCRETİ|DONEM UCRETI|G.NAKDI",
            "780.??", "FİNANSMAN GİDERİ", "NEGATIF", 100, "Gayrinakdi kredi dönem ücreti"),
        ("PES TOPLCP|AKPOS PES TOPL|POS DUZELTME|POS DÜZELTME",
            pos_kodu, "POS TAHSİLATI DÜZELTME", "HER İKİSİ", 90, "POS hakediş düzeltme"),
        ("KOMISYON|KOMİSYON",              "780.??", "FİNANSMAN GİDERİ", "NEGATIF", 100, ""),
        ("BSMV",                           "780.??", "FİNANSMAN GİDERİ", "NEGATIF", 100, ""),
        ("EFT|MBL-EFT|HAVALE ÜCRETİ|HAVALE MASRAFI|MOBİL EFT",
            "780.??", "FİNANSMAN GİDERİ", "NEGATIF", 100, ""),
        ("MASRAF|İŞLEM ÜCRETİ|ISLEM UCRETI|HİZMET BEDELİ",
            "780.??", "FİNANSMAN GİDERİ", "NEGATIF", 100, ""),
        ("HESAP İŞLETİM|HESAP ISLETIM|EKSTRE ÜCRETİ|HESAP BAKIM|İŞLETİM ÜCRETİ|MESAJ ÜCRETİ|MESAJ UCRETI|SMS ÜCRET",
            "780.??", "FİNANSMAN GİDERİ", "NEGATIF", 100, ""),
        ("KART AİDAT|KART ÜCRET",          "780.??", "FİNANSMAN GİDERİ", "NEGATIF", 100, ""),
        ("KREDİ FAİZ|KREDI FAIZ|KREDİ FAİZİ|FAİZ TAHSİL|FAİZ TAHAKKUK",
            "780.??", "FİNANSMAN GİDERİ", "NEGATIF", 100, "Kredi faizi → 780"),
        ("FAİZ|FAIZ",                      "780.??", "FİNANSMAN GİDERİ", "NEGATIF", 90,  "Faiz gideri (pozitifse 642 faiz geliri olabilir)"),
        ("STOPAJ",                         "193.??", "STOPAJ", "NEGATIF", 80,  "Banka faiz stopajı — 193 alt hesabı"),
        # --- POS tahsilatı → 108.xx (kesin) ---
        ("AKPOS|POS ODE|POS PES|PES ODE|POS ÖDEME|FAST POS ODEME|POS HAKEDİŞ|POS HAKEDİS|POS ALACAK|ÜYE İŞYERİ|UYELIK ISYERI|PEŞİNSATIŞ|PESINSATIS|PEŞİN SATIŞ|POS TAHSİL|İŞYERİ NO|ISYERI NO|POS SATIŞ|POS SATIS",
            pos_kodu, pos_ad, "POZITIF", 100, ""),
        # --- Kasa ↔ banka ---
        ("PARA YATIRMA|NAKİT YATIRMA|NAKIT YATIRMA",
            "100.02", "GÜNLÜK NAKİT TAHSİLATLAR", "POZITIF", 100,
            "Kasaya yatan günlük nakit (100.02)"),
        # --- Personel (yargı gerektirir → sarı) ---
        ("MAAS AVANS|MAAŞ AVANS|MASŞ AVANS|MASS AVANS|AVANS ÖDEME|PERSONEL AVANS|AVANS",
            "335.??", "PERSONEL NET MAAŞ ÖDEMESİ", "NEGATIF", 100,
            "Personel maaş avansı (335)"),
        ("MAAS|MAAŞ|MASŞ|MMAŞ|MMAS|UCRET ODEME|ÜCRET ÖDEME",
            "335.??", "PERSONEL NET MAAŞ ÖDEMESİ", "NEGATIF", 100,
            "Personel net maaş ödemesi (335)"),
        # --- SGK / Vergi ---
        ("SGK|PRİM ÖDEME|SOSYAL GUVENLIK",
            "361.02", "SGK İŞVEREN PAYI", "NEGATIF", 65, "SGK prim ödemesi — teyit et (361.01/361.02/361.03)"),
        ("KDV ODEME|KDV ÖDEME|MUHTASAR|VERGİ ÖDEME|VERGI ODEME|GELIR VERGISI|VERGİ TRANSFER|VERGI TRANSFER|VERGİ TAHSİL|GİB|G.İ.B|HAZİNE VE MALİYE",
            "360.07", "VERGİ ÖDEMESİ", "NEGATIF", 60, "Vergi ödemesi — doğru 360.xx alt hesabı seç"),
        # --- Kullanıcı talimatı özel hesaplar ---
        ("ANA PARA TAHSILAT|ANA PARA TAHSİLAT|ÖDEME PLANI ANA PARA|ODEME PLANI ANA PARA|ANAPARA TAHSIL",
            "999", "GEÇİCİ HESAP", "HER İKİSİ", 90, "Kredi anapara — şimdilik 999 geçici hesap"),
        ("HGS",
            "760.02", "ARAÇ GİDERLERİ", "NEGATIF", 95, "HGS geçiş/yükleme → araç gideri"),
        ("FON ALIMI|FON SATIMI|FON SATIŞ|FON SATIS|FON ALIM",
            "118.01.001.0003", "AKBANK FON HESABI", "HER İKİSİ", 90, "Yatırım fonu alım/satım → 118 (Akbank fon)"),
        ("DBSODMSHELL|DBS ODM SHELL|DBSODMS|DBS SHELL",
            "320.01.021.0000", "SHELL&TURCAS PETROL A.Ş.", "NEGATIF", 95, "Shell DBS ödemesi → Shell cari"),
        # --- Kısa/özel ünvanlı cariler (token eşleşmesi kısa kaldığı için kuralla) ---
        ("MYT PETROL|MYT ",
            "320.490", "MYT PETROL ÜRÜNLERİ SAN. VE TİC. LTD.ŞTİ", "NEGATIF", 90, "Myt Petrol (320.490)"),
        # --- Kodsuz (sadece uyarı — manuel doldurulacak) ---
        ("HACIZ|İCRA|ICRA",
            "335.??", "PERSONEL NET MAAŞ ÖDEMESİ", "NEGATIF", 100, "Maaş haczi (335)"),
        ("KREDİ KARTI ODEME|KREDİ KARTI ÖDEME|KREDI KARTI",
            "329.02", "AKBANK 5526...1755 KK", "NEGATIF", 100,
            "Akbank KK ödemesi (329.02)"),
        ("TRCELL|TURKCELL",
            "320.01.023.0007", "Turkcell Iletisim Hizmetleri A.S.", "NEGATIF", 70,
            "Turkcell telefon ödemesi — cari üzerinden (alternatif: 770 haberleşme gideri)"),
        ("TÜRK TELEKOM|TURK TELEKOM|TTNET|TTLKOM|TT MOBIL|TTMOBIL",
            "320.01.023.0003", "TÜRK TELEKOM A.Ş.", "NEGATIF", 70, "Türk Telekom ödemesi"),
        ("MKK ÜCRET|MKK UCRET|MERKEZİ KAYIT",
            "780.??", "FİNANSMAN GİDERİ", "NEGATIF", 90, "MKK (Merkezi Kayıt) ücreti"),
        ("KESİNTİ VE EKLER|KESINTI VE EKLER|KESİNTİ|KESINTI",
            "780.??", "FİNANSMAN GİDERİ", "NEGATIF", 90, "Banka kesintisi → finansman gideri"),
        ("ÇEKTKSZ|ÇEK TAKAS|CEK TAKAS|VERİLEN ÇEK|VERILEN CEK|ÇEK ÖDEME|ÇEK ÇIKIŞ|ÇEK İADE",
            "103.??", "VERİLEN ÇEKLER (banka adına göre)", "NEGATIF", 90, "Çek ÖDEMESİ → o bankanın 103 alt hesabı"),
        ("ALINAN ÇEK|ALINAN CEK|MÜŞTERİ ÇEK|MUSTERI CEK|ÇEK TAHSİL|CEK TAHSIL|ÇEK GİRİŞ|GELEN ÇEK|ÇEK BORDRO",
            "101.??", "ALINAN ÇEKLER (banka adına göre)", "POZITIF", 90, "Çek TAHSİLATI → o bankanın 101 alt hesabı"),
        ("VODAFONE|TELEFON|İNTERNET|HABERLESME",
            "", "", "NEGATIF", 25, "Haberleşme gideri (770.01.003.0006) — kontrol et"),
    ]

    satir = len(aciklama) + 3
    for i, (kelime, kod, hesap_ad, isaret, guven, not_) in enumerate(baslangic, 1):
        ws.cell(row=satir, column=1, value=i).alignment = ORTA
        ws.cell(row=satir, column=2, value=kelime).alignment = SOL
        ws.cell(row=satir, column=3, value=kod).alignment = ORTA
        ws.cell(row=satir, column=4, value=hesap_ad).alignment = SOL
        ws.cell(row=satir, column=5, value=isaret).alignment = ORTA
        ws.cell(row=satir, column=6, value=guven).alignment = ORTA
        ws.cell(row=satir, column=7, value=not_).alignment = SOL
        ws.cell(row=satir, column=8, value="E").alignment = ORTA
        for j in range(1, 9):
            ws.cell(row=satir, column=j).border = KENAR
        satir += 1

    _otomatik_genislik(ws, 8)
    ws.freeze_panes = f"A{len(aciklama) + 3}"
    wb.save(cikti)
    return len(baslangic)


# ----------------------------------------------------------------------------- #
# Ana fonksiyon
# ----------------------------------------------------------------------------- #
def tam_hesap_plani_kaydet(satirlar, cikti: Path) -> int:
    """TAM hesap planını (tüm kod+ad+bakiye) kalıcı kaydeder — 06_hesap_plani.xlsx.
    Motor 780/335/770/760/193/642/679 vb. alt hesabını buradan okur, ezbere değil."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hesap Plani"
    _stil_basliklar(ws, 1, ["Hesap Kodu", "Hesap Adı", "Bakiye"])
    satir = 2
    for kod, ad, bakiye in satirlar:
        if not kod:
            continue
        ws.cell(row=satir, column=1, value=str(kod)).alignment = ORTA
        ws.cell(row=satir, column=2, value=str(ad)).alignment = SOL
        ws.cell(row=satir, column=3, value=bakiye).alignment = ORTA
        for j in range(1, 4):
            ws.cell(row=satir, column=j).border = KENAR
        satir += 1
    _otomatik_genislik(ws, 3)
    ws.freeze_panes = "A2"
    wb.save(cikti)
    return satir - 2


def firma_kur(firma_kodu: str, hesap_plani_yolu: Path) -> dict:
    firma_dizini = FIRMA_DIZIN / firma_kodu
    firma_dizini.mkdir(parents=True, exist_ok=True)

    satirlar = _hesap_planini_oku(hesap_plani_yolu)

    # POS tahsilatı hesabı tespiti: GERÇEK 108 POS hesabı varsa 108.?? (IŞIK/akaryakıt), yoksa 127.
    # ('Posta Pulları' POS sayılmaz — _gercek_pos_mu POSTA'yı eler.)
    pos_108_var = any(str(k).startswith("108.") and _gercek_pos_mu(str(a)) for k, a, _ in satirlar)
    if pos_108_var:
        pos_kodu, pos_ad = "108.??", "POS TAHSİLATI (banka adına göre)"
    else:
        # 127 yaprak hesabını bul (yoksa parent 127)
        k127 = next((k for k, a, _ in satirlar if str(k).startswith("127.") and str(k).count(".") >= 1), "127")
        pos_kodu, pos_ad = k127, "DİĞER TİCARİ ALACAKLAR (POS)"

    sonuc = {}
    sonuc["banka"] = banka_hesaplari_uret(satirlar, firma_dizini / "01_banka_hesaplari.xlsx")
    sonuc["pos"] = pos_hesaplari_uret(satirlar, firma_dizini / "02_pos_hesaplari.xlsx")
    sonuc["cari"] = cari_eslesme_uret(satirlar, firma_dizini / "03_cari_eslesme.xlsx")
    sonuc["kural"] = kural_sozlugu_uret(firma_dizini / "04_kural_sozlugu.xlsx", pos_kodu, pos_ad)
    sonuc["ortak"] = ortaklar_uret(satirlar, firma_dizini / "05_ortaklar.xlsx")
    # 06 — TAM HESAP PLANI (KALICI): tüm kodları sakla ki bir daha kaynak gerekmesin.
    # Motor 780/335/770/760/193/642/679 alt hesabını buradan okur.
    sonuc["plan"] = tam_hesap_plani_kaydet(satirlar, firma_dizini / "06_hesap_plani.xlsx")
    sonuc["pos_hesabi"] = pos_kodu
    return sonuc


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Kullanım: python firma_kur.py <FIRMA_KODU> <HESAP_PLANI_XLSX>")
        sys.exit(1)
    kod = sys.argv[1]
    yol = Path(sys.argv[2])
    s = firma_kur(kod, yol)
    print(f"[OK] {kod} kuruldu.")
    print(f"     Banka  : {s['banka']} kayıt")
    print(f"     POS    : {s['pos']} kayıt")
    print(f"     Cari   : {s['cari']} kayıt")
    print(f"     Kural  : {s['kural']} kayıt")
