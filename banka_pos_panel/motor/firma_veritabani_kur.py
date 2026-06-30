# -*- coding: utf-8 -*-
"""
Firma veritabanı kurulumu — 'banka excel.xlsx'teki tüm firmaları panele ekler.

Her firma için:
  firmalar/<FIRMA_KODU>/01_banka_hesaplari.xlsx  (banka + IBAN; hesap kodu BOŞ, hesap planı gelince dolar)
  firmalar/<FIRMA_KODU>/02_pos_hesaplari.xlsx     (boş - hesap planı gelince)
  firmalar/<FIRMA_KODU>/03_cari_eslesme.xlsx      (boş - hesap planı gelince)
  firmalar/<FIRMA_KODU>/04_kural_sozlugu.xlsx     (varsayılan kural seti)
Ayrıca: firma_veritabani.xlsx (tüm firma + IBAN ana listesi).
"""
from __future__ import annotations
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .firma_kur import (FIRMA_DIZIN, kural_sozlugu_uret, _stil_basliklar,
                        _otomatik_genislik, BASLIK_DOLGU, BASLIK_FONT, KENAR, ORTA, SOL)

BASE = Path(__file__).resolve().parent.parent


def _kod(ad: str) -> str:
    """'AKÇA AĞIZ' → 'AKCA_AGIZ' (klasör adı)."""
    t = ad.upper().translate(str.maketrans("İIŞĞÜÖÇ", "IISGUOC"))
    t = re.sub(r"[^A-Z0-9]+", "_", t).strip("_")
    return t


def _banka_kisa(banka_adi: str) -> str:
    m = re.match(r"([A-ZÇĞİÖŞÜ]+)", banka_adi.upper())
    return m.group(1) if m else banka_adi.split()[0]


def _firmalari_oku(banka_excel: Path) -> dict[str, list[tuple[str, str]]]:
    wb = openpyxl.load_workbook(banka_excel, data_only=True)
    ws = wb["Sayfa1"]
    firma = None
    out: dict[str, list[tuple[str, str]]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        f, banka, iban = row[1], row[2], row[3]
        if f and str(f).strip():
            firma = str(f).strip()
        if firma and banka and str(banka).strip():
            out.setdefault(firma, []).append((str(banka).strip(), str(iban or "").strip().replace(" ", "")))
    return out


def _banka_hesaplari_yaz(firma_dizini: Path, hesaplar: list[tuple[str, str]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Banka Hesaplari"
    aciklama = [
        "BU FİRMANIN BANKA HESAPLARI — 'banka excel.xlsx' firma veritabanından alındı.",
        "→ 'Hesap Kodu' BOŞ: bu firmanın hesap planı (102.xx) gelince doldurulacak.",
        "→ IBAN'lar dolu; ekstre yüklenince banka IBAN'dan tanınır.",
    ]
    for i, t in enumerate(aciklama, 1):
        c = ws.cell(row=i, column=1, value=t)
        c.font = Font(italic=True, color="595959", size=10)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
    basliklar = ["Hesap Kodu", "Banka", "Hesap Adı", "Hesap No", "IBAN", "Anahtar Kelimeler"]
    _stil_basliklar(ws, len(aciklama) + 2, basliklar)
    satir = len(aciklama) + 3
    for banka_adi, iban in hesaplar:
        hesap_no = ""
        m = re.search(r"(\d{6,})", iban)
        if m:
            hesap_no = m.group(1)
        ws.cell(row=satir, column=1, value="").alignment = ORTA          # Hesap Kodu (boş)
        ws.cell(row=satir, column=2, value=_banka_kisa(banka_adi)).alignment = SOL
        ws.cell(row=satir, column=3, value=banka_adi).alignment = SOL
        ws.cell(row=satir, column=4, value=hesap_no).alignment = SOL
        ws.cell(row=satir, column=5, value=iban).alignment = SOL
        ws.cell(row=satir, column=6, value="").alignment = SOL
        for j in range(1, 7):
            ws.cell(row=satir, column=j).border = KENAR
        satir += 1
    _otomatik_genislik(ws, 6)
    wb.save(firma_dizini / "01_banka_hesaplari.xlsx")


def _bos_tablo(yol: Path, baslik: str, basliklar: list[str]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = baslik
    c = ws.cell(row=1, column=1, value="Bu firmanın hesap planı gelince doldurulacak.")
    c.font = Font(italic=True, color="595959", size=10)
    _stil_basliklar(ws, 3, basliklar)
    _otomatik_genislik(ws, len(basliklar))
    wb.save(yol)


def _veritabani_master_yaz(firmalar: dict, hedef: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Firma Veritabani"
    basliklar = ["No", "Firma", "Firma Kodu", "Banka", "IBAN"]
    _stil_basliklar(ws, 1, basliklar)
    r = 2
    no = 0
    for firma, hesaplar in sorted(firmalar.items()):
        no += 1
        for j, (banka, iban) in enumerate(hesaplar):
            ws.cell(row=r, column=1, value=no if j == 0 else "").alignment = ORTA
            ws.cell(row=r, column=2, value=firma if j == 0 else "").alignment = SOL
            ws.cell(row=r, column=3, value=_kod(firma) if j == 0 else "").alignment = SOL
            ws.cell(row=r, column=4, value=banka).alignment = SOL
            ws.cell(row=r, column=5, value=iban).alignment = SOL
            for k in range(1, 6):
                ws.cell(row=r, column=k).border = KENAR
            r += 1
    _otomatik_genislik(ws, 5)
    ws.freeze_panes = "A2"
    wb.save(hedef)


def kur(banka_excel: Path) -> dict:
    firmalar = _firmalari_oku(banka_excel)
    sonuc = {"firma": 0, "hesap": 0, "atlanan": []}
    for firma, hesaplar in firmalar.items():
        kod = _kod(firma)
        d = FIRMA_DIZIN / kod
        if (d / "03_cari_eslesme.xlsx").exists() and kod == "ISIK_PETROL":
            sonuc["atlanan"].append(firma)  # mevcut IŞIK'a dokunma
            continue
        d.mkdir(parents=True, exist_ok=True)
        _banka_hesaplari_yaz(d, hesaplar)
        if not (d / "02_pos_hesaplari.xlsx").exists():
            _bos_tablo(d / "02_pos_hesaplari.xlsx", "POS Hesaplari",
                       ["Hesap Kodu", "POS Adı", "Banka", "Anahtar Kelimeler"])
        if not (d / "03_cari_eslesme.xlsx").exists():
            _bos_tablo(d / "03_cari_eslesme.xlsx", "Cari Eslesme",
                       ["Hesap Kodu", "Tip (120/320)", "Cari Adı", "Anahtar Kelimeler", "Aktif (E/H)"])
        if not (d / "04_kural_sozlugu.xlsx").exists():
            kural_sozlugu_uret(d / "04_kural_sozlugu.xlsx")
        sonuc["firma"] += 1
        sonuc["hesap"] += len(hesaplar)
    _veritabani_master_yaz(firmalar, BASE / "firma_veritabani.xlsx")
    return sonuc


if __name__ == "__main__":
    import sys
    yol = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"C:/Users/omery/OneDrive/Desktop/banka excel.xlsx")
    s = kur(yol)
    print(f"[OK] {s['firma']} firma kuruldu, {s['hesap']} banka hesabı.")
    if s["atlanan"]:
        print(f"     Atlanan (mevcut): {s['atlanan']}")
