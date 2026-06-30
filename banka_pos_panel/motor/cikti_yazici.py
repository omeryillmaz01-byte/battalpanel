# -*- coding: utf-8 -*-
"""
Mikro çıktı Excel'i üretir.

Tek sayfa: Tarih | Açıklama | Tutar | Hesap Kodu (döviz: +Kur +TL Tutar)
Başlık 1. satırda, veri 2. satırdan başlar. Üst bilgi satırı YOK.
KONTROL ve AÇILACAK CARİLER sayfaları YOK — Mikro tek sayfadan veri aktarır.
TUTAR 0.00 olan satırlar dahil edilmez.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .eslestirici import Eslesme


def tr_upper(s):
    """Türkçe imla kurallarına uygun BÜYÜK HARF: i→İ, ı→I, ş→Ş, ğ→Ğ, ü→Ü, ö→Ö, ç→Ç."""
    return (s or "").replace("i", "İ").upper()


BASLIK_DOLGU = PatternFill("solid", fgColor="305496")
BASLIK_FONT = Font(bold=True, color="FFFFFF", size=11)
MANUEL_DOLGU = PatternFill("solid", fgColor="F8CBAD")     # turuncu — manuel kontrol
DUSUK_DOLGU = PatternFill("solid", fgColor="FFE699")      # sarı — düşük güven
TAMAM_DOLGU = PatternFill("solid", fgColor="C6EFCE")      # yeşil — tam eşleşme
KENAR = Border(*(Side(style="thin", color="BFBFBF"),) * 4)
SAG = Alignment(horizontal="right", vertical="center")
ORTA = Alignment(horizontal="center", vertical="center", wrap_text=True)
SOL = Alignment(horizontal="left", vertical="center", wrap_text=True)


def mikro_excel_yaz(
    eslesmeler: list[Eslesme],
    cikti_yolu: Path,
    firma: str,
    banka: str,
    hesap_kodu: str,
    acilis: float | None = None,
    kapanis: float | None = None,
    cariler: list | None = None,
    doviz: str = "",
) -> Path:
    # TUTAR 0.00 ve "SIL" işaretli satırları filtrele
    eslesmeler = [e for e in eslesmeler
                  if abs(e.hareket.tutar) > 0.001 and e.hesap_kodu != "SIL"]

    wb = openpyxl.Workbook()
    dvz = (doviz or "").upper() in ("USD", "EUR", "GBP")

    ws = wb.active
    ws.title = "HAREKETLER"

    # Başlık 1. satırda — üst bilgi satırı YOK
    bas_satir = 1
    son_kol = 6 if dvz else 4
    if dvz:
        basliklar = ["Tarih", "Açıklama", f"Döviz Tutar ({doviz})", "Kur", "TL Tutar", "Hesap Kodu"]
    else:
        basliklar = ["Tarih", "Açıklama", "Tutar", "Hesap Kodu"]
    for j, b in enumerate(basliklar, 1):
        h = ws.cell(row=bas_satir, column=j, value=b)
        h.fill = BASLIK_DOLGU
        h.font = BASLIK_FONT
        h.alignment = ORTA
        h.border = KENAR
    ws.row_dimensions[bas_satir].height = 26

    if dvz:
        from .tcmb_kur import onceki_gun_kuru, virman_kuru

    for i, e in enumerate(eslesmeler):
        r = bas_satir + 1 + i
        tar = e.hareket.tarih
        try:
            g, a, y = tar.split("/")
            ws.cell(row=r, column=1, value=datetime(int(y), int(a), int(g))).number_format = "DD/MM/YYYY"
        except Exception:
            ws.cell(row=r, column=1, value=tar)
        ws.cell(row=r, column=1).alignment = ORTA
        ws.cell(row=r, column=2, value=tr_upper(f"{tar}- {e.hareket.aciklama}")).alignment = SOL

        if dvz:
            dt = ws.cell(row=r, column=3, value=e.hareket.tutar)
            dt.number_format = "#,##0.00;-#,##0.00;0.00"; dt.alignment = SAG
            kur = onceki_gun_kuru(doviz, tar)
            virman_mi = "VIRMAN" in e.hareket.aciklama.upper() or "VİRMAN" in e.hareket.aciklama.upper()
            kc = ws.cell(row=r, column=4, value=round(kur, 4) if kur else None)
            kc.number_format = "#,##0.0000"; kc.alignment = SAG
            tl = round(e.hareket.tutar * kur, 2) if kur else None
            tc = ws.cell(row=r, column=5, value=tl)
            tc.number_format = "#,##0.00;-#,##0.00;0.00"; tc.alignment = SAG
            hk = ws.cell(row=r, column=6, value=e.hesap_kodu)
            hk.alignment = ORTA; hk.font = Font(bold=True)
            if virman_mi:
                kc.comment = None
                ws.cell(row=r, column=4).font = Font(color="C00000")
        else:
            tt = ws.cell(row=r, column=3, value=e.hareket.tutar)
            tt.number_format = "#,##0.00;-#,##0.00;0.00"; tt.alignment = SAG
            if e.hareket.tutar < 0:
                tt.font = Font(color="C00000")
            hk = ws.cell(row=r, column=4, value=e.hesap_kodu)
            hk.alignment = ORTA; hk.font = Font(bold=True)

        if not e.hesap_kodu or e.guven == 0 or e.hesap_kodu == "999":
            renk = MANUEL_DOLGU
        elif e.guven < 80:
            renk = DUSUK_DOLGU
        else:
            renk = TAMAM_DOLGU
        for j in range(1, son_kol + 1):
            cell = ws.cell(row=r, column=j)
            cell.border = KENAR
            cell.fill = renk

    genis = [12, 60, 16, 12, 16, 18] if dvz else [12, 70, 16, 18]
    for i, w in enumerate(genis, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{bas_satir + 1}"

    # KONTROL ve AÇILACAK CARİLER sayfaları YOK — tek sayfa

    cikti_yolu = Path(cikti_yolu)
    cikti_yolu.parent.mkdir(parents=True, exist_ok=True)
    wb.save(cikti_yolu)
    return cikti_yolu
