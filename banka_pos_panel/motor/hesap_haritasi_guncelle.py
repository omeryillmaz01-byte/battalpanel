# -*- coding: utf-8 -*-
"""
IŞIK PETROL banka hesap haritası — banka_hesaplari.xlsx'e doğru hesap no + IBAN yazar.
Ekstrelerden ve hesap planından teyit edilen kesin eşleme.
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment

BASE = Path(__file__).resolve().parent.parent
FIRMA = BASE / "firmalar" / "ISIK_PETROL"

# 102.xx → (banka, tam hesap no, IBAN boşluksuz)
HARITA = {
    "102.02": ("GARANTİ",   "6299525",  "TR110006200063700006299525"),
    "102.03": ("ZİRAAT",    "815001",   ""),  # IBAN PDF'ten gelecek
    "102.04": ("YAPIKREDİ", "67103094", "TR460006701000000067103094"),
    "102.05": ("AKBANK",    "183316",   "TR210004600271888000183316"),
    "102.06": ("AKBANK",    "178445",   "TR060004600271888000178445"),
    "102.07": ("AKBANK",    "157047",   "TR200004600271888000157047"),
    "102.08": ("AKBANK",    "28620",    "TR900004600271888000028620"),
    "102.09": ("AKBANK",    "157419",   "TR640004600271888000157419"),
    "102.10": ("GARANTİ",   "6296193",  "TR560006200063700006296193"),
    "102.12": ("HALKBANK",  "10260184", "TR740001200956800010260184"),
    "102.13": ("AKBANK",    "44994",    "TR210004600271888000044994"),
}


def guncelle() -> int:
    yol = FIRMA / "01_banka_hesaplari.xlsx"
    wb = openpyxl.load_workbook(yol)
    ws = wb.active
    bas = -1
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] and "HESAP KODU" in str(row[0]).upper():
            bas = i
            break
    sayac = 0
    for r in range(bas + 1, ws.max_row + 1):
        kod = ws.cell(row=r, column=1).value
        if not kod:
            continue
        kod = str(kod).strip()
        if kod in HARITA:
            banka, no, iban = HARITA[kod]
            ws.cell(row=r, column=4, value=no).alignment = Alignment(horizontal="left")    # Hesap No
            ws.cell(row=r, column=5, value=iban).alignment = Alignment(horizontal="left")  # IBAN
            sayac += 1
    wb.save(yol)
    return sayac


if __name__ == "__main__":
    print(f"Güncellenen hesap: {guncelle()}")
