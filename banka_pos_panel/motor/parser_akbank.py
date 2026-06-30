# -*- coding: utf-8 -*-
"""
Akbank kurumsal internet şubesi → "Hesap Hareketleri" Excel parser'ı.

Akbank dosyasının yapısı:
  Satır 1 : 'Vadesiz Hesap Hareketleriniz'
  Satır 2 : Ad Soyad/Ünvan | <firma>
  Satır 3 : Şube           | <şube>
  Satır 4 : Hesap No       | 0271-0183316
  Satır 5 : Döviz Cinsi    | TL
  Satır 6 : IBAN           | TR21 0004 6002 7188 8000 1833 16
  Satır 7 : Kullanılabilir Bakiye | 477,56 TL    ← KAPANIŞ
  Satır 8 : Tarih Aralığı  | 01.06.2026 - 24.06.2026
  Satır 10: Tarih | Saat | Tutar | Bakiye | Borç/Alacak | Açıklama | Fiş/Dekont No
  Satır 11+: veri

Tutar zaten işaretli geliyor (B → negatif, A → pozitif).
Bakiye yürüyen bakiye — son satırın bakiyesi = kapanış, ilk satırın (bakiye - tutar) = açılış.
"""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import openpyxl

from .eslestirici import Hareket
from .parser import EkstreSonuc, sayi_parse


def akbank_uygun_mu(yol: Path) -> bool:
    """Excel'in Akbank formatında olup olmadığını sezgisel test eder."""
    try:
        wb = openpyxl.load_workbook(yol, data_only=True, read_only=True)
        ws = wb.active
        ilk_satir = next(ws.iter_rows(values_only=True, max_row=1))
        ust = str(ilk_satir[0] or "").upper()
        return "HESAP HAREKETLER" in ust
    except Exception:
        return False


def akbank_parse(yol: Path) -> EkstreSonuc:
    wb = openpyxl.load_workbook(yol, data_only=True)
    ws = wb.active

    meta: dict[str, str] = {}
    veri_basliklari_satiri = -1
    rows = list(ws.iter_rows(values_only=True))

    # Header bul
    for i, row in enumerate(rows):
        if not row or row[0] is None:
            continue
        anahtar = str(row[0]).strip()
        if anahtar == "Tarih" and row[1] and "Saat" in str(row[1]):
            veri_basliklari_satiri = i
            break
        if row[1] is not None:
            meta[anahtar] = str(row[1]).strip()

    if veri_basliklari_satiri < 0:
        return EkstreSonuc([], not_="Akbank tablo başlığı bulunamadı.", kaynak_dosya=str(yol))

    # Meta'dan çıkar
    hesap_no = meta.get("Hesap No", "")
    iban = meta.get("IBAN", "").replace(" ", "")
    banka_adi = "AKBANK"
    kapanis = sayi_parse(meta.get("Kullanılabilir Bakiye", ""))

    hareketler: list[Hareket] = []
    for row in rows[veri_basliklari_satiri + 1:]:
        if not row or row[0] is None:
            continue
        tarih_s = str(row[0]).strip()
        if not re.match(r"\d{1,2}[./]\d{1,2}[./]\d{4}", tarih_s):
            # alt bilgi satırları (Akbank T.A.Ş., Genel Müdürlük vs.)
            continue
        try:
            g, a, y = re.split(r"[./]", tarih_s)
            tarih = datetime(int(y), int(a), int(g)).strftime("%d/%m/%Y")
        except Exception:
            continue
        tutar = sayi_parse(row[2])
        if tutar is None:
            continue
        aciklama = str(row[5] or "").strip()
        fis = str(row[6] or "").strip() if len(row) > 6 else ""
        hareketler.append(Hareket(tarih=tarih, aciklama=aciklama, tutar=tutar, referans=fis))

    # Açılış bakiyesini hesapla — ilk satır = en güncel; son veri satırının önceki bakiyesini bul
    acilis = None
    if hareketler and len(rows) > veri_basliklari_satiri + len(hareketler):
        # Akbank en üstte en yeni tarih var. Son satır en eski → onun (bakiye - tutar) açılış.
        son_veri_satiri = rows[veri_basliklari_satiri + len(hareketler)]
        # ama row index hesaplaması karışık olabilir — basit yaklaşım:
        # Sondaki veri satırını bul
        son_bakiye, son_tutar = None, None
        for row in reversed(rows):
            if not row or row[0] is None:
                continue
            tarih_s = str(row[0]).strip()
            if not re.match(r"\d{1,2}[./]\d{1,2}[./]\d{4}", tarih_s):
                continue
            son_bakiye = sayi_parse(row[3])
            son_tutar = sayi_parse(row[2])
            break
        if son_bakiye is not None and son_tutar is not None:
            acilis = round(son_bakiye - son_tutar, 2)

    # Akbank en yeni tarihten başlıyor — kronolojik sıralayalım (eskiden yeniye)
    def _t(h: Hareket):
        g, a, y = h.tarih.split("/")
        return (int(y), int(a), int(g))

    hareketler.sort(key=_t)

    return EkstreSonuc(
        hareketler=hareketler,
        banka_adi=banka_adi,
        hesap_no=hesap_no,
        acilis=acilis,
        kapanis=kapanis,
        kaynak_dosya=str(yol),
    )
