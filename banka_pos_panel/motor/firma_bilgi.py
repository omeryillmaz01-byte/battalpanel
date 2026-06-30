# -*- coding: utf-8 -*-
"""
Firma bilgi deposu — MERSİS / ticaret sicil / vergi / ortak / faaliyet bilgileri.

Her firma için: firmalar/<FIRMA>/firma_bilgi.json
Ayrıca tüm firmaların özeti: firma_bilgi_ozet.xlsx (tek bakışta).

Bu modül "her şeyi barındıran panel"in firma künye katmanıdır.
"""
from __future__ import annotations
import json
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
FIRMA_DIZIN = BASE / "firmalar"

ALANLAR = [
    "unvan", "vkn", "tc", "vergi_dairesi", "vergi_turu",
    "mersis_no", "ticaret_sicil_no", "kurulus_tarihi", "sermaye",
    "faaliyet_kodu", "faaliyet_adi",
    "adres", "il", "ilce",
    "kep", "ortaklar", "yetkililer",
    "uyumsoft_kullanici", "uyumsoft_sifre",
    "ebeyanname_sifre", "vergi_dairesi_kullanici", "vergi_dairesi_sifre",
    "sgk_adi", "sgk_sicil_no", "sgk_yetkili", "sgk_kullanici_kodu",
    "sgk_sistem_sifresi", "sgk_isyeri_sifresi",
]


def bilgi_kaydet(firma_kodu: str, bilgi: dict) -> Path:
    """Firma künye bilgisini JSON olarak kaydeder (mevcutla birleştirir)."""
    d = FIRMA_DIZIN / firma_kodu
    d.mkdir(parents=True, exist_ok=True)
    yol = d / "firma_bilgi.json"
    mevcut = bilgi_oku(firma_kodu)
    # sadece dolu gelen alanları güncelle
    for k, v in bilgi.items():
        if v not in (None, "", [], {}):
            mevcut[k] = v
    yol.write_text(json.dumps(mevcut, ensure_ascii=False, indent=2), encoding="utf-8")
    return yol


def bilgi_oku(firma_kodu: str) -> dict:
    yol = FIRMA_DIZIN / firma_kodu / "firma_bilgi.json"
    if yol.exists():
        try:
            return json.loads(yol.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def tum_bilgiler() -> dict[str, dict]:
    out = {}
    if FIRMA_DIZIN.exists():
        for fd in sorted(FIRMA_DIZIN.iterdir()):
            if fd.is_dir():
                b = bilgi_oku(fd.name)
                if b:
                    out[fd.name] = b
    return out


def ozet_excel_yaz(hedef: Path | None = None) -> Path:
    """Tüm firmaların künye özetini tek Excel'e döker."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    hedef = hedef or (FIRMA_DIZIN / "firma_bilgi_ozet.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Firma Künye"

    basliklar = ["Firma Kodu", "Ünvan", "VKN/TC", "Vergi Dairesi", "Vergi Türü",
                 "Faaliyet Kodu", "Faaliyet Adı", "MERSİS No", "Tic.Sicil No",
                 "Kuruluş", "Sermaye", "KEP", "Adres", "İl", "İlçe"]
    hdr_fill = PatternFill("solid", fgColor="305496")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    sari = PatternFill("solid", fgColor="FFFF00")
    kenar = Border(*(Side(style="thin", color="BFBFBF"),) * 4)

    for j, b in enumerate(basliklar, 1):
        c = ws.cell(row=1, column=j, value=b)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    r = 2
    for kod, b in tum_bilgiler().items():
        vals = [
            kod,
            b.get("unvan", ""),
            b.get("vkn") or b.get("tc", ""),
            b.get("vergi_dairesi", ""),
            b.get("vergi_turu", ""),
            b.get("faaliyet_kodu", ""),
            b.get("faaliyet_adi", ""),
            b.get("mersis_no", ""),
            b.get("ticaret_sicil_no", ""),
            b.get("kurulus_tarihi", ""),
            b.get("sermaye", ""),
            b.get("kep", ""),
            b.get("adres", ""),
            b.get("il", ""),
            b.get("ilce", ""),
        ]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = kenar
            if not v and j > 1:
                c.fill = sari
        r += 1

    genislik = [22, 40, 14, 16, 20, 12, 45, 20, 14, 14, 18, 22, 55, 10, 14]
    for i, w in enumerate(genislik, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(hedef)
    return hedef


if __name__ == "__main__":
    print("Firma bilgi modülü hazır. Alanlar:", ", ".join(ALANLAR))
    print("Kayıtlı künye:", len(tum_bilgiler()), "firma")
