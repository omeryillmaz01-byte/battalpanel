# -*- coding: utf-8 -*-
"""
IŞIK PETROL referans Excel'lerine ek bilgiler yamar:
  - 01_banka_hesaplari.xlsx → bilinen IBAN'ları doldurur
  - 04_kural_sozlugu.xlsx   → Akbank'a özel anahtar kelime eşleşmelerini ekler

Tek seferlik yardımcı: panele "Akbank entegrasyonu" yapılırken çağrıldı.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side

BASE = Path(__file__).resolve().parent.parent
FIRMA = BASE / "firmalar" / "ISIK_PETROL"

KENAR = Border(*(Side(style="thin", color="BFBFBF"),) * 4)
SOL = Alignment(horizontal="left", vertical="center", wrap_text=True)
ORTA = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Hesap no → IBAN (ekran görüntülerinden ve Akbank dosyasından)
IBAN_HARITASI = {
    "157047": "TR20 0004 6002 7188 8000 1570 47",  # 102.07 SHELL&TURCAS
    "28620":  "TR90 0004 6002 7188 8000 0286 20",  # 102.08 POS HESABI
    "183316": "TR21 0004 6002 7188 8000 1833 16",  # 102.05 Market
    "157419": "TR64 0004 6002 7188 8000 1574 19",  # 102.09 PURFON/OPET (FON)
    "44994":  "TR21 0004 6002 7188 8000 0449 94",  # 102.13
    "178445": "TR06 0004 6002 7188 8000 1784 45",  # 102.06
}


def iban_yama() -> int:
    yol = FIRMA / "01_banka_hesaplari.xlsx"
    wb = openpyxl.load_workbook(yol)
    ws = wb.active
    # Başlık satırını bul
    bas = -1
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] and "HESAP KODU" in str(row[0]).upper():
            bas = i
            break
    if bas < 0:
        return 0
    sayac = 0
    for r in range(bas + 1, ws.max_row + 1):
        hesap_no_cell = ws.cell(row=r, column=4)
        iban_cell = ws.cell(row=r, column=5)
        if not hesap_no_cell.value:
            continue
        # Son parça, baştaki sıfırlar atılmış
        son_parca = str(hesap_no_cell.value).split("-")[-1].lstrip("0")
        if son_parca in IBAN_HARITASI and not iban_cell.value:
            iban_cell.value = IBAN_HARITASI[son_parca]
            iban_cell.alignment = SOL
            iban_cell.border = KENAR
            sayac += 1
    wb.save(yol)
    return sayac


# Akbank'a özel kural ekleri
AKBANK_KURALLARI = [
    # AKPOS hakedişi — 108.xx (bu bankaya bağlı POS)
    ("AKPOS|POS ODE|POS PES|POS ÖDEME|PES ODE",  "108.??", "POS TAHSİLATI (banka adına göre)", "POZITIF"),
    # MBL-HAV → mobil havale, MBL-VİRMAN, MBL-EFT vs.
    ("MBL-EFT|MOBİL EFT|MBL EFT",                "780.??", "FİNANSMAN GİDERİ", "NEGATIF"),
    # Hesap işletim ücreti vs. (Akbank açıklamaları)
    ("HESAP ISLETIM|HESAP İŞLETİM|İŞLETİM ÜCRETİ", "780.??", "FİNANSMAN GİDERİ", "NEGATIF"),
    # Para yatırma → genelde 100 KASA'dan gelir, manuel kontrol için boş
    # NOT: 'Para Yatırma' kural sözlüğüne girmedi — manuel kalsın ki kullanıcı bilinçli karar versin.
]


def kural_yama() -> int:
    yol = FIRMA / "04_kural_sozlugu.xlsx"
    wb = openpyxl.load_workbook(yol)
    ws = wb.active
    # Başlık + sıra bul
    bas = -1
    max_sira = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] and "SIRA" in str(row[0]).upper():
            bas = i
            continue
        if bas > 0 and row and row[0]:
            try:
                max_sira = max(max_sira, int(row[0]))
            except (ValueError, TypeError):
                pass
    if bas < 0:
        return 0

    # Mevcut anahtarları al
    mevcut = set()
    for r in range(bas + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if v:
            mevcut.add(str(v).strip())

    sira = max_sira + 1
    eklenen = 0
    yazma_satiri = ws.max_row + 1
    for kelime, kod, ad, isaret in AKBANK_KURALLARI:
        if kelime in mevcut:
            continue
        ws.cell(row=yazma_satiri, column=1, value=sira).alignment = ORTA
        ws.cell(row=yazma_satiri, column=2, value=kelime).alignment = SOL
        ws.cell(row=yazma_satiri, column=3, value=kod).alignment = ORTA
        ws.cell(row=yazma_satiri, column=4, value=ad).alignment = SOL
        ws.cell(row=yazma_satiri, column=5, value=isaret).alignment = ORTA
        ws.cell(row=yazma_satiri, column=6, value="E").alignment = ORTA
        for j in range(1, 7):
            ws.cell(row=yazma_satiri, column=j).border = KENAR
        yazma_satiri += 1
        sira += 1
        eklenen += 1
    wb.save(yol)
    return eklenen


if __name__ == "__main__":
    i = iban_yama()
    k = kural_yama()
    print(f"IBAN doldurulan satır: {i}")
    print(f"Eklenen Akbank kuralı : {k}")
