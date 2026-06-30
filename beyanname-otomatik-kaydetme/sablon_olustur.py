"""
MÜKELLEF VERGİ ŞABLONU oluşturucu.

Proje klasöründeki tüm uretilen_*.csv dosyalarını tarar.
Her firma için hangi vergileri verdiğini ve MUHSGK için periyodu (AYLIK / 3 AYLIK) tespit eder.
Çıktı: 'MUKELLEF VERGI ŞABLONU.xlsx'
"""
import csv as csvmod
import re
import sys
from pathlib import Path
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from indir import mukellef_listesi_yukle, klasor_bul, _norm, _tr, _kw, _sim

BASE = Path(__file__).parent
TURLER = ["KDV1","KDV2","MUHSGK","GGECICI","KGECICI","KURUMLAR","GELIR","POSET","DAMGA"]
GORUNEN = {"POSET":"GEKAP"}
# GIB tür kodu → şablon sütun adı normalizasyonu
TUR_NORM = {"GELIR1001E":"GELIR","KURUMLARP":"KURUMLAR"}

def donem_periyot(donem_str):
    """Vergilendirme_Donemi'nden periyot çıkar.
    '01/2025-01/2025' → AYLIK
    '01/2025-03/2025' → 3 AYLIK
    '01/2025-12/2025' → YILLIK
    """
    aylar = re.findall(r'(\d{2})/(\d{4})', str(donem_str))
    if len(aylar) < 2: return "AYLIK"
    bas_ay, bas_yil = int(aylar[0][0]), int(aylar[0][1])
    bit_ay, bit_yil = int(aylar[-1][0]), int(aylar[-1][1])
    if bas_ay == bit_ay and bas_yil == bit_yil:
        return "AYLIK"
    fark = (bit_yil - bas_yil) * 12 + (bit_ay - bas_ay) + 1
    if fark >= 12: return "YILLIK"
    if fark == 3: return "3 AYLIK"
    return f"{fark} AYLIK"

def main():
    # Mükellef listesi
    taner_set, omer_set = mukellef_listesi_yukle(BASE)
    tum_mukellef = [(f, "TANER") for f in sorted(taner_set)] + [(f, "ÖMER") for f in sorted(omer_set)]
    print(f"  Mükellef: {len(tum_mukellef)} firma (TANER {len(taner_set)} + ÖMER {len(omer_set)})")

    # CSV'leri tara
    # firma_normli → {tur: set(periyotlar)}
    firma_turleri = defaultdict(lambda: defaultdict(set))
    csv_yollari = sorted(BASE.glob("uretilen_*.csv"))
    print(f"  Taranan CSV: {len(csv_yollari)} dosya")

    for csv_yol in csv_yollari:
        try:
            with open(csv_yol, encoding="utf-8-sig") as f:
                for r in csvmod.DictReader(f, delimiter=";"):
                    ad = str(r.get("Ad_Soyad","") or "").strip().upper()
                    tur = str(r.get("Beyanname_Turu","") or "").strip().upper()
                    donem = str(r.get("Vergilendirme_Donemi","") or "")
                    if not ad or not tur: continue
                    tur = TUR_NORM.get(tur, tur)
                    if tur not in TURLER: continue
                    fn = _norm(ad)
                    periyot = donem_periyot(donem)
                    firma_turleri[fn][tur].add(periyot)
        except Exception as e:
            print(f"  [!] {csv_yol.name}: {e}")

    print(f"  Tespit edilen firma: {len(firma_turleri)}")

    # Her mükellef için CSV'de bulunan en yakın eşleşmeyi bul
    def eslesme_bul(mukellef):
        mn = _norm(mukellef)
        if mn in firma_turleri: return mn
        # Kısmi eşleşme
        for fn in firma_turleri:
            if fn and (fn in mn or mn in fn): return fn
        # Anahtar kelime ile fuzzy
        mkw = _kw(mukellef)
        if len(mkw) >= 2:
            for fn in firma_turleri:
                fkw = set(re.findall(r'[A-Z0-9]+', fn))
                e = sum(1 for kw in mkw[:2] if kw in fkw or any(len(kw)>3 and _sim(kw,h)>=0.82 for h in fkw))
                if e >= 2: return fn
        return None

    # Excel oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Mükellef Vergi Şablonu"

    baslik_fill = PatternFill("solid", fgColor="1F2937")
    baslik_font = Font(bold=True, color="FFFFFF", size=11)
    aktif_fill  = PatternFill("solid", fgColor="DDF2D5")
    aylik_font  = Font(bold=True, color="006100")
    uc_aylik_font = Font(bold=True, color="0050C0")
    yillik_font = Font(bold=True, color="7C3AED")
    merkez = Alignment(horizontal="center", vertical="center")
    sol = Alignment(horizontal="left", vertical="center")
    kenar = Border(*[Side(style="thin", color="999999")] * 4)

    # Başlık satırı
    basliklar = ["NO:", "FİRMA ÜNVANI", "SAHİBİ"]
    for t in TURLER:
        basliklar.append(GORUNEN.get(t, t))
    basliklar += ["NOT"]
    for i, h in enumerate(basliklar, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = baslik_fill; cell.font = baslik_font
        cell.alignment = merkez; cell.border = kenar
    ws.row_dimensions[1].height = 24

    # Satırlar
    for i, (firma, sahip) in enumerate(tum_mukellef, 2):
        ws.cell(row=i, column=1, value=i-1).alignment = merkez
        ws.cell(row=i, column=2, value=firma).alignment = sol
        ws.cell(row=i, column=3, value=sahip).alignment = merkez

        esl = eslesme_bul(firma)
        for j, t in enumerate(TURLER, 4):
            cell = ws.cell(row=i, column=j)
            cell.alignment = merkez; cell.border = kenar
            if esl and t in firma_turleri[esl]:
                periyotlar = firma_turleri[esl][t]
                if t == "MUHSGK":
                    if "3 AYLIK" in periyotlar:
                        cell.value = "3 AYLIK"; cell.font = uc_aylik_font
                    elif "AYLIK" in periyotlar:
                        cell.value = "AYLIK"; cell.font = aylik_font
                    else:
                        cell.value = ", ".join(periyotlar)
                    cell.fill = aktif_fill
                elif t in ("KURUMLAR","GELIR"):
                    cell.value = "✓"; cell.fill = aktif_fill; cell.font = yillik_font
                elif t in ("GGECICI","KGECICI"):
                    cell.value = "3 AYLIK"; cell.fill = aktif_fill; cell.font = uc_aylik_font
                else:
                    cell.value = "✓"; cell.fill = aktif_fill; cell.font = aylik_font
        # Diğer hücrelere kenar
        for j in range(1, len(basliklar)+1):
            ws.cell(row=i, column=j).border = kenar

    # Sütun genişlikleri
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 10
    for j in range(4, 4+len(TURLER)):
        ws.column_dimensions[get_column_letter(j)].width = 12
    ws.column_dimensions[get_column_letter(4+len(TURLER))].width = 25
    ws.freeze_panes = "D2"

    cikti = BASE / "MUKELLEF VERGI ŞABLONU.xlsx"
    wb.save(cikti)
    print(f"\n[OK] Şablon yazıldı: {cikti}")

if __name__ == "__main__":
    main()
