# -*- coding: utf-8 -*-
"""
EBYN Otomasyon Paneli — v3.1 (2026-06-27)
========================================
Tüm beyanname işlerini tek pencereden, butonla yapmak için masaüstü arayüzü.

v3 özellikleri (tam liste):
  - GIB her zaman canlı sorgu (bayat veri tuzağı yok)
  - Çoklu yükleme aralığı: dönem+1 ayından bugüne (geç yüklemeler dahil)
  - Form hazırsa beklemeden geç (hız)
  - Disk CSV + GIB birleştirme (BynOID dedupe)
  - Ek-temizleyen eşleştirme (LTD.ŞTİ./A.Ş./ADİ ORT. farkı bozmaz)
  - Şablon-öncelikli kurated ünvan (GİZEM GÖKER vs ADİ ORT. çakışma yok)
  - Yükleme Zamanı + Hata Durumu ayrı kolonlar
  - Akıllı sıralama (Verildi üstte yeşil, Verilmedi altta kırmızı)
  - Excel otomatik filtre (renge/değere göre filtrele-sırala)
  - Tüm beklenen firmalar her ay kontrolde (3 aylık dahil)
  - Durum filtresi panelden seçilir (varsayılan: Onaylandı)
  - Pasif mükellef otomatik filtre (PASIF MUKELLEFLER.xlsx)
  - Yeni firma otomatik mükellef listesine eklenir

Çift tıkla çalıştırmak için: PANEL.bat
"""
__version__ = "3.1"

import os
import sys
import json
import threading
import queue
import subprocess
import traceback
import datetime
import time
from pathlib import Path
from urllib.request import urlopen

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

BASE = Path(__file__).resolve().parent
INDIR_PY = BASE / "indir.py"
CHROME_PROFIL = BASE / "chrome_profil"
DEBUG_PORT = 9222

RENK_BG    = "#1e1e2e"
RENK_PANEL = "#272739"
RENK_YAZI  = "#e6e6e6"
RENK_YESIL = "#28a745"
RENK_MAVI  = "#3b82f6"
RENK_KIRMIZI = "#e23b3b"
RENK_LOG_BG = "#11111b"

# Panel etiketi → GIB beyanname kodu (form select + indir.py --tur + CSV eşleşme)
TUR_KODLARI = {
    "KDV1": "KDV1", "KDV2": "KDV2", "MUHSGK": "MUHSGK",
    "GGECİCİ": "GGECICI", "KGECİCİ": "KGECICI",
    "KURUMLAR": "KURUMLAR", "GELİR": "GELIR1001E",
    "GEKAP": "POSET", "DAMGA": "DAMGA",
}

# Dosyaların kaydedileceği klasörler (indir.py ile aynı — TANER / ÖMER)
KAYIT_YERLERI = {
    "KDV1": [
        r"C:\Users\omery\iCloudDrive\SMMM ÖMER YILMAZ\1-) MUHASEBE\1-) OFİS MUHASEBE\4-) BEYAN VE TAHAKKUKLAR\1-) KDV BEYAN & THK",
        r"C:\Users\omery\iCloudDrive\SMMM ÖMER YILMAZ\1-) MUHASEBE\2-) ÖMER MUHASEBE\1-) BEYAN VE TAHAKKUKLAR\2026 YILI\1-) KDV BYN & THK",
    ],
    "MUHSGK": [
        r"C:\Users\omery\iCloudDrive\SMMM ÖMER YILMAZ\1-) MUHASEBE\1-) OFİS MUHASEBE\4-) BEYAN VE TAHAKKUKLAR\2-) MUHTASAR BYN & THK",
        r"C:\Users\omery\iCloudDrive\SMMM ÖMER YILMAZ\1-) MUHASEBE\2-) ÖMER MUHASEBE\1-) BEYAN VE TAHAKKUKLAR\2026 YILI\2-) MUHTASAR BYN & THK",
    ],
}


def chrome_bul():
    adaylar = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
    ]
    for c in adaylar:
        if os.path.exists(c):
            return c
    return None


try:
    from tkinterdnd2 import TkinterDnD
    _BASE_TK = TkinterDnD.Tk
except Exception:
    _BASE_TK = tk.Tk

class Panel(_BASE_TK):
    def __init__(self):
        super().__init__()
        self.title("EBYN Otomasyon Paneli")
        self.geometry("860x680")
        self.minsize(760, 600)
        self.configure(bg=RENK_BG)

        self.kuyruk = queue.Queue()
        self.calisiyor = False
        self._dur_iste = False   # toplu/indirme durdurma bayrağı
        self.surec = None

        self._stil()
        self._arayuz()
        self.after(150, self._kuyruk_isle)
        self.after(500, self.baglanti_kontrol)

    # ---------- stil ----------
    def _stil(self):
        s = ttk.Style(self)
        try:
            s.theme_use("clam")
        except Exception:
            pass
        s.configure("TFrame", background=RENK_BG)
        s.configure("Panel.TFrame", background=RENK_PANEL)
        s.configure("TLabel", background=RENK_BG, foreground=RENK_YAZI, font=("Segoe UI", 10))
        s.configure("Baslik.TLabel", background=RENK_BG, foreground="#ffffff", font=("Segoe UI Semibold", 16))
        s.configure("Alt.TLabel", background=RENK_PANEL, foreground=RENK_YAZI, font=("Segoe UI", 10))
        s.configure("Bolum.TLabel", background=RENK_BG, foreground="#9aa0c0", font=("Segoe UI Semibold", 11))
        s.configure("TButton", font=("Segoe UI", 10), padding=6)
        s.configure("Yesil.TButton", font=("Segoe UI Semibold", 11))
        s.configure("TCombobox", fieldbackground="#ffffff", font=("Segoe UI", 10))
        s.configure("TEntry", font=("Segoe UI", 10))

    # ---------- arayüz ----------
    def _arayuz(self):
        ttk.Label(self, text="EBYN Otomasyon Paneli", style="Baslik.TLabel").pack(anchor="w", padx=18, pady=(14, 2))
        ttk.Label(self, text="Beyanname & Tahakkuk + SGK indirme — tek tıkla").pack(anchor="w", padx=18, pady=(0, 10))

        # --- Chrome bölümü ---
        cf = tk.Frame(self, bg=RENK_PANEL)
        cf.pack(fill="x", padx=16, pady=6)
        tk.Button(cf, text="1) Chrome'u Başlat & Giriş Yap", command=self.chrome_baslat,
                  bg=RENK_MAVI, fg="white", relief="flat", font=("Segoe UI Semibold", 10),
                  activebackground="#2563eb", activeforeground="white", padx=12, pady=8, cursor="hand2"
                  ).pack(side="left", padx=10, pady=10)
        self.durum_isik = tk.Label(cf, text="●", bg=RENK_PANEL, fg=RENK_KIRMIZI, font=("Segoe UI", 14))
        self.durum_isik.pack(side="left", padx=(10, 2))
        self.durum_yazi = tk.Label(cf, text="Bağlı değil", bg=RENK_PANEL, fg=RENK_YAZI, font=("Segoe UI", 10))
        self.durum_yazi.pack(side="left")
        tk.Button(cf, text="Bağlantıyı Kontrol Et", command=self.baglanti_kontrol,
                  bg="#3a3a52", fg="white", relief="flat", padx=10, pady=6, cursor="hand2",
                  activebackground="#4a4a6a", activeforeground="white").pack(side="right", padx=10)

        # --- İndirme bölümü ---
        ttk.Label(self, text="İNDİRME", style="Bolum.TLabel").pack(anchor="w", padx=18, pady=(14, 4))
        inf = tk.Frame(self, bg=RENK_PANEL)
        inf.pack(fill="x", padx=16, pady=4)

        satir1 = tk.Frame(inf, bg=RENK_PANEL)
        satir1.pack(fill="x", padx=12, pady=(12, 6))
        tk.Label(satir1, text="Beyanname Türü:", bg=RENK_PANEL, fg=RENK_YAZI, width=16, anchor="w").pack(side="left")
        self.tur_var = tk.StringVar(value="KDV1")
        ttk.Combobox(satir1, textvariable=self.tur_var,
                     values=["KDV1", "KDV2", "MUHSGK", "GGECİCİ", "KGECİCİ", "KURUMLAR", "GELİR", "GEKAP", "DAMGA"],
                     state="readonly", width=14).pack(side="left")

        tk.Label(satir1, text="   Dönem (ay/yıl):", bg=RENK_PANEL, fg=RENK_YAZI).pack(side="left")
        self.ay_var = tk.StringVar()
        self.yil_var = tk.StringVar(value="2026")
        tk.Entry(satir1, textvariable=self.ay_var, width=5, justify="center").pack(side="left", padx=(6, 2))
        tk.Label(satir1, text="/", bg=RENK_PANEL, fg=RENK_YAZI).pack(side="left")
        tk.Entry(satir1, textvariable=self.yil_var, width=7, justify="center").pack(side="left", padx=(2, 6))
        self.liste_ac_btn = tk.Button(satir1, text="🔎 Listeyi Aç", command=self.liste_ac_basla,
                  bg="#0f766e", fg="white", relief="flat", padx=10, cursor="hand2",
                  activebackground="#0d9488", activeforeground="white")
        self.liste_ac_btn.pack(side="left", padx=(8, 4))
        tk.Label(satir1, text=" Durum:", bg=RENK_PANEL, fg=RENK_YAZI).pack(side="left", padx=(4, 2))
        self.durum_filtre_var = tk.StringVar(value="Onaylandı")
        ttk.Combobox(
            satir1,
            textvariable=self.durum_filtre_var,
            values=["Onaylandı", "Tümü", "Hatalı", "Onay bekliyor", "İptal"],
            state="readonly",
            width=13,
        ).pack(side="left", padx=(0, 6))
        self.donem_hint_lbl = tk.Label(satir1, text="", bg=RENK_PANEL, fg="#8a8aa0",
                                       font=("Segoe UI", 9, "italic"))
        self.donem_hint_lbl.pack(side="left")
        self.tur_var.trace_add("write", lambda *a: (self._kayit_yeri_guncelle(), self._donem_hint_guncelle()))

        satir2 = tk.Frame(inf, bg=RENK_PANEL)
        satir2.pack(fill="x", padx=12, pady=6)
        tk.Label(satir2, text="CSV Dosyası:", bg=RENK_PANEL, fg=RENK_YAZI, width=16, anchor="w").pack(side="left")
        self.csv_var = tk.StringVar()
        tk.Entry(satir2, textvariable=self.csv_var, width=58).pack(side="left", padx=(0, 6))
        tk.Button(satir2, text="Seç...", command=self.csv_sec, bg="#3a3a52", fg="white", relief="flat",
                  padx=10, cursor="hand2", activebackground="#4a4a6a", activeforeground="white").pack(side="left")
        self.csv_olustur_btn = tk.Button(satir2, text="📄 CSV Oluştur", command=self.csv_olustur_basla,
                  bg="#6d28d9", fg="white", relief="flat", padx=10, cursor="hand2",
                  activebackground="#7c3aed", activeforeground="white")
        self.csv_olustur_btn.pack(side="left", padx=6)

        satir3 = tk.Frame(inf, bg=RENK_PANEL)
        satir3.pack(fill="x", padx=12, pady=(6, 14))
        self.indir_btn = tk.Button(satir3, text="⬇  İNDİR", command=self.indir_basla,
                                    bg=RENK_YESIL, fg="white", relief="flat", font=("Segoe UI Semibold", 12),
                                    padx=22, pady=8, cursor="hand2",
                                    activebackground="#22963c", activeforeground="white")
        self.indir_btn.pack(side="left")
        self.durdur_btn = tk.Button(satir3, text="Durdur", command=self.indir_durdur, state="disabled",
                                    bg="#5a3a3a", fg="white", relief="flat", padx=14, pady=8, cursor="hand2",
                                    activebackground="#7a4a4a", activeforeground="white")
        self.durdur_btn.pack(side="left", padx=8)

        # Toplu mod: dönem aralığı (Listeyi Aç + CSV Oluştur + İNDİR otomatik zincir)
        satir_t = tk.Frame(inf, bg=RENK_PANEL)
        satir_t.pack(fill="x", padx=12, pady=(0, 10))
        tk.Label(satir_t, text="🔁 Toplu (dönem aralığı):", bg=RENK_PANEL, fg="#ffd479",
                 font=("Segoe UI Semibold", 10), width=22, anchor="w").pack(side="left")
        self.toplu_bas_ay = tk.StringVar(value="1")
        self.toplu_bas_yil = tk.StringVar(value="2025")
        self.toplu_bit_ay = tk.StringVar(value="12")
        self.toplu_bit_yil = tk.StringVar(value="2025")
        tk.Entry(satir_t, textvariable=self.toplu_bas_ay, width=4, justify="center").pack(side="left", padx=(2, 1))
        tk.Label(satir_t, text="/", bg=RENK_PANEL, fg=RENK_YAZI).pack(side="left")
        tk.Entry(satir_t, textvariable=self.toplu_bas_yil, width=6, justify="center").pack(side="left", padx=(1, 4))
        tk.Label(satir_t, text="→", bg=RENK_PANEL, fg=RENK_YAZI).pack(side="left")
        tk.Entry(satir_t, textvariable=self.toplu_bit_ay, width=4, justify="center").pack(side="left", padx=(4, 1))
        tk.Label(satir_t, text="/", bg=RENK_PANEL, fg=RENK_YAZI).pack(side="left")
        tk.Entry(satir_t, textvariable=self.toplu_bit_yil, width=6, justify="center").pack(side="left", padx=(1, 8))
        self.toplu_btn = tk.Button(satir_t, text="🔁 Toplu Çalıştır", command=self.toplu_basla,
                                   bg="#b45309", fg="white", relief="flat", font=("Segoe UI Semibold", 10),
                                   padx=14, pady=6, cursor="hand2",
                                   activebackground="#d97706", activeforeground="white")
        self.toplu_btn.pack(side="left")

        # Kontrol Listesi satırı
        satir_kl = tk.Frame(inf, bg=RENK_PANEL)
        satir_kl.pack(fill="x", padx=12, pady=(0, 8))
        tk.Label(satir_kl, text="📋 Kontrol Listesi:", bg=RENK_PANEL, fg="#7ee787",
                 font=("Segoe UI Semibold", 10), width=22, anchor="w").pack(side="left")
        tk.Label(satir_kl, text="Üst Beyanname Türü + Dönem (ay/yıl) kutularını kullanır",
                 bg=RENK_PANEL, fg="#9aa0c0", font=("Segoe UI", 9)).pack(side="left", padx=(0, 8))
        self.kontrol_btn = tk.Button(satir_kl, text="📋 Excel Oluştur", command=self.kontrol_listesi,
                                     bg="#16a34a", fg="white", relief="flat", font=("Segoe UI Semibold", 10),
                                     padx=14, pady=5, cursor="hand2",
                                     activebackground="#22c55e", activeforeground="white")
        self.kontrol_btn.pack(side="right")
        self.sablon_yenile_btn = tk.Button(satir_kl, text="📊 Şablonu Yenile",
                                           command=self.sablon_yenile,
                                           bg="#0ea5e9", fg="white", relief="flat",
                                           font=("Segoe UI Semibold", 10),
                                           padx=14, pady=5, cursor="hand2",
                                           activebackground="#38bdf8", activeforeground="white")
        self.sablon_yenile_btn.pack(side="right", padx=(0, 6))

        # Kayıt yeri göstergesi
        satir4 = tk.Frame(inf, bg=RENK_PANEL)
        satir4.pack(fill="x", padx=12, pady=(0, 12))
        tk.Label(satir4, text="Kayıt yeri:", bg=RENK_PANEL, fg="#9aa0c0", width=16, anchor="w").pack(side="left")
        self.kayit_yeri_lbl = tk.Label(satir4, text="", bg=RENK_PANEL, fg="#8a8aa0",
                                       font=("Segoe UI", 9), anchor="w", justify="left")
        self.kayit_yeri_lbl.pack(side="left", fill="x", expand=True)
        tk.Button(satir4, text="📁 TANER", command=lambda: self._klasor_ac(0), bg="#3a3a52", fg="white",
                  relief="flat", padx=8, cursor="hand2", activebackground="#4a4a6a",
                  activeforeground="white").pack(side="right", padx=(4, 0))
        tk.Button(satir4, text="📁 ÖMER", command=lambda: self._klasor_ac(1), bg="#3a3a52", fg="white",
                  relief="flat", padx=8, cursor="hand2", activebackground="#4a4a6a",
                  activeforeground="white").pack(side="right", padx=4)

        # Test modu
        satir5 = tk.Frame(inf, bg=RENK_PANEL)
        satir5.pack(fill="x", padx=12, pady=(0, 12))
        self.test_var = tk.BooleanVar(value=True)
        tk.Checkbutton(satir5, text="🧪 Test modu (iCloud yerine test klasörüne indir)", variable=self.test_var,
                       bg=RENK_PANEL, fg="#ffd479", selectcolor=RENK_PANEL, activebackground=RENK_PANEL,
                       activeforeground="#ffd479", font=("Segoe UI", 10), cursor="hand2",
                       command=self._test_durum).pack(side="left")
        # Görünen masaüstü = proje klasörünün üst klasörü (OneDrive\Desktop olabilir)
        self.test_yol_var = tk.StringVar(value=str(BASE.parent / "EBYN_TEST"))
        self.test_entry = tk.Entry(satir5, textvariable=self.test_yol_var, width=38)
        self.test_entry.pack(side="left", padx=(10, 4))
        tk.Button(satir5, text="Seç...", command=self._test_klasor_sec, bg="#3a3a52", fg="white", relief="flat",
                  padx=8, cursor="hand2", activebackground="#4a4a6a", activeforeground="white").pack(side="left")
        tk.Button(satir5, text="📁 Aç", command=self._test_klasor_ac, bg="#3a3a52", fg="white", relief="flat",
                  padx=8, cursor="hand2", activebackground="#4a4a6a", activeforeground="white").pack(side="left", padx=4)

        # Hızlı mod
        satir5b = tk.Frame(inf, bg=RENK_PANEL)
        satir5b.pack(fill="x", padx=12, pady=(0, 12))
        self.atla_var = tk.BooleanVar(value=True)
        tk.Checkbutton(satir5b, text="⚡ Hızlı mod (BYN+THK varsa HTTP atmadan atla)", variable=self.atla_var,
                       bg=RENK_PANEL, fg="#7ee787", selectcolor=RENK_PANEL, activebackground=RENK_PANEL,
                       activeforeground="#7ee787", font=("Segoe UI", 10), cursor="hand2").pack(side="left")

        # --- Log ---
        ttk.Label(self, text="CANLI LOG", style="Bolum.TLabel").pack(anchor="w", padx=18, pady=(12, 4))
        self.log = scrolledtext.ScrolledText(self, bg=RENK_LOG_BG, fg="#d4d4d4", insertbackground="white",
                                             font=("Consolas", 9), relief="flat", height=14, wrap="word")
        self.log.pack(fill="both", expand=True, padx=16, pady=(0, 14))
        self.log.configure(state="disabled")
        self.yaz("Panele hoş geldin! 1) Chrome'u başlat, GIB'e gir. 2) CSV seç, türü seç, İNDİR'e bas.\n")

        self.surec = None
        self._kayit_yeri_guncelle()
        self._donem_hint_guncelle()

    def _donem_hint_guncelle(self):
        p = self._periyot(self._kod())
        if p == "ceyrek":
            self.donem_hint_lbl.configure(text="← çeyrek: 1=1-3, 2=4-6, 3=7-9, 4=10-12  /  yıl")
        elif p == "yil":
            self.donem_hint_lbl.configure(text="← yıllık: ay'ı boş bırak, sadece yıl yaz")
        else:
            self.donem_hint_lbl.configure(text="← ay (1-12) / yıl")

    def _kayit_yeri_guncelle(self):
        tur = self.tur_var.get()
        yerler = KAYIT_YERLERI.get(tur)
        if yerler:
            tip = "MUHTASAR" if tur == "MUHSGK" else "KDV"
            self.kayit_yeri_lbl.configure(
                text=f"{tip} klasörleri → firma sahibine göre TANER (Ofis) / ÖMER, ay alt klasörüne")
        else:
            self.kayit_yeri_lbl.configure(text="(indir.py içindeki yollara)")

    def _test_durum(self):
        self.test_entry.configure(state="normal" if self.test_var.get() else "disabled")

    def _test_klasor_sec(self):
        yol = filedialog.askdirectory(title="Test klasörü seç", initialdir=str(Path.home() / "Desktop"))
        if yol:
            self.test_yol_var.set(yol)

    def _test_klasor_ac(self):
        yol = self.test_yol_var.get().strip()
        try:
            os.makedirs(yol, exist_ok=True)
            os.startfile(yol)
        except Exception as e:
            messagebox.showinfo("Test klasörü", f"{yol}\n\n({e})")

    def _klasor_ac(self, idx):
        yerler = KAYIT_YERLERI.get(self.tur_var.get())
        if not yerler or idx >= len(yerler):
            return
        yol = yerler[idx]
        try:
            os.startfile(yol)
        except Exception as e:
            messagebox.showinfo("Klasör", f"Klasör:\n{yol}\n\n(Açılamadı: {e})")

    # ---------- log ----------
    def yaz(self, metin):
        self.log.configure(state="normal")
        self.log.insert("end", metin)
        self.log.see("end")
        self.log.configure(state="disabled")

    # ---------- chrome ----------
    def chrome_baslat(self):
        chrome = chrome_bul()
        if not chrome:
            messagebox.showerror("Chrome bulunamadı", "Google Chrome kurulu değil ya da bulunamadı.")
            return
        try:
            subprocess.Popen([
                chrome,
                f"--remote-debugging-port={DEBUG_PORT}",
                "--remote-allow-origins=*",
                f"--user-data-dir={CHROME_PROFIL}",
                "--no-first-run",
                "--no-default-browser-check",
                "https://ebeyanname.gib.gov.tr",
            ])
            self.yaz("[Chrome] Başlatıldı. Açılan pencerede GIB'e giriş yap.\n")
            self.after(2500, self.baglanti_kontrol)
        except Exception as e:
            messagebox.showerror("Hata", f"Chrome başlatılamadı:\n{e}")

    def baglanti_kontrol(self):
        try:
            with urlopen(f"http://localhost:{DEBUG_PORT}/json", timeout=2) as r:
                hedefler = json.loads(r.read().decode("utf-8", "replace"))
            gib = any("gib.gov.tr" in (t.get("url", "")) for t in hedefler)
            if gib:
                self._durum(True, "Bağlı — GIB açık")
            else:
                self._durum(True, "Chrome açık ama GIB sekmesi yok")
        except Exception:
            self._durum(False, "Bağlı değil")

    def _durum(self, ok, metin):
        self.durum_isik.configure(fg=RENK_YESIL if ok else RENK_KIRMIZI)
        self.durum_yazi.configure(text=metin)

    # ---------- csv ----------
    def csv_sec(self):
        yol = filedialog.askopenfilename(
            title="CSV seç", initialdir=str(Path.home() / "Downloads"),
            filetypes=[("CSV dosyası", "*.csv"), ("Tüm dosyalar", "*.*")])
        if yol:
            self.csv_var.set(yol)
            # Türü dosya adından tahmin et
            ad = Path(yol).name.lower()
            if "muhsgk" in ad or "muhtasar" in ad:
                self.tur_var.set("MUHSGK")
            elif "kdv" in ad:
                self.tur_var.set("KDV1")

    def _kod(self):
        """Panel etiketini GIB koduna çevir (GEKAP→POSET vb.)."""
        t = self.tur_var.get().strip()
        return TUR_KODLARI.get(t, t)

    def _durum_filtre(self):
        """Panel durum seçimini GIB radio değerine çevir.
        GIB: 0=Hatalı, 1=Onay bekliyor, 2=Onaylandı, 3=İptal.
        ALL seçilirse durum filtresi kapalı gider."""
        secim = (self.durum_filtre_var.get() if hasattr(self, "durum_filtre_var") else "Onaylandı").strip()
        harita = {
            "Tümü": ("ALL", "Tümü"),
            "Hatalı": ("0", "Hatalı"),
            "Onay bekliyor": ("1", "Onay bekliyor"),
            "Onaylandı": ("2", "Onaylandı"),
            "İptal": ("3", "İptal"),
        }
        return harita.get(secim, ("2", "Onaylandı"))

    # ---------- liste aç / filtrele (Chrome'dan, CDP) ----------
    def liste_ac_basla(self):
        if self.calisiyor:
            return

        kod = self._kod()   # GIB kodu (GEKAP→POSET)
        ay = self.ay_var.get().strip()
        yil = self.yil_var.get().strip()
        p = self._periyot(kod)
        if not yil.isdigit():
            messagebox.showwarning("Yıl gerekli", "Yıl yaz (örn. 2026).")
            return
        if p != "yil" and not ay.isdigit():
            mesaj = ("Çeyrek yaz: 1=Oca-Mar, 2=Nis-Haz, 3=Tem-Eyl"
                     if p == "ceyrek" else "Ay yaz (1-12).")
            messagebox.showwarning("Dönem gerekli", mesaj)
            return
        try:
            yukleme_bas, yukleme_bit, _filtre, ozet = self._donem_bilgisi(kod, ay or "1", yil)
        except ValueError as e:
            messagebox.showwarning("Dönem hatalı", str(e))
            return

        durum_kodu, durum_etiket = self._durum_filtre()
        snippet_js = self._liste_ac_snippet(kod, 0, int(yil), yukleme_bas, yukleme_bit, durum_kodu)

        self.calisiyor = True
        self._toplu_buton_durum("disabled")
        self.durdur_btn.configure(state="disabled")
        self.yaz("\n" + "=" * 60 + "\n")
        self.yaz(f"[Liste Aç] {self.tur_var.get()} ({kod}) — dönem {ozet}, GIB formu dolduruluyor...\n")
        self.yaz(f"Yükleme tarihi aralığı: {yukleme_bas:%d/%m/%Y} → {yukleme_bit:%d/%m/%Y}, durum: {durum_etiket}.\n")
        self.yaz("=" * 60 + "\n")

        t = threading.Thread(target=self._liste_ac_thread, args=(snippet_js,), daemon=True)
        t.start()

    def _liste_ac_thread(self, snippet_js):
        try:
            sonuc = self._cdp_calistir(snippet_js)
            self.kuyruk.put(f"[✓] {sonuc or 'Sorgu gönderildi.'}\n")
            self.kuyruk.put(("__LISTEBITTI__", True))
        except Exception:
            self.kuyruk.put("[HATA] " + traceback.format_exc() + "\n")
            self.kuyruk.put(("__LISTEBITTI__", False))

    def _yukleme_araligi(self, ay, yil):
        # Birinci (varsayılan) yükleme ayı — geriye uyumluluk için
        ranges = self._yukleme_araliklari(ay, yil)
        return ranges[0] if ranges else (datetime.date.today(), datetime.date.today())

    def _yukleme_araliklari(self, ay, yil):
        """Dönemin yükleme aralıklarını döndür (max 1 aylık parçalar halinde).
        Beyanname dönem AYINDAN SONRA verilir (ör. Mayıs KDV → Haziran).
        Dönem+1 ayından bugüne kadar tüm aylar — boş ay sorgusu yapma, hız için."""
        bugun = datetime.date.today()
        # Dönemden BİR SONRAKİ aydan başla (o ay verilemez, boşa sorgu olmasın)
        araliklar = []
        if ay == 12:
            cur_ay, cur_yil = 1, yil + 1
        else:
            cur_ay, cur_yil = ay + 1, yil
        while True:
            bas = datetime.date(cur_yil, cur_ay, 1)
            if bas > bugun: break
            if cur_ay == 12:
                son = datetime.date(cur_yil, 12, 31)
            else:
                son = datetime.date(cur_yil, cur_ay + 1, 1) - datetime.timedelta(days=1)
            bit = max(bas, min(son, bugun))
            araliklar.append((bas, bit))
            if cur_ay == 12:
                cur_ay, cur_yil = 1, cur_yil + 1
            else:
                cur_ay += 1
            if len(araliklar) >= 8: break  # güvenlik
        return araliklar

    def _periyot(self, kod):
        if kod in ("GGECICI", "KGECICI"):
            return "ceyrek"
        if kod in ("KURUMLAR", "KURUMLARP", "GELIR", "GELIR1001E"):
            return "yil"
        return "ay"

    def _donem_bilgisi(self, kod, ay_str, yil_str):
        """(yukleme_bas, yukleme_bit, indir_filtre[list], özet[str]) döndürür.
        ay_str: aylık→ay(1-12), çeyrek→çeyrek(1-3), yıllık→önemsiz."""
        yil = int(yil_str)
        p = self._periyot(kod)
        bugun = datetime.date.today()

        if p == "ay":
            ay = int(ay_str)
            if not 1 <= ay <= 12:
                raise ValueError("Ay 1-12 arasında olmalı.")
            bas, bit = self._yukleme_araligi(ay, yil)
            return bas, bit, ["--donem", f"{ay:02d}/{yil}"], f"{ay:02d}/{yil}"

        if p == "ceyrek":
            q = int(ay_str)
            if q not in (1, 2, 3, 4):
                raise ValueError("Çeyrek 1-4 olmalı (1=1-3, 2=4-6, 3=7-9, 4=10-12).")
            ceyrek_son = {1: 3, 2: 6, 3: 9, 4: 12}[q]
            don_bas = {1: 1, 2: 4, 3: 7, 4: 10}[q]
            if q == 4:
                deadline = datetime.date(yil + 1, 2, 17)
            else:
                deadline = datetime.date(yil, ceyrek_son + 2, 17)
            bas = deadline - datetime.timedelta(days=23)
            bit = deadline + datetime.timedelta(days=7)
            if bugun >= bas:
                bit = min(bit, bugun)
            return bas, bit, ["--donem", f"{don_bas:02d}/{yil}"], \
                f"{q}. çeyrek ({don_bas:02d}-{ceyrek_son:02d}/{yil})"

        # yıllık (Kurumlar) — ertesi yıl Nisan'da beyan
        bas = datetime.date(yil + 1, 4, 1)
        bit = datetime.date(yil + 1, 4, 30)
        if bugun >= bas:
            bit = min(bit, bugun)
        return bas, bit, ["--yil", str(yil)], f"{yil} yılı (yıllık)"

    def _forma_git(self):
        """Beyanname Ara formuna git. Form ZATEN varsa hiçbir şey yapma (hız)."""
        nav_js = """
        (function(){
          // Form zaten hazırsa dokunma — en hızlı yol
          if (document.querySelector('#taxReturnSearchForm')) return 'hazir';
          var norm = function(s){ return (s||'').replace(/\\s+/g,' ').trim().toLocaleLowerCase('tr'); };
          // Yöntem 1: yedek HTML enjekte et
          try {
            if (typeof beyannameAraFormuHTMLKodu === 'string' && beyannameAraFormuHTMLKodu) {
              var t = document.getElementById('mainWindow_content');
              if (t) { t.innerHTML = beyannameAraFormuHTMLKodu; return 'html_inject'; }
            }
          } catch(e) {}
          // Yöntem 2: menü linki
          var adaylar = Array.from(document.querySelectorAll('a, td, span, li'))
            .filter(function(el){ return norm(el.textContent) === 'beyanname ara'; });
          if (adaylar.length) { adaylar[0].click(); return 'menu_click'; }
          // Yöntem 3: GIB fonksiyonu
          try { if (typeof menuBeyannameAra==='function'){ menuBeyannameAra(); return 'fn'; } } catch(e) {}
          return 'not_found';
        })()
        """
        try:
            sonuc = self._cdp_calistir(nav_js)
            if sonuc == 'hazir':
                return True  # Form zaten orada — bekleme gereksiz
            if sonuc == 'not_found':
                self._cdp_calistir("location.reload()")
                time.sleep(3.0)
                self._cdp_calistir(nav_js)
        except Exception:
            pass
        return False

    def _liste_ac_snippet(self, tur, ay, yil, yukleme_bas, yukleme_bit, durum_kodu="ALL"):
        bas = yukleme_bas
        bit = yukleme_bit
        return f"""
(async () => {{
  const wait = ms => new Promise(r => setTimeout(r, ms));
  const setChecked = (el, val) => {{
    if (!el) return;
    el.checked = !!val;
    el.dispatchEvent(new Event('change', {{bubbles:true}}));
  }};
  const setAllChecked = (selector, val) => {{
    document.querySelectorAll(selector).forEach(el => setChecked(el, val));
  }};
  const setValue = (el, val) => {{
    if (!el) return;
    el.value = String(val);
    el.dispatchEvent(new Event('input', {{bubbles:true}}));
    el.dispatchEvent(new Event('change', {{bubbles:true}}));
  }};
  const byName = name => document.querySelector(`[name="${{name}}"]`);
  const findForm = () => document.querySelector('#taxReturnSearchForm');

  let form = findForm();

  // Yedek HTML varsa enjekte et
  if (!form) {{
    try {{
      if (typeof beyannameAraFormuHTMLKodu === 'string' && beyannameAraFormuHTMLKodu) {{
        const target = document.getElementById('mainWindow_content');
        if (target) {{ target.innerHTML = beyannameAraFormuHTMLKodu; await wait(500); form = findForm(); }}
      }}
    }} catch (e) {{}}
  }}

  if (!form) {{
    throw new Error('Form hazır değil — _forma_git çağrıldıktan sonra tekrar dene.');
  }}

  setAllChecked('input[name="sorguTipiN"], #sorguTipiN', false);
  setAllChecked('input[name="sorguTipiT"], #sorguTipiT', false);
  setAllChecked('input[name="sorguTipiV"], #sorguTipiV', false);
  // Durum filtresi panelden gelir. ALL ise kapalı; diğerlerinde GIB radio seçilir.
  const durumKodu = {json.dumps(durum_kodu)};
  setAllChecked('input[name="sorguTipiD"], #sorguTipiD', durumKodu !== 'ALL');
  setAllChecked('input[name="durum"]', false);
  if (durumKodu !== 'ALL') {{
    const durumRadio = [...document.querySelectorAll('input[name="durum"]')]
      .find(el => String(el.value) === String(durumKodu));
    setChecked(durumRadio, true);
  }}
  setValue(document.querySelector('#vergiNo'), '');
  setValue(document.querySelector('#tcKimlikNo'), '');

  setChecked(document.querySelector('#sorguTipiB'), true);
  const turSelect = document.querySelector('#btbeyannameTanim');
  if (!turSelect) throw new Error('Beyanname türü alanı bulunamadı.');
  const tur = {json.dumps(tur)};
  if (![...turSelect.options].some(o => o.value === tur)) {{
    throw new Error(`Beyanname türü listede yok: ${{tur}}`);
  }}
  setValue(turSelect, tur);

  // Vergilendirme Dönemi SEÇİLMİYOR (GIB hata veriyor). İşareti kaldır.
  setChecked(document.querySelector('#sorguTipiP'), false);

  // Sadece Yükleme Tarih Aralığı ile sorgulanır (Beyanname Türü + isteğe bağlı Durum + tarih).
  setChecked(document.querySelector('#sorguTipiZ'), true);
  setValue(byName('baslangicTarihiGun'), '{bas.day:02d}');
  setValue(byName('baslangicTarihiAy'), '{bas.month:02d}');
  setValue(byName('baslangicTarihiYil'), '{bas.year}');
  setValue(byName('baslangicTarihi'), '{bas:%Y%m%d}');
  setValue(byName('bitisTarihiGun'), '{bit.day:02d}');
  setValue(byName('bitisTarihiAy'), '{bit.month:02d}');
  setValue(byName('bitisTarihiYil'), '{bit.year}');
  setValue(byName('bitisTarihi'), '{bit:%Y%m%d}');

  // GIB'in KENDİ Sorgula akışını çalıştır (elle enjekte ETME!).
  // Böylece liste GIB'in normal sayfalama durumuyla yüklenir ve
  // snippet'in ">>" ile sayfa geçişi düzgün çalışır.
  const _alert = window.alert; window.alert = m => {{ window.__sonAlert = m; }};
  try {{
    if (typeof taxReturnSearchFormPost === 'function') {{
      taxReturnSearchFormPost();
    }} else {{
      const b = document.getElementById('sorgulaButon');
      if (b) b.click();
      else throw new Error('Sorgula düğmesi/fonksiyonu bulunamadı.');
    }}
  }} finally {{
    setTimeout(() => {{ window.alert = _alert; }}, 4000);
  }}

  // Liste düşene kadar bekle (form kaybolup satırlar gelir)
  for (let i = 0; i < 40; i++) {{
    await wait(500);
    if (window.__sonAlert) {{
      const m = window.__sonAlert; window.__sonAlert = '';
      throw new Error('GIB uyarısı: ' + m);
    }}
    const rows = document.querySelectorAll('tr[id^="row"]').length;
    if (rows && !findForm()) {{
      const sayac = [...document.querySelectorAll('td, b, font, span')]
        .map(el => el.textContent || '')
        .find(t => /\\d+\\s*-\\s*\\d+\\s*\\/\\s*\\d+/.test(t));
      return 'Liste açıldı: ' + (sayac ? sayac.trim() : rows + ' satır');
    }}
  }}
  return 'Sorgu gönderildi. Liste düşmediyse Chrome\\'da kontrol et.';
}})()
"""

    # ---------- CSV oluştur (Chrome'dan, CDP) ----------
    def csv_olustur_basla(self):
        if self.calisiyor:
            return
        kod = self._kod()
        snippet_dosya = BASE / ("snippet_muhtasar_sgk.js" if kod == "MUHSGK" else "snippet_kdv.js")
        if not snippet_dosya.exists():
            messagebox.showerror("Snippet yok", f"Bulunamadı:\n{snippet_dosya}")
            return
        try:
            snippet_js = snippet_dosya.read_text(encoding="utf-8")
        except Exception as e:
            messagebox.showerror("Hata", str(e))
            return

        self.calisiyor = True
        self._toplu_buton_durum("disabled")
        self.durdur_btn.configure(state="disabled")
        self.yaz("\n" + "=" * 60 + "\n")
        self.yaz(f"[CSV Oluştur] {self.tur_var.get()} ({kod}) — Chrome'dan veri çekiliyor...\n")
        if kod == "MUHSGK":
            self.yaz("(Muhtasar SGK taraması uzun sürer — sabırla bekle, ilerleme aşağıda akar)\n")
        self.yaz("Not: Chrome'da ilgili beyanname LİSTESİ sayfası açık olmalı.\n")
        self.yaz("=" * 60 + "\n")

        t = threading.Thread(target=self._csv_olustur_thread, args=(snippet_js, kod), daemon=True)
        t.start()

    def _yeni_mukellef_ekle(self, csv_yol):
        """CSV'deki yeni firmaları mükellef Excel'ine SAHİBİ='YENİ' olarak ekle.
        Mevcut (aktif+pasif+yeni) hiçbir firma ile eşleşmeyenleri ekler."""
        try:
            import csv as csvmod
            import pandas as pd
            sys.path.insert(0, str(BASE))
            from indir import klasor_bul

            xl = BASE / "TANER BATTAL MÜKELLEF LİSTESİ.xlsx"
            if not xl.exists():
                return
            df = pd.read_excel(str(xl))
            firma_kol = next((c for c in df.columns
                              if "FİRMA" in str(c).upper() or "FIRMA" in str(c).upper() or "UNVAN" in str(c).upper()),
                             None)
            sahip_kol = next((c for c in df.columns
                              if "SAHİBİ" in str(c).upper() or "SAHIBI" in str(c).upper()),
                             None)
            if not firma_kol or not sahip_kol:
                return

            tum_set = set(str(x).strip().upper() for x in df[firma_kol].dropna()
                          if str(x).strip() and str(x).strip().upper() != "NAN")

            # CSV'deki benzersiz firma adları
            csv_firmalar = []
            gor = set()
            with open(csv_yol, encoding="utf-8-sig") as f:
                for r in csvmod.DictReader(f, delimiter=";"):
                    ad = str(r.get("Ad_Soyad", "") or "").strip().upper()
                    if ad and ad not in gor:
                        gor.add(ad); csv_firmalar.append(ad)

            # klasor_bul-style fuzzy match: tum_set'i taner_set olarak ver
            yeni = []
            for firma in csv_firmalar:
                if klasor_bul(firma, tum_set, set()) is None:
                    yeni.append(firma)

            if not yeni:
                return

            no_kol = next((c for c in df.columns if str(c).upper().startswith("NO")), None)
            max_no = 0
            if no_kol is not None:
                try:
                    max_no = int(pd.to_numeric(df[no_kol], errors="coerce").max())
                except Exception:
                    max_no = len(df)

            yeni_rows = []
            for i, firma in enumerate(yeni, 1):
                row = {firma_kol: firma, sahip_kol: "YENİ"}
                if no_kol is not None:
                    row[no_kol] = max_no + i
                yeni_rows.append(row)
            df_yeni = pd.DataFrame(yeni_rows)
            df_birlesik = pd.concat([df, df_yeni], ignore_index=True)
            try:
                df_birlesik.to_excel(str(xl), index=False)
                self.kuyruk.put(f"[+] {len(yeni)} yeni mükellef eklendi (SAHİBİ='YENİ' — Excel'i kontrol et)\n")
                for f in yeni[:8]:
                    self.kuyruk.put(f"    • {f}\n")
                if len(yeni) > 8:
                    self.kuyruk.put(f"    ... ve {len(yeni)-8} firma daha\n")
            except PermissionError:
                self.kuyruk.put(f"[!] Mükellef Excel açık, yeni firmalar eklenemedi: {xl.name}\n")
        except Exception as e:
            self.kuyruk.put(f"[!] Yeni mükellef ekleme hatası: {e}\n")

    def _csv_olustur_thread(self, snippet_js, kod):
        try:
            csv_metni = self._cdp_calistir(snippet_js)
            if not csv_metni or not str(csv_metni).strip():
                self.kuyruk.put("[HATA] CSV boş döndü. Doğru sayfada mısın / oturum açık mı?\n")
                self.kuyruk.put(("__CSVBITTI__", None))
                return
            ad = "uretilen_muhtasar_sgk.csv" if kod == "MUHSGK" else f"uretilen_{kod.lower()}.csv"
            yol = BASE / ad
            yol.write_text(csv_metni, encoding="utf-8-sig")
            satir = csv_metni.count("\n")
            self.kuyruk.put(f"[✓] CSV kaydedildi ({satir} satır): {yol}\n")
            self._yeni_mukellef_ekle(yol)
            self.kuyruk.put(("__CSVBITTI__", str(yol)))
        except Exception:
            self.kuyruk.put("[HATA] " + traceback.format_exc() + "\n")
            self.kuyruk.put(("__CSVBITTI__", None))

    def _chrome_ws(self):
        with urlopen(f"http://localhost:{DEBUG_PORT}/json", timeout=5) as r:
            hedefler = json.loads(r.read().decode("utf-8", "replace"))
        ws_url = next((t["webSocketDebuggerUrl"] for t in hedefler
                       if t.get("type") == "page" and "gib.gov.tr" in t.get("url", "")), None) \
            or next((t["webSocketDebuggerUrl"] for t in hedefler if t.get("type") == "page"), None)
        if not ws_url:
            raise RuntimeError("Chrome'da açık sayfa bulunamadı. Önce Chrome'u başlat / giriş yap.")
        return ws_url

    def _cdp_calistir(self, snippet_js):
        import websocket  # indir.py ile aynı bağımlılık
        ws = websocket.create_connection(self._chrome_ws(), timeout=20, max_size=None)
        ws.settimeout(1200)  # 20 dk (muhtasar uzun sürebilir)
        # GIB'in createModal içindeki `debugger;` komutu DevTools açıkken snippet'i
        # durdurabiliyor → tüm duraklatmaları atla (anti-debugger bypass).
        ws.send(json.dumps({"id": 1, "method": "Debugger.enable"}))
        ws.send(json.dumps({"id": 2, "method": "Debugger.setSkipAllPauses", "params": {"skip": True}}))
        ws.send(json.dumps({"id": 3, "method": "Runtime.enable"}))
        ws.send(json.dumps({"id": 100, "method": "Runtime.evaluate", "params": {
            "expression": snippet_js, "awaitPromise": True, "returnByValue": True, "userGesture": True
        }}))
        csv_metni = None
        try:
            while True:
                msg = json.loads(ws.recv())
                if msg.get("method") == "Runtime.consoleAPICalled":
                    args = msg.get("params", {}).get("args", [])
                    text = " ".join(str(a.get("value", "")) for a in args if "value" in a)
                    if text.strip():
                        self.kuyruk.put("  " + text + "\n")
                elif msg.get("id") == 100:
                    res = msg.get("result", {})
                    if "exceptionDetails" in res:
                        det = res["exceptionDetails"]
                        raise RuntimeError("Snippet JS hatası: " + str(det.get("exception", det))[:200])
                    csv_metni = res.get("result", {}).get("value")
                    break
        finally:
            try:
                ws.close()
            except Exception:
                pass
        return csv_metni

    # ---------- indirme ----------
    def indir_basla(self):
        if self.calisiyor:
            return
        csv = self.csv_var.get().strip()
        if not csv or not Path(csv).exists():
            messagebox.showwarning("CSV gerekli", "Önce geçerli bir CSV dosyası seç.")
            return
        if not INDIR_PY.exists():
            messagebox.showerror("Hata", f"indir.py bulunamadı:\n{INDIR_PY}")
            return

        kod = self._kod()
        # Dönem filtresi YOK: listede ne varsa hepsi iner; her biri kendi
        # yıl/ay/tür klasörüne gider (yıl klasörü olduğu için karışmaz).
        cmd = [sys.executable, str(INDIR_PY), "--csvpath", csv, "--tur", kod]
        if self.test_var.get():
            test_yol = self.test_yol_var.get().strip()
            if test_yol:
                cmd += ["--cikti", test_yol]
        if self.atla_var.get():
            cmd += ["--atla-mevcut"]

        self.calisiyor = True
        self._dur_iste = False
        self._toplu_buton_durum("disabled")
        self.durdur_btn.configure(state="normal")
        self.yaz("\n" + "=" * 60 + "\n")
        self.yaz(f"[İndirme başladı] tür={self.tur_var.get()}  csv={Path(csv).name}\n")
        self.yaz("=" * 60 + "\n")

        t = threading.Thread(target=self._surec_calistir, args=(cmd,), daemon=True)
        t.start()

    def _surec_calistir(self, cmd):
        try:
            env = dict(os.environ)
            env["PYTHONIOENCODING"] = "utf-8"
            env["PYTHONUTF8"] = "1"
            env["PYTHONUNBUFFERED"] = "1"   # çıktı anlık aksın (tamponlama olmasın)
            self.surec = subprocess.Popen(
                cmd, cwd=str(BASE), stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                stdin=subprocess.DEVNULL, env=env, bufsize=1, text=True,
                encoding="utf-8", errors="replace",
                creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
            )
            for satir in self.surec.stdout:
                self.kuyruk.put(satir)
            self.surec.wait()
            self.kuyruk.put(("__BITTI__", self.surec.returncode))
        except Exception:
            self.kuyruk.put("[HATA] " + traceback.format_exc() + "\n")
            self.kuyruk.put(("__BITTI__", -1))

    def indir_durdur(self):
        self._dur_iste = True   # toplu döngü de dursun
        if self.surec and self.surec.poll() is None:
            try:
                self.surec.terminate()
                self.yaz("\n[Durduruldu]\n")
            except Exception:
                pass

    # ---------- Toplu mod (dönem aralığı: Listeyi Aç + CSV + İNDİR zinciri) ----------
    def _donem_listesi(self, ba, by, ea, ey, periyot="ay"):
        res = []
        if periyot == "ceyrek":
            q, y = ba, by
            while (y, q) <= (ey, ea) and len(res) < 20:
                res.append((q, y))
                q += 1
                if q > 4:
                    q = 1
                    y += 1
        elif periyot == "yil":
            for y in range(by, ey + 1):
                res.append((1, y))
                if len(res) >= 20:
                    break
        else:
            y, m = by, ba
            while (y, m) <= (ey, ea) and len(res) < 60:
                res.append((m, y))
                m += 1
                if m > 12:
                    m = 1
                    y += 1
        return res

    def sablon_yenile(self):
        """Tüm uretilen_*.csv dosyalarını tarayıp MUKELLEF VERGI ŞABLONU.xlsx'i yeniden oluştur."""
        if self.calisiyor:
            messagebox.showinfo("Meşgul", "Başka bir işlem çalışıyor.")
            return
        try:
            cmd = [sys.executable, str(BASE / "sablon_olustur.py")]
            r = subprocess.run(cmd, cwd=str(BASE), capture_output=True, text=True,
                             encoding="utf-8", errors="replace",
                             creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0))
            self.yaz("\n[ŞABLON YENİLE]\n")
            self.yaz(r.stdout)
            if r.stderr:
                self.yaz("[stderr] " + r.stderr)
            if r.returncode == 0:
                self.yaz("[✓] Şablon yenilendi: MUKELLEF VERGI ŞABLONU.xlsx\n")
                try:
                    os.startfile(str(BASE / "MUKELLEF VERGI ŞABLONU.xlsx"))
                except Exception: pass
        except Exception as e:
            messagebox.showerror("Hata", f"Şablon yenileme hatası:\n{e}")

    @staticmethod
    def _sablonu_oku(tur, ay, yil):
        """MUKELLEF VERGI ŞABLONU.xlsx'i oku, belirli tür+dönem için BEKLENEN firmaları döndür.
        Dönüş: (beklenen_set, periyot_str)
        Periyot: 'AYLIK' veya '3 AYLIK' veya 'YILLIK' — gerçek kontrol için.
        """
        import pandas as pd
        sab = BASE / "MUKELLEF VERGI ŞABLONU.xlsx"
        if not sab.exists():
            return None, None
        try:
            df = pd.read_excel(str(sab))
        except Exception:
            return None, None

        kolonlar = {str(c).strip().upper(): c for c in df.columns}
        firma_kol = (kolonlar.get("FİRMA ÜNVANI") or kolonlar.get("FIRMA UNVANI")
                     or kolonlar.get("FIRMA ÜNVANI"))
        if firma_kol is None: return None, None

        # Tür → şablon sütun adı
        tur_upper = (tur or "").upper().strip()
        tur_kol = None
        if tur_upper == "POSET":
            tur_kol = kolonlar.get("GEKAP") or kolonlar.get("POSET")
        elif tur_upper in kolonlar:
            tur_kol = kolonlar[tur_upper]
        if tur_kol is None: return None, None

        # Periyot belirleme — türe göre
        periyot = "AYLIK"
        if tur_upper in ("GGECICI", "KGECICI", "POSET"):
            periyot = "3 AYLIK"
        elif tur_upper == "KURUMLAR":
            periyot = "YILLIK"
        elif tur_upper == "MUHSGK":
            # MUHSGK için: ay 1,4,7,10 ise hem AYLIK hem 3 AYLIK firmalar
            # (3 AYLIK firmaların 1-3 dönemi 03/yıl tek ay olarak gelmez,
            #  ama biz şablona göre ayırırız)
            periyot = "BOTH"

        beklenen = set()
        for _, row in df.iterrows():
            firma = str(row.get(firma_kol, "")).strip().upper()
            if not firma or firma == "NAN": continue
            deger = str(row.get(tur_kol, "")).strip().upper()
            if not deger or deger == "NAN": continue
            if "YOK" in deger or "KDV YOK" in deger: continue  # muafiyet

            if tur_upper == "MUHSGK":
                if "3 AYLIK" in deger or "ÜÇ" in deger:
                    beklenen.add((firma, "3 AYLIK"))
                elif "AYLIK" in deger:
                    beklenen.add((firma, "AYLIK"))
                else:
                    beklenen.add((firma, "AYLIK"))  # default
            else:
                if deger in ("✓", "VAR", "EVET", "+", "X"):
                    beklenen.add((firma, periyot))
                else:
                    # Periyot belirten herhangi bir değer
                    beklenen.add((firma, periyot))

        return beklenen, periyot

    def kontrol_listesi(self):
        """Seçili tür+dönem için GIB'den canlı veri çekip Excel kontrol listesi üret.
        1) Listeyi Aç (GIB'de form doldur+Sorgula)
        2) Snippet çalıştır → CSV
        3) Mükellef listesi ile karşılaştır
        4) ✅ Yeşil = verildi, ❌ Kırmızı = verilmedi"""
        if self.calisiyor:
            messagebox.showinfo("Meşgul", "Başka bir işlem çalışıyor, bitmesini bekle.")
            return
        try:
            sys.path.insert(0, str(BASE))
            from openpyxl import Workbook
            from indir import mukellef_listesi_yukle
        except Exception as e:
            messagebox.showerror("Hata", f"Modüller yüklenemedi:\n{e}")
            return

        tur = self._kod()
        try:
            ay = int(self.ay_var.get().strip())
            yil = int(self.yil_var.get().strip())
            if not (1 <= ay <= 12 and 2000 <= yil <= 2100):
                raise ValueError
        except Exception:
            messagebox.showwarning("Dönem", "Üstteki Dönem alanına geçerli ay/yıl gir (örn: 5/2025).")
            return

        # Kontrol listesinde HAFİF snippet kullan (SGK bilgisi gerekmiyor, çok hızlı)
        snippet_dosya = BASE / "snippet_kdv.js"
        if not snippet_dosya.exists():
            messagebox.showerror("Snippet yok", f"Bulunamadı:\n{snippet_dosya}")
            return
        try:
            snippet_js = snippet_dosya.read_text(encoding="utf-8")
        except Exception as e:
            messagebox.showerror("Hata", str(e))
            return

        self.calisiyor = True
        self._dur_iste = False
        self._toplu_buton_durum("disabled")
        self.yaz("\n" + "=" * 60 + "\n")
        self.yaz(f"[KONTROL LİSTESİ] {tur} — {ay:02d}/{yil}  (GIB'den canlı çekiliyor — hafif mod)\n")
        self.yaz("=" * 60 + "\n")
        t = threading.Thread(target=self._kontrol_listesi_thread,
                             args=(tur, ay, yil, snippet_js), daemon=True)
        t.start()

    def _kontrol_listesi_thread(self, tur, ay, yil, snippet_js):
        import csv as csvmod
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from indir import mukellef_listesi_yukle, _norm, _tr, _kw, _sim
            import re
        except Exception as e:
            self.kuyruk.put(f"[HATA] Modül import: {e}\n")
            self.kuyruk.put(("__KONTROLBITTI__", None))
            return

        try:
            # 1) Listeyi Aç — TÜM yükleme aralıklarını sorgula (geç yüklemeler dahil)
            try:
                araliklar = self._yukleme_araliklari(ay, yil)
            except Exception:
                import datetime
                son_gun = 28 if ay == 2 else 30 if ay in (4,6,9,11) else 31
                araliklar = [(datetime.date(yil, ay, 1), datetime.date(yil, ay, son_gun))]

            # ── ADIM 1: Disk CSV'lerini yükle (güvenilir, hızlı) ──────────────
            csv_satirlari = []
            csvyol = BASE / f"uretilen_{tur.lower()}_{yil}{ay:02d}.csv"
            if csvyol.exists():
                try:
                    with open(csvyol, encoding="utf-8-sig") as f:
                        csv_satirlari = list(csvmod.DictReader(f, delimiter=";"))
                    self.kuyruk.put(f"[✓] Disk CSV: {csvyol.name} — {len(csv_satirlari)} satır\n")
                except Exception as e:
                    self.kuyruk.put(f"[!] Disk CSV okuma: {e}\n")

            # ── ADIM 2: GIB canlı sorgu (opsiyonel, hata olursa disk verisiyle devam) ──
            # GIB CANLI SORGU HER ZAMAN ÇALIŞIR — disk CSV sadece taban veri.
            # (Disk eski kalabilir: bugün yazılmış olsa bile içi birkaç gün öncesine ait
            #  olabilir. En güncel beyannameleri yakalamak için GIB'i mutlaka sorgula.)
            import datetime as _dt2
            mevcut_oidler = {(r.get("BynOID","") or "").strip() for r in csv_satirlari}
            gib_yeni = []
            if True:
                for idx, (bas, bit) in enumerate(araliklar, 1):
                    if self._dur_iste: break
                    self.kuyruk.put(f"[GIB Sorgu {idx}/{len(araliklar)}] {bas:%d/%m/%Y} → {bit:%d/%m/%Y}\n")
                    try:
                        hazir = self._forma_git()
                        if not hazir: time.sleep(0.8)
                        self._cdp_calistir(self._liste_ac_snippet(tur, ay, yil, bas, bit))
                        time.sleep(0.6)
                        tablo_js = "(function(){return document.querySelectorAll('tr[id^=\"row\"]').length;})()"
                        for _ in range(15):
                            n = self._cdp_calistir(tablo_js)
                            try: n = int(n or 0)
                            except: n = 0
                            if n > 0: break
                            time.sleep(0.3)
                        time.sleep(0.3)
                        csv_metni = self._cdp_calistir(snippet_js)
                        if csv_metni and len(csv_metni.splitlines()) > 1:
                            yeni = list(csvmod.DictReader(csv_metni.splitlines(), delimiter=";"))
                            for r in yeni:
                                oid = (r.get("BynOID","") or "").strip()
                                if oid not in mevcut_oidler:
                                    gib_yeni.append(r)
                                    if oid: mevcut_oidler.add(oid)
                            self.kuyruk.put(f"  ✓ +{len(yeni)} GIB satır ({len(gib_yeni)} yeni)\n")
                    except Exception as e:
                        self.kuyruk.put(f"  [!] GIB sorgu atlandı: {e}\n")

            if gib_yeni:
                csv_satirlari.extend(gib_yeni)
                self.kuyruk.put(f"[✓] GIB'den {len(gib_yeni)} yeni kayıt eklendi\n")
                # Güncellenmiş CSV'yi diske yaz
                try:
                    baslik = "Ad_Soyad;VK_No;Beyanname_Turu;Vergi_Dairesi;Vergilendirme_Donemi;Sube_No;Yukleme_Zamani;Durum;SGK_Durum;BynOID;ThkOID;IhbOID;Klasor"
                    with open(csvyol, "w", encoding="utf-8-sig", newline="") as f:
                        f.write(baslik + "\r\n")
                        for r in csv_satirlari:
                            f.write(";".join(f'"{str(r.get(k,"")).replace(chr(34), chr(34)+chr(34))}"'
                                            for k in baslik.split(";")) + "\r\n")
                    self._yeni_mukellef_ekle(csvyol)
                except Exception as e:
                    self.kuyruk.put(f"[!] CSV yazma: {e}\n")

            self.kuyruk.put(f"[i] Toplam: {len(csv_satirlari)} kayıt (disk + GIB)\n")

            # Sadece istenen tür kayıtları al (dönem filtresi alta taşındı)
            csv_satirlari = [
                r for r in csv_satirlari
                if str(r.get("Beyanname_Turu", "") or "").strip().upper() == tur
            ]

            # MUHSGK/GGECICI/KGECICI için: GEÇMİŞ TÜM ÇEYREKLERİ sorgula
            ceyrek_kayitlari = []
            ucaylik_turler = ("MUHSGK","GGECICI","KGECICI")
            if tur.upper() in ucaylik_turler:
                import datetime as _dt
                ceyrek_no = (ay - 1) // 3 + 1
                for q in range(1, ceyrek_no + 1):
                    if self._dur_iste: break
                    q_bas_ay = {1:1, 2:4, 3:7, 4:10}[q]
                    q_bit_ay = {1:3, 2:6, 3:9, 4:12}[q]
                    donem_bas_str = f"{q_bas_ay:02d}/{yil}"
                    donem_bit_str = f"{q_bit_ay:02d}/{yil}"

                    # Önce disk CSV'lerinden Q verilerini al (hızlı)
                    q_disk = []
                    for csv_f in sorted(BASE.glob(f"uretilen_{tur.lower()}_{yil}*.csv")):
                        try:
                            with open(csv_f, encoding="utf-8-sig") as f:
                                for r in csvmod.DictReader(f, delimiter=";"):
                                    vd = str(r.get("Vergilendirme_Donemi","") or "")
                                    if donem_bas_str in vd and donem_bit_str in vd:
                                        oid = (r.get("BynOID","") or "").strip()
                                        if oid not in mevcut_oidler:
                                            q_disk.append(r)
                                            if oid: mevcut_oidler.add(oid)
                        except: pass
                    if q_disk:
                        ceyrek_kayitlari.extend(q_disk)
                        self.kuyruk.put(f"[i] Q{q} disk: {len(q_disk)} kayıt\n")
                        continue  # Disk'te bulunduysa GIB sorgusuna gerek yok

                    # Disk'te yoksa GIB'e sor
                    upload_ay = q_bit_ay + 1 if q_bit_ay < 12 else 1
                    upload_yil = yil + 1 if q_bit_ay == 12 else yil
                    q_bas = _dt.date(upload_yil, upload_ay, 1)
                    son_g = 28 if upload_ay == 2 else 30 if upload_ay in (4,6,9,11) else 31
                    q_bit = _dt.date(upload_yil, upload_ay, son_g)
                    self.kuyruk.put(f"[i] Q{q} GIB sorgu {q_bas:%d/%m/%Y}→{q_bit:%d/%m/%Y}\n")
                    try:
                        hazir = self._forma_git()
                        if not hazir: time.sleep(0.8)
                        self._cdp_calistir(self._liste_ac_snippet(tur, ay, yil, q_bas, q_bit))
                        time.sleep(0.6)
                        tablo_js = "(function(){return document.querySelectorAll('tr[id^=\"row\"]').length;})()"
                        for _ in range(15):
                            n = self._cdp_calistir(tablo_js)
                            try: n = int(n or 0)
                            except: n = 0
                            if n > 0: break
                            time.sleep(0.3)
                        time.sleep(0.3)
                        q_csv = self._cdp_calistir(snippet_js)
                        if q_csv and str(q_csv).strip():
                            q_rows = list(csvmod.DictReader(q_csv.splitlines(), delimiter=";"))
                            q_filtreli = [r for r in q_rows
                                if str(r.get("Beyanname_Turu","") or "").strip().upper() == tur
                                and donem_bas_str in str(r.get("Vergilendirme_Donemi","") or "")
                                and donem_bit_str in str(r.get("Vergilendirme_Donemi","") or "")]
                            ceyrek_kayitlari.extend(q_filtreli)
                            self.kuyruk.put(f"[i] Q{q} GIB: {len(q_filtreli)} kayıt\n")
                    except Exception as e:
                        self.kuyruk.put(f"[!] Q{q} GIB atlandı: {e}\n")

            # Aylık dönem filtresi: TAM aylık (XX/YYYY-XX/YYYY, başlangıç=bitiş)
            donem_str = f"{ay:02d}/{yil}"
            aylik_donem_tam = f"{donem_str}-{donem_str}"  # ör: "03/2025-03/2025"
            yil_str = str(yil)
            aylik_kayitlari = []
            csv_3aylik = []
            for r in csv_satirlari:
                vd = str(r.get("Vergilendirme_Donemi", "") or "").strip()
                if aylik_donem_tam in vd:
                    aylik_kayitlari.append(r)
                elif tur.upper() in ("MUHSGK","GGECICI","KGECICI") and "-" in vd and yil_str in vd:
                    # 3 aylık dönem aynı yıl içinde
                    csv_3aylik.append(r)
            ceyrek_kayitlari.extend(csv_3aylik)
            self.kuyruk.put(f"[i] {tur} {donem_str} aylık: {len(aylik_kayitlari)} | CSV içi 3 aylık: {len(csv_3aylik)}\n")

            # Verildi = Onaylandı VEYA Onay bekliyor (Hatalı/İptal hariç)
            def verildi_mi(r):
                d = str(r.get("Durum","")).upper().strip()
                if "İPTAL" in d or "IPTAL" in d: return False
                if "HATA" in d: return False
                return True  # Onaylandı, Onay bekliyor, vb.
            def hatali_mi(r):
                return "HATA" in str(r.get("Durum","")).upper()

            verilen_aylik = [r for r in aylik_kayitlari if verildi_mi(r)]
            verilen_ceyrek = [r for r in ceyrek_kayitlari if verildi_mi(r)]
            hatali_aylik = [r for r in aylik_kayitlari if hatali_mi(r)]
            hatali_ceyrek = [r for r in ceyrek_kayitlari if hatali_mi(r)]

            from collections import Counter
            durum_dagilimi = Counter(str(r.get("Durum","")).strip() for r in aylik_kayitlari + ceyrek_kayitlari)
            self.kuyruk.put(f"[i] Durum dağılımı: {dict(durum_dagilimi)}\n")
            self.kuyruk.put(f"[i] Verildi (aylık+3aylık): {len(verilen_aylik)+len(verilen_ceyrek)} | Hatalı: {len(hatali_aylik)+len(hatali_ceyrek)}\n")

            csv_adlari = [str(r.get("Ad_Soyad", "")).strip() for r in verilen_aylik if r.get("Ad_Soyad")]
            ceyrek_adlari = [str(r.get("Ad_Soyad", "")).strip() for r in verilen_ceyrek if r.get("Ad_Soyad")]
            # Hatalı firma adları (ayrı rapor için)
            hatali_adlari = [str(r.get("Ad_Soyad", "")).strip() for r in hatali_aylik + hatali_ceyrek if r.get("Ad_Soyad")]
            # Durum + yükleme zamanı eşleştirme için
            firma_durum = {}
            for r in aylik_kayitlari + ceyrek_kayitlari:
                ad = str(r.get("Ad_Soyad","")).strip()
                durum = str(r.get("Durum","")).strip()
                if ad and (ad not in firma_durum or "ONAYLAN" in durum.upper()):
                    firma_durum[ad] = durum

            # Mükellef listesini yükle
            try:
                taner_set, omer_set = mukellef_listesi_yukle(BASE)
            except SystemExit:
                self.kuyruk.put("[HATA] Mükellef listesi yüklenemedi (TANER BATTAL MÜKELLEF LİSTESİ.xlsx)\n")
                self.kuyruk.put(("__KONTROLBITTI__", None))
                return

            # ŞABLON kullanılabilir mi?
            beklenen_set, _periyot = self._sablonu_oku(tur, ay, yil)
            sablon_aktif = beklenen_set is not None and len(beklenen_set) > 0
            # Hukuki ek (LTD.ŞTİ./A.Ş./ADİ ORT.) temizle — şablonda ekli, mükellef listesinde olmayabilir
            def _ek_temizle(ad):
                a = _norm(ad)
                for ek in ("LTDSTI","LTD","ANONIMSIRKETI","AS","ADIORT","ADIORTAKLIGI","STI","SIRKETI","ANONIM"):
                    if a.endswith(ek): a = a[:-len(ek)]
                return a

            if sablon_aktif:
                self.kuyruk.put(f"[i] Şablondan {len(beklenen_set)} 'beklenen' firma alındı\n")
                # Beklenen firmaları ek-temizlenmiş normalize ile tut (ek farkı eşleşmeyi bozmasın)
                beklenen_norm = set()
                for f, p in beklenen_set:
                    beklenen_norm.add(_ek_temizle(f))
                # Set'leri şablona göre filtrele: ek-temizlenmiş eşleşme
                taner_set = {f for f in taner_set if _ek_temizle(f) in beklenen_norm}
                omer_set  = {f for f in omer_set  if _ek_temizle(f) in beklenen_norm}
                self.kuyruk.put(f"[i] Bu dönemde beklenen: TANER {len(taner_set)}, ÖMER {len(omer_set)}\n")
            else:
                # Şablon yoksa MUAF filtresine düş
                try:
                    from indir import mukellef_muaf_set, _norm as _n
                    muaf_norm = mukellef_muaf_set(BASE, tur)
                except Exception:
                    muaf_norm = set()
                def _muaf_mi(firma):
                    if not muaf_norm: return False
                    fn = _n(firma)
                    if fn in muaf_norm: return True
                    return any(m and (m in fn or fn in m) for m in muaf_norm)
                once_t, once_o = len(taner_set), len(omer_set)
                taner_set = {f for f in taner_set if not _muaf_mi(f)}
                omer_set  = {f for f in omer_set  if not _muaf_mi(f)}
                muaf_atlanan = (once_t - len(taner_set)) + (once_o - len(omer_set))
                if muaf_atlanan:
                    self.kuyruk.put(f"[i] {tur} için MUAF olan {muaf_atlanan} firma listeden çıkarıldı\n")
                self.kuyruk.put(f"[!] Şablon henüz yok — mükellef listesinden gidiyor. Daha doğru sonuç için: '📊 Şablonu Yenile' bas.\n")

            # klasor_bul mantığıyla, her mükellef için "verdi mi?" kontrolü
            def eslesiyor_mu(mukellef, csv_ad):
                mn = _norm(mukellef); hn = _norm(csv_ad)
                if not mn or not hn: return False
                if mn in hn or hn in mn: return True
                hw = set(re.findall(r'[A-Z0-9]+', _tr(csv_ad)))
                kws = _kw(mukellef)
                if len(kws) >= 2:
                    e = sum(1 for kw in kws[:2] if kw in hw or any(len(kw) > 3 and _sim(kw, h) >= 0.82 for h in hw))
                    return e >= 2
                return False

            def veren_mi(mukellef, csv_listesi):
                return any(eslesiyor_mu(mukellef, ad) for ad in csv_listesi)

            # Firma → Yükleme Zamanı dict (en son yüklenmiş)
            yukleme_zamani = {}
            for kayit_listesi in [verilen_aylik, verilen_ceyrek]:
                for r in kayit_listesi:
                    ad = str(r.get("Ad_Soyad", "") or "").strip()
                    zaman = str(r.get("Yukleme_Zamani", "") or "").strip()
                    if ad and zaman:
                        # Aynı firma için en son zaman tut
                        if ad not in yukleme_zamani or zaman > yukleme_zamani[ad]:
                            yukleme_zamani[ad] = zaman

            def yukleme_bul(mukellef):
                for ad, zaman in yukleme_zamani.items():
                    if eslesiyor_mu(mukellef, ad):
                        return zaman
                return ""

            # Firma → GIB'deki TAM ünvan (LTD.ŞTİ. / A.Ş. ekini çekmek için)
            tum_gib_adlar = set()
            for r in (aylik_kayitlari + ceyrek_kayitlari + hatali_aylik + hatali_ceyrek):
                ad = str(r.get("Ad_Soyad", "") or "").strip()
                if ad: tum_gib_adlar.add(ad)

            # Şablondaki kurated tam ünvanlar (kullanıcı elle düzeltti — öncelikli)
            sablon_unvan_exact = {}  # tam normalize -> ünvan (çakışmayı önler)
            sablon_unvan = {}        # ek-temizlenmiş -> ünvan (yedek)
            try:
                import pandas as _pd
                _dfs = _pd.read_excel(str(BASE / "MUKELLEF VERGI ŞABLONU.xlsx"),
                                      sheet_name="Mükellef Vergi Şablonu")
                _fk = None
                for _c in _dfs.columns:
                    if str(_c).strip().upper() in ("FİRMA ÜNVANI", "FIRMA UNVANI", "FIRMA ÜNVANI"):
                        _fk = _c; break
                if _fk is not None:
                    for _v in _dfs[_fk].astype(str):
                        _v = _v.strip()
                        if _v and _v != "nan":
                            sablon_unvan_exact[_norm(_v)] = _v
                            sablon_unvan.setdefault(_ek_temizle(_v), _v)
            except Exception:
                pass

            def unvan_ekli(mukellef):
                """Görüntülenecek ünvan: önce TAM eşleşme (çakışma önler), sonra ek-temizlenmiş."""
                tn = sablon_unvan_exact.get(_norm(mukellef))
                if tn: return tn
                tn = sablon_unvan.get(_ek_temizle(mukellef))
                if tn: return tn
                return mukellef

            # Firma → Hata durumu (hatalı beyanname var mı, hangi tarih)
            def hata_bul(mukellef):
                for r in (hatali_aylik + hatali_ceyrek):
                    ad = str(r.get("Ad_Soyad", "") or "").strip()
                    if ad and eslesiyor_mu(mukellef, ad):
                        zaman = str(r.get("Yukleme_Zamani", "") or "").strip()
                        return f"⚠ HATALI {zaman}".strip()
                return ""

            # TEK EXCEL: MUKELLEF VERGI ŞABLONU.xlsx içine yeni sayfa ekle
            from openpyxl import load_workbook
            sablon_yolu = BASE / "MUKELLEF VERGI ŞABLONU.xlsx"
            if not sablon_yolu.exists():
                self.kuyruk.put("[HATA] MUKELLEF VERGI ŞABLONU.xlsx yok. Önce '📊 Şablonu Yenile' bas.\n")
                self.kuyruk.put(("__KONTROLBITTI__", None))
                return
            try:
                wb = load_workbook(sablon_yolu)
            except PermissionError:
                self.kuyruk.put(f"[HATA] Şablon açık olabilir, kapat ve tekrar dene: {sablon_yolu.name}\n")
                self.kuyruk.put(("__KONTROLBITTI__", None))
                return

            sayfa_adi = f"{tur} {ay:02d}-{yil}"
            if sayfa_adi in wb.sheetnames:
                del wb[sayfa_adi]
            ws = wb.create_sheet(title=sayfa_adi)

            baslik_fill = PatternFill("solid", fgColor="1F2937")
            baslik_font = Font(bold=True, color="FFFFFF", size=11)
            yesil_fill  = PatternFill("solid", fgColor="C6EFCE")
            yesil_font  = Font(color="006100", bold=True)
            mavi_fill   = PatternFill("solid", fgColor="DBEAFE")
            mavi_font   = Font(color="0050C0", bold=True)
            kirmizi_fill = PatternFill("solid", fgColor="FFC7CE")
            kirmizi_font = Font(color="9C0006", bold=True)
            merkez = Alignment(horizontal="center", vertical="center")
            sol = Alignment(horizontal="left", vertical="center")
            kenar = Border(*[Side(style="thin", color="999999")] * 4)

            # Sütun yapısı: Sahip, Firma, Durum, [Dönem], Yükleme Zamanı, Hata Durumu
            is_muhsgk = (tur.upper() in ("MUHSGK","GGECICI","KGECICI"))
            son_kol = 6 if is_muhsgk else 5
            from openpyxl.utils import get_column_letter
            son_harf = get_column_letter(son_kol)

            ws.merge_cells(f"A1:{son_harf}1")
            ws["A1"] = f"{tur} KONTROL LİSTESİ — {ay:02d}/{yil}"
            ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
            ws["A1"].fill = baslik_fill
            ws["A1"].alignment = merkez
            ws.row_dimensions[1].height = 28

            basliklar_satiri = ["Sahip", "Firma Ünvanı", "Durum"]
            if is_muhsgk: basliklar_satiri.append("Dönem")
            basliklar_satiri.append("Yükleme Zamanı")
            basliklar_satiri.append("Hata Durumu")
            for c, h in enumerate(basliklar_satiri, 1):
                cell = ws.cell(row=2, column=c, value=h)
                cell.fill = baslik_fill; cell.font = baslik_font
                cell.alignment = merkez; cell.border = kenar
            ws.row_dimensions[2].height = 22

            # Önce tüm satırları topla (sıralama için)
            verdi_sayi = 0; vermedi_sayi = 0; ucaylik_sayi = 0
            kayitlar = []
            for sahip_kisa, mset in [("TANER", taner_set), ("ÖMER", omer_set)]:
                for firma in sorted(mset):
                    verdi_aylik = veren_mi(firma, csv_adlari)
                    verdi_3aylik = is_muhsgk and veren_mi(firma, ceyrek_adlari)
                    tur_up = tur.upper()
                    sadece_3aylik = tur_up in ("GGECICI","KGECICI")
                    firma_gercek_durum = ""
                    for ad, d in firma_durum.items():
                        if eslesiyor_mu(firma, ad):
                            firma_gercek_durum = d; break

                    if verdi_aylik and not sadece_3aylik:
                        ek = f" ({firma_gercek_durum})" if firma_gercek_durum and "ONAYLAN" not in firma_gercek_durum.upper() else ""
                        durum, donem = f"✅ Verildi{ek}", "AYLIK"
                        fill, font = (yesil_fill, yesil_font) if "ONAYLAN" in firma_gercek_durum.upper() else (mavi_fill, mavi_font)
                        verdi_sayi += 1; sira = 0
                    elif verdi_3aylik or (sadece_3aylik and verdi_aylik):
                        ek = f" ({firma_gercek_durum})" if firma_gercek_durum and "ONAYLAN" not in firma_gercek_durum.upper() else ""
                        durum, donem = f"✅ Verildi{ek}", "3 AYLIK"
                        fill, font = mavi_fill, mavi_font
                        ucaylik_sayi += 1; sira = 0
                    else:
                        donem = "3 AYLIK" if is_muhsgk else ""
                        durum = "❌ Verilmedi"
                        fill, font = kirmizi_fill, kirmizi_font
                        vermedi_sayi += 1; sira = 1

                    kayitlar.append({
                        "sira": sira, "sk": sahip_kisa, "fg": unvan_ekli(firma),
                        "durum": durum, "donem": donem, "fill": fill, "font": font,
                        "zaman": yukleme_bul(firma) if (verdi_aylik or verdi_3aylik) else "",
                        "hata": hata_bul(firma),
                    })

            # AKILLI SIRALAMA: Verildi (yeşil) üstte, Verilmedi (kırmızı) altta; içinde sahip+ünvan
            kayitlar.sort(key=lambda k: (k["sira"], k["sk"], k["fg"]))

            row = 3
            for k in kayitlar:
                a = ws.cell(row=row, column=1, value=k["sk"])
                b = ws.cell(row=row, column=2, value=k["fg"])
                c = ws.cell(row=row, column=3, value=k["durum"])
                a.alignment = merkez; b.alignment = sol; c.alignment = merkez
                a.border = b.border = c.border = kenar
                a.fill = b.fill = c.fill = k["fill"]; c.font = k["font"]
                if is_muhsgk:
                    d = ws.cell(row=row, column=4, value=k["donem"])
                    d.alignment = merkez; d.border = kenar; d.fill = k["fill"]; d.font = k["font"]
                    yz_col = 5
                else:
                    yz_col = 4
                yz = ws.cell(row=row, column=yz_col, value=k["zaman"])
                yz.alignment = merkez; yz.border = kenar; yz.fill = k["fill"]
                yz.font = Font(color="333333", size=10)
                hc = ws.cell(row=row, column=yz_col+1, value=k["hata"])
                hc.alignment = merkez; hc.border = kenar
                if k["hata"]:
                    hc.fill = PatternFill("solid", fgColor="FFE4B5")
                    hc.font = Font(color="B45309", bold=True, size=10)
                else:
                    hc.fill = k["fill"]; hc.font = Font(color="333333", size=10)
                row += 1

            son_veri_satir = row - 1
            ws.cell(row=row, column=1, value="TOPLAM").font = Font(bold=True)
            ozet = f"Verdi: {verdi_sayi}"
            if is_muhsgk: ozet += f"  |  3 Aylık: {ucaylik_sayi}"
            ozet += f"  |  Vermedi: {vermedi_sayi}  |  Toplam: {verdi_sayi+ucaylik_sayi+vermedi_sayi}"
            ws.cell(row=row, column=2, value=ozet).font = Font(bold=True)

            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 60
            ws.column_dimensions["C"].width = 16
            if is_muhsgk:
                ws.column_dimensions["D"].width = 12
                ws.column_dimensions["E"].width = 22
                ws.column_dimensions["F"].width = 26
            else:
                ws.column_dimensions["D"].width = 22
                ws.column_dimensions["E"].width = 26
            ws.freeze_panes = "A3"
            # AKILLI FİLTRE: başlık satırına otomatik filtre (renge/değere göre filtrele-sırala)
            from openpyxl.utils import get_column_letter as _gcl
            ws.auto_filter.ref = f"A2:{_gcl(son_kol)}{son_veri_satir}"

            # Yeni sayfayı önce göster (aktif sekme yap)
            wb.active = wb.sheetnames.index(sayfa_adi)

            try:
                wb.save(sablon_yolu)
            except PermissionError:
                self.kuyruk.put(f"[HATA] Şablon açık olabilir, kapat ve tekrar dene: {sablon_yolu.name}\n")
                self.kuyruk.put(("__KONTROLBITTI__", None))
                return

            # ÖZ-DOĞRULAMA: Şablonda BEKLENEN ama CSV'de OLMAYAN firmaları log'a yaz
            self.kuyruk.put("\n[ÖZ-DOĞRULAMA]\n")
            try:
                beklenenler_norm = set()
                for f in (taner_set | omer_set):
                    beklenenler_norm.add(_norm(f))
                # CSV'de bulunan firmaları normalize et
                csv_bulunanlar = set()
                for ad in csv_adlari + (ceyrek_adlari if is_muhsgk else []):
                    if ad: csv_bulunanlar.add(_norm(ad))
                eksikler = []
                for f in sorted(taner_set | omer_set):
                    fn = _norm(f)
                    if not any(fn in cn or cn in fn for cn in csv_bulunanlar):
                        eksikler.append(f)
                if eksikler:
                    self.kuyruk.put(f"  ⚠ {len(eksikler)} firma şablonda var ama CSV'de yok:\n")
                    for f in eksikler[:15]:
                        self.kuyruk.put(f"    - {f}\n")
                    if len(eksikler) > 15:
                        self.kuyruk.put(f"    ... ve {len(eksikler)-15} firma daha\n")
                else:
                    self.kuyruk.put("  ✓ Tüm beklenen firmalar CSV'de bulundu\n")
            except Exception as e:
                self.kuyruk.put(f"  [!] Doğrulama hatası: {e}\n")

            self.kuyruk.put(f"\n[KONTROL] Sayfa: '{sayfa_adi}' (içine eklendi)\n")
            self.kuyruk.put(f"  ✅ Verdi:    {verdi_sayi}\n")
            self.kuyruk.put(f"  ❌ Vermedi:  {vermedi_sayi}\n")
            self.kuyruk.put(f"  Toplam:     {verdi_sayi+vermedi_sayi}\n")
            try:
                os.startfile(str(sablon_yolu))
            except Exception as e:
                self.kuyruk.put(f"[i] Excel açma hatası: {e}\n")
            self.kuyruk.put(("__KONTROLBITTI__", str(sablon_yolu)))
        except Exception:
            self.kuyruk.put("[HATA] " + traceback.format_exc() + "\n")
            self.kuyruk.put(("__KONTROLBITTI__", None))

    def toplu_basla(self):
        if self.calisiyor:
            return
        tur = self._kod()
        periyot = self._periyot(tur)
        try:
            ba, by = int(self.toplu_bas_ay.get()), int(self.toplu_bas_yil.get())
            ea, ey = int(self.toplu_bit_ay.get()), int(self.toplu_bit_yil.get())
            if periyot == "ceyrek":
                if not (1 <= ba <= 4 and 1 <= ea <= 4 and 2000 <= by <= 2100 and 2000 <= ey <= 2100):
                    raise ValueError("Çeyrek 1-4, yıl 2025 gibi olmalı.")
            elif periyot == "yil":
                if not (2000 <= by <= 2100 and 2000 <= ey <= 2100):
                    raise ValueError("Yıl 2025 gibi olmalı.")
            else:
                if not (1 <= ba <= 12 and 1 <= ea <= 12 and 2000 <= by <= 2100 and 2000 <= ey <= 2100):
                    raise ValueError("Ay 1-12, yıl 2025 gibi olmalı.")
        except ValueError as ve:
            msg = str(ve) if str(ve) else ("Çeyrek 1-3" if periyot == "ceyrek" else "Ay 1-12, yıl 2025 gibi olmalı.")
            messagebox.showwarning("Aralık hatalı", msg)
            return
        if (ey, ea) < (by, ba):
            messagebox.showwarning("Aralık ters", "Bitiş dönemi başlangıçtan önce olamaz.")
            return
        donemler = self._donem_listesi(ba, by, ea, ey, periyot)
        if not donemler:
            messagebox.showwarning("Aralık boş", "Geçerli bir dönem aralığı gir.")
            return
        test_yol = self.test_yol_var.get().strip() if self.test_var.get() else ""
        yer = "TEST klasörü" if test_yol else "GERÇEK klasörler (iCloud)"
        durum_kodu, durum_etiket = self._durum_filtre()
        if periyot == "ceyrek":
            aralik_str = f"Q{donemler[0][0]}/{donemler[0][1]} → Q{donemler[-1][0]}/{donemler[-1][1]}"
        else:
            aralik_str = f"{donemler[0][0]:02d}/{donemler[0][1]} → {donemler[-1][0]:02d}/{donemler[-1][1]}"
        uyari = (f"{tur} için {len(donemler)} dönem işlenecek:\n"
                 f"{aralik_str}\n\n"
                 f"Her dönem: Listeyi Aç + CSV Oluştur + İNDİR (otomatik).\n"
                 f"Durum filtresi: {durum_etiket}\n"
                 f"Kayıt yeri: {yer}\n\n"
                 f"{'MUHSGK uzun sürer (her ay SGK taraması).' if tur == 'MUHSGK' else ''}\n"
                 f"Başlatılsın mı?")
        if not messagebox.askyesno("Toplu Çalıştır", uyari):
            return

        self._dur_iste = False
        self.calisiyor = True
        self._toplu_buton_durum("disabled")
        self.durdur_btn.configure(state="normal")
        self.yaz("\n" + "=" * 60 + "\n")
        self.yaz(f"[TOPLU] {tur} — {len(donemler)} dönem  ({donemler[0][0]:02d}/{donemler[0][1]} → {donemler[-1][0]:02d}/{donemler[-1][1]})  durum={durum_etiket}\n")
        self.yaz("=" * 60 + "\n")
        t = threading.Thread(target=self._toplu_thread, args=(tur, donemler, test_yol, durum_kodu), daemon=True)
        t.start()

    def _toplu_buton_durum(self, durum):
        for b in (self.indir_btn, self.csv_olustur_btn, self.liste_ac_btn, self.toplu_btn,
                  getattr(self, "kontrol_btn", None), getattr(self, "sablon_yenile_btn", None)):
            if b is not None:
                b.configure(state=durum)

    def _toplu_thread(self, tur, donemler, test_yol, durum_kodu="2"):
        # tur burada GIB kodudur (MUHSGK / KDV1 / POSET ...)
        snippet_dosya = BASE / ("snippet_muhtasar_sgk.js" if tur == "MUHSGK" else "snippet_kdv.js")
        try:
            snippet_js = snippet_dosya.read_text(encoding="utf-8")
        except Exception as e:
            self.kuyruk.put(f"[HATA] Snippet okunamadı: {e}\n")
            self.kuyruk.put(("__TOPLUBITTI__", None))
            return

        periyot = self._periyot(tur)
        basari, hatali, atlanan = 0, 0, 0
        for idx, (ay, yil) in enumerate(donemler, 1):
            if self._dur_iste:
                self.kuyruk.put("\n[Durduruldu] Toplu işlem durduruldu.\n")
                break
            if periyot == "ceyrek":
                donem_etiket = f"Q{ay} ({(ay-1)*3+1:02d}-{ay*3:02d}/{yil})"
            else:
                donem_etiket = f"{ay:02d}/{yil}"
            self.kuyruk.put("\n" + "#" * 60 + "\n")
            self.kuyruk.put(f"# [{idx}/{len(donemler)}]  DÖNEM {donem_etiket}  —  {tur}\n")
            self.kuyruk.put("#" * 60 + "\n")
            try:
                # 1) Listeyi Aç
                if periyot == "ceyrek":
                    bas, bit, _, _ = self._donem_bilgisi(tur, str(ay), str(yil))
                else:
                    bas, bit = self._yukleme_araligi(ay, yil)
                self.kuyruk.put(f"[Liste Aç] yükleme {bas:%d/%m/%Y} → {bit:%d/%m/%Y}\n")
                sonuc = self._cdp_calistir(self._liste_ac_snippet(tur, ay, yil, bas, bit, durum_kodu))
                self.kuyruk.put(f"[✓] {sonuc}\n")
                time.sleep(1.0)
                if self._dur_iste:
                    break

                # 2) CSV Oluştur
                self.kuyruk.put("[CSV Oluştur] Chrome'dan çekiliyor...\n")
                csv_metni = self._cdp_calistir(snippet_js)
                if not csv_metni or csv_metni.count("\n") < 1:
                    self.kuyruk.put("[i] Bu dönemde kayıt yok — atlandı.\n")
                    atlanan += 1
                    continue
                csvyol = BASE / f"uretilen_{tur.lower()}_{yil}{ay:02d}.csv"
                csvyol.write_text(csv_metni, encoding="utf-8-sig")
                self.kuyruk.put(f"[✓] CSV: {csvyol.name}  ({csv_metni.count(chr(10))} satır)\n")
                self._yeni_mukellef_ekle(csvyol)
                if self._dur_iste:
                    break

                # 3) İNDİR — dönem filtresi YOK: listede ne varsa hepsi iner,
                # her biri kendi yıl/ay/tür klasörüne (yıl klasörü karışmayı önler).
                cmd = [sys.executable, str(INDIR_PY), "--csvpath", str(csvyol), "--tur", tur]
                if test_yol:
                    cmd += ["--cikti", test_yol]
                if self.atla_var.get():
                    cmd += ["--atla-mevcut"]
                self.kuyruk.put("[İNDİR] başlıyor...\n")
                kod = self._indir_sync(cmd)
                self.kuyruk.put(f"[İNDİR bitti] çıkış kodu: {kod}\n")
                basari += 1
            except Exception as e:
                hatali += 1
                self.kuyruk.put(f"[HATA] {ay:02d}/{yil}: {e}\n")
                self.kuyruk.put("→ Sonraki döneme geçiliyor...\n")
                continue

        self.kuyruk.put("\n" + "=" * 60 + "\n")
        self.kuyruk.put(f"[TOPLU BİTTİ]  Başarılı: {basari}  |  Boş/atlanan: {atlanan}  |  Hatalı: {hatali}\n")
        self.kuyruk.put("=" * 60 + "\n")
        self.kuyruk.put(("__TOPLUBITTI__", None))

    def _indir_sync(self, cmd):
        """indir.py'yi senkron çalıştır, çıktıyı akıt, çıkış kodunu döndür (toplu için)."""
        env = dict(os.environ)
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUTF8"] = "1"
        env["PYTHONUNBUFFERED"] = "1"
        self.surec = subprocess.Popen(
            cmd, cwd=str(BASE), stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            stdin=subprocess.DEVNULL, env=env, bufsize=1, text=True,
            encoding="utf-8", errors="replace",
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
        for satir in self.surec.stdout:
            self.kuyruk.put(satir)
            if self._dur_iste:
                try:
                    self.surec.terminate()
                except Exception:
                    pass
                break
        self.surec.wait()
        return self.surec.returncode

    def _kuyruk_isle(self):
        try:
            while True:
                item = self.kuyruk.get_nowait()
                if isinstance(item, tuple) and item and item[0] == "__BITTI__":
                    kod = item[1]
                    self.calisiyor = False
                    self._toplu_buton_durum("normal")
                    self.durdur_btn.configure(state="disabled")
                    self.yaz(f"\n[Bitti] çıkış kodu: {kod}\n")
                elif isinstance(item, tuple) and item and item[0] == "__CSVBITTI__":
                    yol = item[1]
                    self.calisiyor = False
                    self._toplu_buton_durum("normal")
                    if yol:
                        self.csv_var.set(yol)   # üretilen CSV'yi indirme kutusuna yaz
                        self.yaz(f"[→] CSV indirme kutusuna yazıldı. Artık 'İNDİR'e basabilirsin.\n")
                elif isinstance(item, tuple) and item and item[0] == "__LISTEBITTI__":
                    self.calisiyor = False
                    self._toplu_buton_durum("normal")
                    self.durdur_btn.configure(state="disabled")
                elif isinstance(item, tuple) and item and item[0] == "__TOPLUBITTI__":
                    self.calisiyor = False
                    self._dur_iste = False
                    self._toplu_buton_durum("normal")
                    self.durdur_btn.configure(state="disabled")
                elif isinstance(item, tuple) and item and item[0] == "__KONTROLBITTI__":
                    self.calisiyor = False
                    self._toplu_buton_durum("normal")
                else:
                    self.yaz(item)
        except queue.Empty:
            pass
        self.after(120, self._kuyruk_isle)

    # ============== Z RAPORU → DEFTER BEYAN (KALDIRILDI — Müşavir Pro'ya taşındı) ==============
    def _zr_dosya_sec_KALDIRILDI(self):
        from tkinter import filedialog
        baslangic = str(Path.home() / "OneDrive" / "Desktop" / "🟡 Z_RAPORU")
        if not Path(baslangic).exists():
            baslangic = str(Path.home() / "Desktop")
        yol = filedialog.askopenfilename(
            initialdir=baslangic,
            title="Z Raporu Excel'i Seç",
            filetypes=[("Excel", "*.xlsx"), ("Tüm dosyalar", "*.*")]
        )
        if yol:
            self.zr_yol_var.set(yol)

    def _zr_drop(self, event):
        """tkinterdnd2 drop handler — dosyayı entry'ye koy."""
        try:
            data = event.data.strip('{}').strip().strip('"')
            if data and Path(data).exists():
                self.zr_yol_var.set(data)
                self.kuyruk.put(f"[Z RAPORU] Excel sürüklendi: {Path(data).name}\n")
        except Exception as e:
            self.kuyruk.put(f"[Z RAPORU] Sürükle hatası: {e}\n")

    def _zr_gonder(self):
        yol = self.zr_yol_var.get().strip().strip('"')
        if not yol or yol.startswith("(Excel"):
            self.kuyruk.put("[Z RAPORU] ! Önce Excel seç veya sürükle\n")
            return
        if not Path(yol).exists():
            self.kuyruk.put(f"[Z RAPORU] ! Dosya bulunamadı: {yol}\n")
            return
        self.zr_btn.configure(state="disabled", text="⏳ Yükleniyor...")
        threading.Thread(target=self._zr_thread, args=(yol,), daemon=True).start()

    def _zr_thread(self, yol):
        import json as _json
        try:
            from openpyxl import load_workbook
            self.kuyruk.put("=" * 60 + "\n")
            self.kuyruk.put(f"[Z RAPORU] 📂 Excel: {Path(yol).name}\n")

            # Chrome'da Defter Beyan sekmesini bul
            try:
                tabs = requests.get(f"http://localhost:9222/json", timeout=5).json()
            except Exception:
                self.kuyruk.put("[Z RAPORU] ❌ Chrome bağlantısı yok. Önce 'Chrome'u Başlat' bas ve defterbeyan.gov.tr'a giriş yap.\n")
                self.zr_btn.configure(state="normal", text="🚀 Defter Beyan'a Gönder")
                return
            db_tabs = [t for t in tabs if "defterbeyan.gov.tr" in t.get("url", "")]
            if not db_tabs:
                self.kuyruk.put("[Z RAPORU] ❌ Defter Beyan sekmesi açık değil. Chrome'da portal.defterbeyan.gov.tr'a giriş yap.\n")
                self.zr_btn.configure(state="normal", text="🚀 Defter Beyan'a Gönder")
                return
            ws = websocket.create_connection(db_tabs[0]["webSocketDebuggerUrl"], timeout=30, max_size=None)

            mid = [1000]
            def js(expr, tm=20):
                mid[0] += 1
                ws.settimeout(tm)
                ws.send(_json.dumps({
                    "id": mid[0], "method": "Runtime.evaluate",
                    "params": {"expression": expr, "returnByValue": True, "awaitPromise": True}
                }))
                while True:
                    try:
                        m = _json.loads(ws.recv())
                    except Exception:
                        return None
                    if m.get("id") == mid[0]:
                        return m.get("result", {}).get("result", {}).get("value")

            # Hangi mükellef seçili tespit et
            mevcut_url = js("location.href") or ""
            self.kuyruk.put(f"[Z RAPORU] 📍 Aktif sayfa: {mevcut_url[:80]}\n")

            # Excel'i oku
            wb = load_workbook(yol, data_only=True)
            firmalar_excelde = [s for s in wb.sheetnames if s != "📋 ANA SAYFA"]
            self.kuyruk.put(f"[Z RAPORU] 📊 Excel'de {len(firmalar_excelde)} firma sayfası bulundu\n")

            # Aktif sayfada hangi mükellef seçili — Defter Beyan üst banner'dan oku
            aktif_mukellef = js("""
            (function(){
              const banner = document.querySelector('.dbs-navbar__content span');
              if (banner) return banner.innerText.replace(/\\s+/g,' ').trim();
              return '';
            })()
            """) or ""
            self.kuyruk.put(f"[Z RAPORU] 👤 DB'de aktif mükellef: {aktif_mukellef[:80]}\n")

            # Aktif mükellefe en yakın firma sayfasını bul
            from indir import _norm
            secilen_sayfa = None
            for s in firmalar_excelde:
                if _norm(s) in _norm(aktif_mukellef) or _norm(aktif_mukellef[:30]) in _norm(s):
                    secilen_sayfa = s
                    break
            if not secilen_sayfa:
                self.kuyruk.put(f"[Z RAPORU] ❌ Excel'de '{aktif_mukellef}' için sayfa yok. Sayfalar: {firmalar_excelde}\n")
                self.zr_btn.configure(state="normal", text="🚀 Defter Beyan'a Gönder")
                ws.close()
                return

            self.kuyruk.put(f"[Z RAPORU] ✓ Eşleşen sayfa: {secilen_sayfa}\n")

            # Sayfadan Z raporlarını oku
            ws_excel = wb[secilen_sayfa]
            raporlar = []
            bas = 13
            for r in range(bas, bas + 31):
                try:
                    z_no = ws_excel.cell(r, 3).value
                    tutar = ws_excel.cell(r, 4).value
                    kdv = ws_excel.cell(r, 5).value
                    nakit = ws_excel.cell(r, 6).value or 0
                    kredi = ws_excel.cell(r, 7).value or 0
                    if z_no and tutar and float(tutar) > 0:
                        raporlar.append({
                            "z_no": int(z_no), "tutar": float(tutar),
                            "kdv_orani": int(kdv) if kdv else 1,
                            "nakit": float(nakit), "kredi": float(kredi),
                        })
                except Exception:
                    pass
            self.kuyruk.put(f"[Z RAPORU] 📋 {len(raporlar)} Z raporu okundu — Defter Beyan'a gönderiliyor...\n")

            ok = 0; fail = 0
            for i, r in enumerate(raporlar, 1):
                params = _json.dumps({
                    "z_no": str(r["z_no"]),
                    "tutar": str(r["tutar"]).replace(".", ","),
                    "nakit": str(r["nakit"]).replace(".", ","),
                    "kredi": str(r["kredi"]).replace(".", ","),
                    "aciklama": f"{r['z_no']} NL. Z RAPORU Mal Satışı",
                })
                snippet = f"""
                (async () => {{
                  const wait = ms => new Promise(res => setTimeout(res, ms));
                  function setReact(el, v) {{
                    if (!el) return false;
                    const s = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                    s.call(el, v);
                    el.dispatchEvent(new Event('input', {{bubbles: true}}));
                    el.dispatchEvent(new Event('change', {{bubbles: true}}));
                    return true;
                  }}
                  const p = {params};
                  try {{
                    setReact(document.querySelector('input[name="siraNo"]'), p.z_no);
                    await wait(400);
                    setReact(document.querySelector('div[name="tutarDiv"] input'), p.tutar);
                    await wait(400);
                    const ac = document.querySelector('input[name="aciklama"]');
                    if (ac) setReact(ac, p.aciklama);
                    await wait(300);
                    const kk = document.querySelector('input[placeholder="Kredi Kartı"]');
                    if (kk) setReact(kk, p.kredi);
                    await wait(300);
                    const nk = document.querySelector('input[placeholder="Nakit"]');
                    if (nk) setReact(nk, p.nakit);
                    await wait(500);
                    const btn = Array.from(document.querySelectorAll('button'))
                      .find(b => b.innerText && b.innerText.includes('Belgeyi Güncelle'));
                    if (btn) {{ btn.click(); await wait(1500); return 'ok'; }}
                    return 'ok_buton_yok';
                  }} catch(e) {{ return 'HATA: ' + e.message; }}
                }})()
                """
                sonuc = js(snippet, 30)
                if sonuc and "ok" in str(sonuc):
                    self.kuyruk.put(f"[Z RAPORU] [{i}/{len(raporlar)}] Z {r['z_no']} — {r['tutar']:.2f} TL ✅\n")
                    ok += 1
                else:
                    self.kuyruk.put(f"[Z RAPORU] [{i}/{len(raporlar)}] Z {r['z_no']} ❌ {sonuc}\n")
                    fail += 1
                time.sleep(2)

            self.kuyruk.put(f"\n[Z RAPORU] 📊 TAMAM: {ok} başarılı, {fail} hatalı / {len(raporlar)} toplam\n")
            self.kuyruk.put("=" * 60 + "\n")
            ws.close()
        except Exception as e:
            self.kuyruk.put(f"[Z RAPORU] ❌ HATA: {e}\n")
        finally:
            self.zr_btn.configure(state="normal", text="🚀 Defter Beyan'a Gönder")


def main():
    try:
        Panel().mainloop()
    except Exception:
        hata = traceback.format_exc()
        try:
            (BASE / "panel_hata.log").write_text(hata, encoding="utf-8")
        except Exception:
            pass
        try:
            import tkinter.messagebox as mb
            mb.showerror("Panel hatası", hata)
        except Exception:
            print(hata)


if __name__ == "__main__":
    main()
