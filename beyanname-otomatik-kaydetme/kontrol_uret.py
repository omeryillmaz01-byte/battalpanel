"""Kontrol sayfasını disk CSV'den doğrudan üret (panel'e bağımlı değil).
Kullanım: python kontrol_uret.py KDV1 5 2026
"""
import sys, csv, re
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))
from indir import _norm, _tr, _kw, _sim, mukellef_listesi_yukle
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent

def eslesir(muk, ad):
    mn=_norm(muk); hn=_norm(ad)
    if not mn or not hn: return False
    if mn in hn or hn in mn: return True
    hw=set(re.findall(r'[A-Z0-9]+',_tr(ad))); kws=_kw(muk)
    if len(kws)>=2:
        e=sum(1 for kw in kws[:2] if kw in hw or any(len(kw)>3 and _sim(kw,h)>=0.82 for h in hw))
        return e>=2
    return False

def hukuki_ek(u):
    u=u.upper()
    if 'ANONİM' in u or 'ANONIM' in u or 'A.Ş' in u or u.endswith(' AŞ'): return 'A.Ş.'
    if 'LİMİTED' in u or 'LIMITED' in u or 'LTD' in u: return 'LTD.ŞTİ.'
    if 'ADİ ORT' in u or 'ADI ORT' in u or 'ORTAKLIĞI' in u or 'ORTAKLIGI' in u: return 'ADİ ORT.'
    return ''

def main():
    tur = sys.argv[1].upper()
    ay = int(sys.argv[2]); yil = int(sys.argv[3])
    is_muhsgk = tur in ("MUHSGK","GGECICI","KGECICI")

    taner, omer = mukellef_listesi_yukle(BASE)

    # Şablondan beklenen (YOK hariç)
    df = pd.read_excel('MUKELLEF VERGI ŞABLONU.xlsx', sheet_name='Mükellef Vergi Şablonu')
    kol = {str(c).strip().upper():c for c in df.columns}
    fk = kol.get('FİRMA ÜNVANI')
    tk = kol.get('GEKAP') if tur=='POSET' else kol.get(tur)
    def ek_temizle(ad):
        a=_norm(ad)
        for ek in ("LTDSTI","LTD","ANONIMSIRKETI","AS","ADIORT","ADIORTAKLIGI","STI","SIRKETI","ANONIM"):
            if a.endswith(ek): a=a[:-len(ek)]
        return a

    # Tam ünvanları al (şablonda artık LTD.ŞTİ./A.Ş. ekli)
    tamunvan_exact = {}   # tam normalize -> ünvan (öncelikli, çakışmayı önler)
    tamunvan = {}         # ek-temizlenmiş -> ünvan (yedek eşleşme)
    beklenen = set()
    for _,r in df.iterrows():
        f=str(r.get(fk,'')).strip()
        if not f or f=='NAN': continue
        tamunvan_exact[_norm(f)] = f
        tamunvan.setdefault(ek_temizle(f), f)  # ilk geleni koru (çakışmada üzerine yazma)
        if tk is None: continue
        v=str(r.get(tk,'')).strip().upper()
        if not v or v=='NAN' or 'YOK' in v: continue
        beklenen.add(ek_temizle(f))

    taner = {f for f in taner if ek_temizle(f) in beklenen}
    omer  = {f for f in omer  if ek_temizle(f) in beklenen}

    # Disk CSV(ler)
    rows = []
    cf = BASE / f"uretilen_{tur.lower()}_{yil}{ay:02d}.csv"
    if cf.exists():
        with open(cf,encoding='utf-8-sig') as f:
            rows = list(csv.DictReader(f,delimiter=';'))

    def durum_grup(r):
        d=str(r.get('Durum','')).upper()
        if 'İPTAL' in d or 'IPTAL' in d: return 'iptal'
        if 'HATA' in d: return 'hatali'
        return 'verildi'

    donem_tam = f"{ay:02d}/{yil}-{ay:02d}/{yil}"
    yil_str = str(yil)
    aylik_ver=[]; ceyrek_ver=[]; hatali=[]
    for r in rows:
        vd=str(r.get('Vergilendirme_Donemi','') or '')
        # Sadece kontrol edilen döneme ait kayıtlar (aylık tam dönem veya bu yıl 3 aylık)
        bu_donem = (donem_tam in vd) or (is_muhsgk and '-' in vd and yil_str in vd)
        if not bu_donem: continue
        g=durum_grup(r)
        if g=='iptal': continue
        if g=='hatali':
            hatali.append(r); continue
        if donem_tam in vd:
            aylik_ver.append(r)
        elif is_muhsgk and '-' in vd and yil_str in vd:
            ceyrek_ver.append(r)

    aylik_ad=[str(r.get('Ad_Soyad','')).strip() for r in aylik_ver]
    ceyrek_ad=[str(r.get('Ad_Soyad','')).strip() for r in ceyrek_ver]
    tum_gib={str(r.get('Ad_Soyad','')).strip() for r in (aylik_ver+ceyrek_ver+hatali) if r.get('Ad_Soyad')}

    def veren(muk,liste): return any(eslesir(muk,a) for a in liste)
    def yukleme(muk):
        best=''
        for r in (aylik_ver+ceyrek_ver):
            a=str(r.get('Ad_Soyad','')).strip(); z=str(r.get('Yukleme_Zamani','')).strip()
            if a and z and eslesir(muk,a) and z>best: best=z
        return best
    def hata(muk):
        for r in hatali:
            a=str(r.get('Ad_Soyad','')).strip()
            if a and eslesir(muk,a):
                return f"⚠ HATALI {str(r.get('Yukleme_Zamani','')).strip()}".strip()
        return ''
    def unvan(muk):
        # Önce TAM eşleşme (GİZEM GÖKER vs GİZEM GÖKER ADİ ORT. çakışmasını önler)
        tn=tamunvan_exact.get(_norm(muk))
        if tn: return tn
        tn=tamunvan.get(ek_temizle(muk))
        if tn: return tn
        return muk

    # Excel
    sp = BASE/'MUKELLEF VERGI ŞABLONU.xlsx'
    wb = load_workbook(sp)
    sayfa = f"{tur} {ay:02d}-{yil}"
    if sayfa in wb.sheetnames: del wb[sayfa]
    ws = wb.create_sheet(sayfa)

    bfill=PatternFill('solid',fgColor='1F2937'); bfont=Font(bold=True,color='FFFFFF',size=11)
    yfill=PatternFill('solid',fgColor='C6EFCE'); yfont=Font(color='006100',bold=True)
    mfill=PatternFill('solid',fgColor='DBEAFE'); mfont=Font(color='0050C0',bold=True)
    kfill=PatternFill('solid',fgColor='FFC7CE'); kfont=Font(color='9C0006',bold=True)
    ofill=PatternFill('solid',fgColor='FFE4B5'); ofont=Font(color='B45309',bold=True,size=10)
    mrk=Alignment(horizontal='center',vertical='center'); sol=Alignment(horizontal='left',vertical='center')
    kn=Border(*[Side(style='thin',color='999999')]*4)

    son = 6 if is_muhsgk else 5
    ws.merge_cells(f"A1:{get_column_letter(son)}1")
    ws['A1']=f"{tur} KONTROL LİSTESİ — {ay:02d}/{yil}"
    ws['A1'].font=Font(bold=True,size=14,color='FFFFFF'); ws['A1'].fill=bfill; ws['A1'].alignment=mrk
    ws.row_dimensions[1].height=28

    bl=['Sahip','Firma Ünvanı','Durum']
    if is_muhsgk: bl.append('Dönem')
    bl+=['Yükleme Zamanı','Hata Durumu']
    for c,h in enumerate(bl,1):
        cell=ws.cell(2,c,h); cell.fill=bfill; cell.font=bfont; cell.alignment=mrk; cell.border=kn
    ws.row_dimensions[2].height=22

    # Önce tüm satırları topla (sıralama için)
    vs=0; ms=0; xs=0
    kayitlar=[]
    for sk,ms_set in [('TANER',taner),('ÖMER',omer)]:
        for firma in sorted(ms_set):
            va=veren(firma,aylik_ad); v3=is_muhsgk and veren(firma,ceyrek_ad)
            s3=tur in ('GGECICI','KGECICI')
            if va and not s3:
                durum,donem='✅ Verildi','AYLIK'; fl,ft=yfill,yfont; vs+=1; sira=0
            elif v3 or (s3 and va):
                durum,donem='✅ Verildi','3 AYLIK'; fl,ft=mfill,mfont; xs+=1; sira=0
            else:
                durum,donem='❌ Verilmedi',('3 AYLIK' if is_muhsgk else ''); fl,ft=kfill,kfont; ms+=1; sira=1
            kayitlar.append({'sira':sira,'sk':sk,'fg':unvan(firma),'durum':durum,'donem':donem,
                             'fl':fl,'ft':ft,'z':(yukleme(firma) if (va or v3) else ''),'ht':hata(firma)})

    # AKILLI SIRALAMA: Verildi (yeşil) üstte, Verilmedi (kırmızı) altta; içinde sahip+ünvan
    kayitlar.sort(key=lambda k:(k['sira'], k['sk'], k['fg']))

    row=3
    for k in kayitlar:
        a=ws.cell(row,1,k['sk']); b=ws.cell(row,2,k['fg']); c=ws.cell(row,3,k['durum'])
        a.alignment=mrk; b.alignment=sol; c.alignment=mrk
        a.border=b.border=c.border=kn; a.fill=b.fill=c.fill=k['fl']; c.font=k['ft']
        if is_muhsgk:
            d=ws.cell(row,4,k['donem']); d.alignment=mrk; d.border=kn; d.fill=k['fl']; d.font=k['ft']; yc=5
        else: yc=4
        zc=ws.cell(row,yc,k['z']); zc.alignment=mrk; zc.border=kn; zc.fill=k['fl']; zc.font=Font(color='333333',size=10)
        hc=ws.cell(row,yc+1,k['ht']); hc.alignment=mrk; hc.border=kn
        if k['ht']: hc.fill=ofill; hc.font=ofont
        else: hc.fill=k['fl']; hc.font=Font(color='333333',size=10)
        row+=1

    son_veri_satir = row - 1
    ws.cell(row,1,'TOPLAM').font=Font(bold=True)
    oz=f"Verdi: {vs}"
    if is_muhsgk: oz+=f"  |  3 Aylık: {xs}"
    oz+=f"  |  Vermedi: {ms}  |  Toplam: {vs+xs+ms}"
    ws.cell(row,2,oz).font=Font(bold=True)

    ws.column_dimensions['A'].width=12; ws.column_dimensions['B'].width=58
    ws.column_dimensions['C'].width=16
    if is_muhsgk:
        ws.column_dimensions['D'].width=12; ws.column_dimensions['E'].width=22; ws.column_dimensions['F'].width=26
    else:
        ws.column_dimensions['D'].width=22; ws.column_dimensions['E'].width=26
    ws.freeze_panes='A3'
    # AKILLI FİLTRE: başlık satırına otomatik filtre (renge/değere göre filtrele-sırala aktif)
    ws.auto_filter.ref = f"A2:{get_column_letter(son)}{son_veri_satir}"
    wb.active=wb.sheetnames.index(sayfa)
    wb.save(sp)
    print(f"[OK] '{sayfa}' yazildi: Verdi {vs}, 3 Aylik {xs}, Vermedi {ms}, Hatali {len(hatali)} | siralama+filtre aktif")

if __name__=='__main__':
    main()
